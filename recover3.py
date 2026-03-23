# -*- coding: utf-8 -*-
"""
Created on Thu Feb 26 10:03:03 2026

@author: rjguz
"""

import streamlit as st
import pandas as pd
import datetime
import uuid
import firebase_admin
from firebase_admin import credentials, firestore
import altair as alt
from io import BytesIO
from fpdf import FPDF
from PIL import Image
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
import extra_streamlit_components as stx

# ==== CONFIGURACIÓN INICIAL ====
st.set_page_config(
    page_title="Bitácora de Mantenimiento",
    page_icon="imagenes/logoo.png",
    layout="wide"
)

# ==== COOKIE MANAGER (SESIÓN INDEPENDIENTE POR DISPOSITIVO) ====
if "cookie_manager" not in st.session_state:
    st.session_state.cookie_manager = stx.CookieManager()

cookie_manager = st.session_state.cookie_manager

# ==== INICIALIZAR FIRESTORE ====
if "db" not in st.session_state:
    if not firebase_admin._apps:
        firebase_config = dict(st.secrets["firebase"])
        cred = credentials.Certificate(firebase_config)
        firebase_admin.initialize_app(cred)
    st.session_state.db = firestore.client()

# ==== CSS GLOBAL (DARK MODE INDUSTRIAL) ====
def aplicar_estilos_globales():
    st.markdown("""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;600;700&display=swap');

        /* Fondo principal oscuro */
        .main, .block-container {
            background-color: #0E1117;
            color: white;
            font-family: 'Montserrat', sans-serif;
        }

        /* Contenedores de registros */
        .registro-container {
            background-color: #1E1E1E;
            padding: 20px;
            border-radius: 10px;
            border: 1px solid #444;
            margin-bottom: 15px;
            color: white;
        }

        /* Header de sesión */
        .session-header {
            background: linear-gradient(135deg, #001a33 0%, #003366 100%);
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border: 1px solid #00d4ff;
        }

        .session-info {
            color: white;
        }

        .session-info h3 {
            color: #00d4ff;
            margin: 0;
            font-size: 1.2rem;
        }

        /* Selector dual elegante */
        .dual-selector {
            display: flex;
            background-color: #1E1E1E;
            border-radius: 10px;
            padding: 5px;
            margin: 20px 0;
            border: 2px solid #00d4ff;
        }

        .dual-selector button {
            flex: 1;
            padding: 12px 24px;
            border: none;
            background: transparent;
            color: white;
            font-weight: 600;
            cursor: pointer;
            border-radius: 8px;
            transition: all 0.3s;
        }

        .dual-selector button.active {
            background: linear-gradient(135deg, #00d4ff 0%, #0099cc 100%);
            color: #0E1117;
        }

        .dual-selector button:hover {
            background-color: #2a2a2a;
        }

        /* Títulos */
        h1, h2, h3 {
            color: #00d4ff;
            font-family: 'Montserrat', sans-serif;
        }

        /* Botones */
        .stButton>button {
            background: linear-gradient(135deg, #00d4ff 0%, #0099cc 100%);
            color: white;
            border: none;
            border-radius: 8px;
            padding: 10px 20px;
            font-weight: 600;
            transition: all 0.3s;
        }

        .stButton>button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(0, 212, 255, 0.4);
        }

        /* Inputs */
        .stTextInput>div>div>input,
        .stSelectbox>div>div>select,
        .stTextArea>div>div>textarea {
            background-color: #1E1E1E;
            color: white;
            border: 1px solid #444;
            border-radius: 5px;
        }

        /* Tabs */
        .stTabs [data-baseweb="tab-list"] {
            background-color: #1E1E1E;
            border-radius: 10px;
            padding: 5px;
        }

        .stTabs [data-baseweb="tab"] {
            color: white;
            font-weight: 600;
        }

        .stTabs [aria-selected="true"] {
            background-color: #00d4ff;
            color: #0E1117;
            border-radius: 8px;
        }

        /* Expanders */
        .streamlit-expanderHeader {
            background-color: #1E1E1E;
            color: white;
            border: 1px solid #444;
            border-radius: 8px;
        }

        /* Alertas */
        .stSuccess {
            background-color: #1a4d2e;
            color: #4ade80;
            border-left: 4px solid #4ade80;
        }

        .stError {
            background-color: #4d1a1a;
            color: #f87171;
            border-left: 4px solid #f87171;
        }

        .stInfo {
            background-color: #1a3a4d;
            color: #60a5fa;
            border-left: 4px solid #60a5fa;
        }

        .stWarning {
            background-color: #4d3a1a;
            color: #fbbf24;
            border-left: 4px solid #fbbf24;
        }
        </style>
    """, unsafe_allow_html=True)

# ==== FUNCIONES DE VALIDACIÓN ====
def validar_usuario(mx, password):
    if not mx or not password:
        return None
    mx_id = mx.upper().strip()
    doc_ref = st.session_state.db.collection("empleados").document(mx_id)
    doc = doc_ref.get()
    if doc.exists:
        data = doc.to_dict()
        if str(data.get("password")) == str(password):
            return data
    return None

# ==== FUNCIONES FIRESTORE - TÉCNICOS ====
@st.cache_data(ttl=600)
def cargar_datos_tecnicos_cached(usuario, role, area):
    # Eliminamos normalizar_actividades() de aquí para evitar bucles de carga
    db = st.session_state.db
    if role == "admin":
        docs = db.collection("actividades").where("role", "==", "tecnico").stream()
    elif role != "tecnico":
        docs = db.collection("actividades").where("role", "==", "tecnico").where("area", "==", area).stream()
    else:
        docs = db.collection("actividades").where("Usuario", "==", usuario).stream()
    
    data = [doc.to_dict() for doc in docs]
    return pd.DataFrame(data) if data else pd.DataFrame(columns=["ID", "Usuario", "Nombre", "Actividad", "Tipo", "Número de orden", "Turno", "Supervisor", "area", "Línea", "Máquina", "Inicio", "Fin", "Duración (min)"])

@st.cache_data(ttl=300) # Caché de 5 min para la lista de empleados
def cargar_empleados_cached():
    docs = st.session_state.db.collection("empleados").stream()
    return [doc.to_dict() for doc in docs]

def cargar_datos_tecnicos_base(usuario, role, area):
    normalizar_actividades()
    if role == "admin":
        docs = (
            st.session_state.db.collection("actividades")
            .where("role", "==", "tecnico")
            .stream()
        )
    elif role != "tecnico":
        docs = (
            st.session_state.db.collection("actividades")
            .where("role", "==", "tecnico")
            .where("area", "==", area)
            .stream()
        )
    else:
        docs = (
            st.session_state.db.collection("actividades")
            .where("Usuario", "==", usuario)
            .stream()
        )
    data = [doc.to_dict() for doc in docs]
    return pd.DataFrame(data) if data else pd.DataFrame(columns=[
        "ID", "Usuario", "Nombre", "Actividad", "Tipo", "Número de orden", "Turno", 
        "Supervisor", "area", "Línea", "Máquina", "Inicio", "Fin", "Duración (min)"
    ])

def cargar_datos_tecnicos():
    if st.session_state.get("usar_cache", True):
        return cargar_datos_tecnicos_cached(
            st.session_state["usuario"],
            st.session_state.get("role", "tecnico"),
            st.session_state.get("area", "General")
        )
    else:
        return cargar_datos_tecnicos_base(
            st.session_state["usuario"],
            st.session_state.get("role", "tecnico"),
            st.session_state.get("area", "General")
        )

def invalidar_cache_tecnicos():
    cargar_datos_tecnicos_cached.clear()

def normalizar_actividades():
    if not st.session_state.get("actividades_normalizadas", False):
        db = st.session_state.db
        actividades = db.collection("actividades").stream()
        for doc in actividades:
            data = doc.to_dict()
            actualizado = False
            for campo in ["Turno", "Supervisor", "Número de orden"]:
                if campo not in data:
                    data[campo] = ""
                    actualizado = True
            if actualizado:
                db.collection("actividades").document(doc.id).set(data)
        st.session_state["actividades_normalizadas"] = True

def guardar_registro_tecnico(registro):
    registro["role"] = st.session_state.get("role", "tecnico")
    registro["area"] = st.session_state.get("area", "General")
    registro["tipo_rol"] = "tecnico"
    st.session_state.db.collection("actividades").document(registro["ID"]).set(registro)
    invalidar_cache_tecnicos()

def eliminar_registro_tecnico(registro_id):
    st.session_state.db.collection("actividades").document(registro_id).delete()
    invalidar_cache_tecnicos()

# ==== FUNCIONES FIRESTORE - SUPERVISORES ====
@st.cache_data(ttl=600)
def cargar_datos_supervisores_cached(usuario, role, area):
    if role == "supervisor":
        docs = (
            st.session_state.db.collection("registros_supervisores")
            .where("Usuario", "==", usuario)
            .stream()
        )
    elif role == "admin":  # ← AGREGAR ESTO
            docs = (
                st.session_state.db.collection("registros_supervisores")
                .stream()  # Admin ve TODOS los registros sin filtro de área
            )
    else:
        docs = (
            st.session_state.db.collection("registros_supervisores")
            .where("area", "==", area)
            .stream()
        )
    data = [doc.to_dict() for doc in docs]
    return pd.DataFrame(data) if data else pd.DataFrame(columns=[
        "ID", "Usuario", "Nombre", "area", "Supervisor", "Fecha", "Turno", "Jornada",
        "Jornada_Normal", "Cursos_Otros", "Inasistencias", "Total_Esperado", 
        "Total_Real", "Horas_Disponibles", "Eventos_Programados"
    ])

def cargar_datos_supervisores_base(usuario, role, area):
    if role == "supervisor":
        docs = (
            st.session_state.db.collection("registros_supervisores")
            .where("Usuario", "==", usuario)
            .stream()
        )
        
    elif role == "admin":  # ← AGREGAR ESTO
        docs = (
            st.session_state.db.collection("registros_supervisores")
            .stream()
            )
    else:
        docs = (
            st.session_state.db.collection("registros_supervisores")
            .where("area", "==", area)
            .stream()
        )
    data = [doc.to_dict() for doc in docs]
    return pd.DataFrame(data) if data else pd.DataFrame(columns=[
        "ID", "Usuario", "Nombre", "area", "Supervisor", "Fecha", "Turno", "Jornada",
        "Jornada_Normal", "Cursos_Otros", "Inasistencias", "Total_Esperado", 
        "Total_Real", "Horas_Disponibles", "Eventos_Programados"
    ])

def cargar_datos_supervisores():
    if st.session_state.get("usar_cache", True):
        return cargar_datos_supervisores_cached(
            st.session_state["usuario"],
            st.session_state.get("role", "supervisor"),
            st.session_state.get("area", "General")
        )
    else:
        return cargar_datos_supervisores_base(
            st.session_state["usuario"],
            st.session_state.get("role", "supervisor"),
            st.session_state.get("area", "General")
        )

def invalidar_cache_supervisores():
    cargar_datos_supervisores_cached.clear()

def guardar_registro_supervisor(registro):
    registro["role"] = "supervisor"
    registro["area"] = st.session_state.get("area", "General")
    registro["tipo_rol"] = "supervisor"
    st.session_state.db.collection("registros_supervisores").document(registro["ID"]).set(registro)
    invalidar_cache_supervisores()

def eliminar_registro_supervisor(registro_id):
    st.session_state.db.collection("registros_supervisores").document(registro_id).delete()
    invalidar_cache_supervisores()

# ==== AUTOLOGIN CON COOKIES ====
if "usuario" not in st.session_state:
    cookies = cookie_manager.get_all()
    if cookies and "session_user" in cookies:
        mx_id = cookies["session_user"].upper().strip()
        doc_ref = st.session_state.db.collection("empleados").document(mx_id)
        doc = doc_ref.get()
        if doc.exists:
            data = doc.to_dict()
            if data.get("mantener_sesion", False) == True:
                st.session_state["usuario"] = mx_id
                st.session_state["role"] = data.get("role", "tecnico")
                st.session_state["area"] = data.get("area", "General")
                st.session_state["nombre"] = data.get("nombre", "")
                st.session_state["unidad"] = data.get("unidad", "")
                st.session_state["business_unit"] = data.get("business_unit", "")
                st.session_state["emp_no"] = data.get("emp_no", "")
                st.session_state["last_login"] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                st.session_state.db.collection("empleados").document(mx_id).update({
                    "last_login": st.session_state["last_login"]
                })
                st.rerun()
            else:
                try:
                    cookie_manager.delete("session_user")
                except Exception:
                    pass

# ==== PANTALLA DE LOGIN ====
if "usuario" not in st.session_state:
    st.markdown("""
        <style>
        html, body, .main, .block-container {
            height: 100vh !important;
            overflow: hidden !important;
            margin: 0 !important;
            padding: 0 !important;
        }
        .left-panel {
            background: linear-gradient(135deg, #000b1a 0%, #001a33 100%);
            height: 100vh;
            padding: 10% 5%;
            color: white;
            font-family: 'Montserrat', sans-serif;
            position: relative;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }
        .left-panel h1 {
            font-size: 3.8rem;
            font-weight: 700;
            line-height: 1;
            margin-bottom: 20px;
        }
        .left-panel h1 span {
            color: #00d4ff;
        }
        .left-panel p {
            font-size: 1.1rem;
            color: rgba(255,255,255,0.6);
            max-width: 450px;
            border-left: 3px solid #0044cc;
            padding-left: 20px;
        }
        .login-title {
            text-align: center;
            margin-top: 40px;
            margin-bottom: 30px;
        }
        .login-title h2 {
            font-family: 'Montserrat', sans-serif;
            font-size: 1.6rem;
            color: #00d4ff;
            text-transform: uppercase;
            letter-spacing: 2px;
            margin: 0;
        }

        @media (max-width: 768px) {
            .stColumns {
                display: block !important;
            }
            .left-panel {
                display: none !important;
            }
        }
        </style>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1.2, 1])

    with col1:
        st.markdown("""
            <div class="left-panel">
                <h1>Purem by Eberspächer<br><span>Mantenimiento</span></h1>
                <p>Plataforma de Gestión y Mantenimiento Industrial para la Planta Ramos Arizpe.</p>
                <p>Developed by: Juan Rodrigo Guzmán Martínez</p>
            </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
            <div class="login-title">
                <h2>Iniciar Sesión</h2>
            </div>
        """, unsafe_allow_html=True)

        username = st.text_input("ID de Usuario (MX)", placeholder="ej. MX351544", key="login_usuario")
        password = st.text_input("Contraseña (número)", type="password", placeholder="••••", key="login_password")
        recordar = st.checkbox("Recordarme en este dispositivo", key="login_recordar")
        login_button = st.button("Acceder al Sistema", use_container_width=True, key="login_boton")

        if login_button:
            if not username or not password:
                st.warning("⚠️ Por favor coloque sus datos para poder ingresar ⚠️")
            else:
                user_data = validar_usuario(username, password)
                if user_data:
                    mx_id = username.upper().strip()
                    st.session_state["usuario"] = mx_id
                    st.session_state["role"] = user_data.get("role", "tecnico")
                    st.session_state["area"] = user_data.get("area", "General")
                    st.session_state["nombre"] = user_data.get("nombre", "")
                    st.session_state["unidad"] = user_data.get("unidad", "")
                    st.session_state["business_unit"] = user_data.get("business_unit", "")
                    st.session_state["emp_no"] = user_data.get("emp_no", "")
                    st.session_state["last_login"] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")

                    st.session_state.db.collection("empleados").document(mx_id).update({
                    "last_login": st.session_state["last_login"],
                    "mantener_sesion": recordar
                    })

                    if recordar:
                        try:
                            cookie_manager.set(
                                "session_user", 
                                mx_id, 
                                expires_at=datetime.datetime.now() + datetime.timedelta(days=30)
                            )
                        except Exception as e:
                            st.warning(f"No se pudo guardar la sesión en cookies: {e}")

                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos ❌")

        st.markdown("""
            <p style="font-size: 0.9rem; color: rgba(255,255,255,0.4); text-align: center; margin-top: 60px;">
                SISTEMA PRIVADO DE USO EXCLUSIVO.<br>
                CUALQUIER INTENTO DE ACCESO NO AUTORIZADO SERÁ MONITOREADO.
            </p>
            <p style="font-size: 1.0rem; color: rgba(255,255,255,0.3); text-align: center; margin-top: 20px;">
                PUREM BY EBERSPÄCHER - RAMOS ARIZPE<br>
                DEVELOPED BY: JUAN RODRIGO GUZMÁN MARTÍNEZ
            </p>
        """, unsafe_allow_html=True)

# ==== INTERFAZ PRINCIPAL ====
else:
    aplicar_estilos_globales()

    with st.sidebar:
        st.title("👤 Sesión")
        st.write(f"**{st.session_state['nombre']}**")
        st.write(f"{st.session_state['usuario']} - {st.session_state['role']}")
        st.write(f"Área: {st.session_state['area']}")
        if st.button("🚪 Cerrar sesión", use_container_width=True):
            try:
                st.session_state.db.collection("empleados").document(st.session_state["usuario"]).update({
                    "mantener_sesion": False
                })
            except Exception:
                pass

            try:
                cookie_manager.delete("session_user")
            except Exception:
                pass

            st.session_state.clear()
            st.rerun()
            
    ##########
    st.title(f"Bitácora de Mantenimiento - Usuario: {st.session_state['usuario']} | {st.session_state['nombre']}")
    
    if st.session_state["role"] == "tecnico":
        tab1, tab2, tab3, tab4 = st.tabs(["📝 Registro", "📋 Historial", "📊 Rendimiento", "📂 Descargas"])
    
        with tab1:
            st.subheader("Registrar Actividad")
    
            tecnico_nombre = st.session_state.get("nombre", "")
            # Técnico (fuera del formulario para que no se duplique)
            st.text_input("Técnico:", value=tecnico_nombre, disabled=True, key="tecnico_display")
    
            # Mostrar tipo de mantenimiento fuera del formulario para permitir rerun inmediato
            tipo_mantenimiento = st.selectbox(
                "Tipo de Mantenimiento:",
                ["seleccionar", "Autonomo", "Preventivo (maquina operando)", "Correctivo (DOWNTIME)", "Mejora", "5s"],
                key="tipo_mantenimiento"
            )
    
            # Unidad / Línea / Máquina se muestran también fuera del form para actualización inmediata
            unidad = ""
            linea = ""
            maquina = ""
    
            if tipo_mantenimiento != "seleccionar":
                unidad = st.selectbox("Unidad:", ["EPU2-GATS", "EPU3-FORD", "EPU4", "FACILIDADES"], key="unidad_select")
    
                datos_cascada = {
                    "EPU2-GATS": {
                        "GATS L1": ["NC1507-OP.10", "NC1519-SA10A", "NC1810-SA10B", "NC1508-OP.20", "NC1505-SA20A", "NC1506-SA20B", "NC1503-SA20C", "SA25", "NC1509-OP.30", "NC1510-OP.40", "NC1808-OP.45", "NC1511-OP.50", "NC1512-OP.60", "NC1513-OP.70/80", "NC1514-OP.90", "NC1515-OP.92", "NC1516-OP.95", "NC1517-OP.100", "NC1518-OP.110", "DRESS OUT-OP.120", "ROBOT PERCEPTRON", "REWORK OP.20", "REWORK OP.45", "REWORK OP.50", "REWORK OP.60", "REWORK OP.70/80", "REWORK OP.95", "REWORK OP.110", "LEAK TEST OP.45", "LEAK TEST OP.50", "LEAK TEST OP.60", "LEAK TEST OP.70 / 80", "LEAK TEST OP.95", "LEAK TEST OP.110", "TRACK 1", "TRACK 2", "TRACK 3", "TRACK 4", "TRACK 5", "TRACK 6", "CONVEYOR", "POLIPASTO OP.50", "OP.5"],
                        "GATS L2": ["NC1679-OP.10", "NC1678-OP.20", "NC1681-SA20A", "NC1682-SA20B", "SA25", "NC1680-OP.30", "NC1685-OP.40", "NC1683-OP.45", "NC1689-OP.50", "NC1684-OP.60", "NC1690-OP.70/80", "NC1686-OP.90", "NC1687-OP.92", "NC1688-OP.95", "NC1691-OP.100", "NC1692-OP.110", "DRESS OUT-OP.120", "REWORK OP.20B - 2", "REWORK OP.45", "REWORK OP.50", "REWORK OP.60", "REWORK OP.70/80", "REWORK OP.95", "REWORK OP.110", "LEAK TEST OP.20", "LEAK TEST OP.45", "LEAK TEST OP.50", "LEAK TEST OP.60", "LEAK TEST OP.70 / 80", "LEAK TEST OP.95", "LEAK TEST OP.110", "TRACK 1", "TRACK 2", "TRACK 3", "TRACK 4", "TRACK 5", "TRACK 6", "CONVEYOR", "POLIPASTO OP.50", "ROBOT PERCEPTRON", "OP.5"],
                        "CANNING": ["CV5", "PROTO 1", "HAAS SPINER 1", "HAAS SPINER 2"],
                        "SOLDADURA Y ROLADO": ["SHELL LINE WEIL LASER #1"],
                        "CORTE LASER": ["LASER CUT #1"],
                        "SUB ASSY": ["NC1500 SA60D", "SA50ABC", "NC1502 SA60A-C", "NC1808 SA70C", "NC1811 SA90 SERVICE PANEL", "NC607 SA50E/F", "GLUING Estación"]
                    },
                    "EPU3-FORD": {
                        "TRIFLOW": ["NC1769 OP.20A", "NC1769 OP.20B", "NC1846 OP.10A", "NC1846 OP.10B", "NC1819 OP.50A", "NC1819 OP.50B", "NC1818 OP30A", "NC1818 OP30B", "OP.40.1 PRENSA VERICAL", "OP. 40.2 PRENSA VERICAL", "OP.60 SIZER  SCR", "OP.70 LEAK TESTER #3", "OP.75 RW", "OP.80 CHEKING FIXTURE"],
                        "P708": ["NC1766 OP.10A", "NC1766 OP.10B", "NC1817 OP.20A", "NC1817 OP.20B", "NC1770 OP.40A/40B", "OP.50 FINAL LEAK  #4", "OP.60 INSPECCIÓN CHECKING FIXTURE", "OP.80 RETRABAJO", "OP.45 NC1840", "NC1816-OP.10 SCR", "OP.20 FINAL LEAK #2 SCR", "OP.25 SCR RW", "OP.30 INSPECCIÓN CHEKING FIXTURE SCR"],
                        "H567": ["NC1188 OP.40", "NC1166  OP.30", "NC1757 OP.20", "NC 604 OP.10", "PRUEBA DE FUBAS OP.50", "CKECKING FIXTURE OP.60", "PRUEBA DE FUBAS OP.80", "CHECKING FIXTURE OP-90", "RETRABAJO"],
                        "MPC REAR": ["NC1804  OP.10", "OP-15 RB 80 ER-V", "NC1805  OP. 20", "NC1481 OP.30", "OP.40 W-100 LT", "OP.50 UNDERFLOOR", "OP.25 STATIONARY WELDING MACHINE", "RW OP.45", "UPPER ASSY CRIMP", "LOWER ASSY CRIMP", "OP-30 MELTON 20766", "W-80/90 LASER ETCH", "LASER ETCH #2", "NC1307 OP. 30.2", "NC1304 OP. 20.2", "OP.40.2 LEAK TEST", "OP.45 BAND-IT"],
                        "MPC FRONT": ["NC1571 OP.10", "NC1571 OP.40", "NC1573  OP.50", "NC1573  OP.60", "NC1572 OP.20", "OP.70 LEAK TESTER", "RW OP-75", "OP.80 FINAL ASSY CRIMP", "OP.90 FINAL CRIMPING", "OP.100 NUT RUNNER", "NC1847", "OP.70.2 LEAK TEST"],
                        "RESONADOR": ["OP.10 LOCK SEAMER", "OP.20 SIZER", "OP.30 SPINNER", "OP.40 SPOT WELDER", "OP.50 LEAK TESTER"],
                        "CANNING": ["L1 CANNING- KIRSCHENHOFER #3", "L2 CANNING- KIRSCHENHOFER #4", "KIRSCHENHOFER #2"],
                        "MPC CANNING L1": ["20-A LMM", "20-B PESAJE", "20-C PRENSA", "20-D CALIBRADO", "20-E GBD", "20-F MARCADORA LASER"],
                        "SOLDADURA Y ROLADO": ["WEIL LASER #1", "WEIL LASER #2", "WEIL LASER #3", "SIZER DOUBLE #1", "SIZER DOUBLE #2", "WEIL PLASMA #1", "WEIL PLASMA #2", "PRESS 50 TON #1", "PRESS 50 TON #2"],
                        "CORTE LASER": ["LASER #2", "LASER #3", "LASER #4", "PROTO-1", "TAB BENDING MACHINE"]
                    },
                    "EPU4": {
                        "AUDI": ["CELDA NC1854", "OP.50 LEAK TEST", "OP. 20A LEAK TEST FINAL", "OP. 20B LEAK TEST ESTACION V6", "OP.40 CRIMPING -10000409-1", "OP. 30 CRIMPING SPOT WELDING 10000409-2", "OP.35 BANDIG STATION"],
                        "BMW": ["CELDA NC386", "CELDA NC240", "CELDA NC237", "CELDA NC602", "BORRIES STATION_L1 (ESTACION DE MARCADO)", "BORRIES STATION_L2 (ESTACION DE MARCADO)", "LEAK TEST DOCKING STATION_MFLR LINE 2", "LEAK TEST DOCKING STATION_MFLR LINE 1", "INVENIO OP1-WRAP BRICK", "SIZER SINGLE RB80 IO LINE 1", "BANDING IT STATION", "PRE MOUTING INSULATION SHELLS", "ESTACION DE PRE-ENSAMBLE", "SIZING MACHINE LINE 2", "OP. 30 CRIMPING SPOT WELDING 10000409-2"],
                        "VW": {
                            "VW EA211": ["CELDA NC1876 OP. 10.1", "CELDA NC1876 OP. 10.2", "CELDA NC1877 OP. 20.1","CELDA NC1877 OP. 20.2", "CELDA NC1878 OP. 30.1", "CELDA NC1878 OP. 30.2", "iES SIZER SINGLE RB80 IO/18172-01", "iES FLARE  RB80 SF+SF VS 18174-01", "iES NOTCHER RB80 ATN", "OP.50 GAGE", "OP. 40.1 PIPE TORQUE", "OP. 40.2 LEAK TESTER", "OP. 60.1 BRACKETS TORQUE", "OP. 60.2 MARCADORA LASER"],
                            "VW EA888": ["CELDA NC1871 OP.10.1", "CELDA NC1871 OP.10.2", "CELDA  NC1872 OP.20.1", "CELDA  NC1872 OP.20.2", "CELDA  NC1873 OP.30.1", "CELDA  NC1873 OP.30.2", "CELDA  NC1874 OP.40.1", "CELDA  NC1874 OP.40.2", "CELDA NC1965 OP. 10.1-ATLAS", "CELDA NC1965 OP. 10.2-ATLAS", "CELDA NC1966 OP. 20.1- ATLAS", "CELDA NC1966 OP. 20.2- ATLAS", "CELDA NC1967 OP. 30.1- ATLAS", "CELDA NC1967 OP. 30.2- ATLAS", "CELDA NC1968 OP. 40.1- ATLAS", "CELDA NC1968 OP. 40.2- ATLAS", "CD WELDER", "LEAK TEST CD WELDER", "OP. 50.1 PIPE TORQUE", "OP.50.2 LEAK TESTER", "OP.60 CRIMPING MACHINE 48284-1", "OP.70 GAGE", "OP. 80.1 BRACKETS TORQUE 1", "OP. 80.2 BRACKETS TORQUE 2", "OP. 80.3 MARCADORA LASER"],
                            "SOLDADURA Y ROLADO": ["WEIL LASER #4", "SIZER DOUBLE  RB150 IO+IO"],
                            "CANNING": ["WEIL LASER #4", "SIZER DOUBLE  RB150 IO+IO", "KIRSCHENHOFER #1", "KIRSCHENHOFER #5"]
                        },
                        "CUMMINS": ["CELDA NC1303", "LEAK TEST 1 (CELDA)", "iES SIZER DOUBLE RB250 IO+IO/18506-01", "LEAK TEST 2 (CANNING)", "FLEXMASTER 400/1250", "ROLLFORMING MACHINE 400/600", "PLASNISHER (ALISADO DE COSTURAS 3855)", "KIRSCHENHOFER #6"]
                    },
                    "FACILIDADES": {}
                }
    
                if unidad == "FACILIDADES":
                    linea = st.text_input("Linea (especificar):", key="linea_fac")
                    maquina = st.text_input("Máquina (especificar):", key="maquina_fac")
                else:
                    lineas_disponibles = list(datos_cascada.get(unidad, {}).keys())
                    linea = st.selectbox("Linea:", lineas_disponibles, key="linea_select")
    
                    if unidad == "EPU4" and linea == "VW":
                        area_vw = st.selectbox("Area:", list(datos_cascada["EPU4"]["VW"].keys()), key="area_vw")
                        maquina = st.selectbox("Maquina:", datos_cascada["EPU4"]["VW"][area_vw], key="maquina_select_vw")
                    else:
                        maquinas_disponibles = datos_cascada.get(unidad, {}).get(linea, [])
                        maquina = st.selectbox("Maquina:", maquinas_disponibles, key="maquina_select")
    
            else:
                # mantener variables definidas aunque no se muestren
                unidad = ""
                linea = ""
                maquina = ""
    
            # turno y supervisor (solo aquí, con keys únicas)
            turno = st.selectbox("Turno:", ["1", "2", "3"], key="turno_select")
            supervisores_disponibles = ["Ramiro Carrillo", "Juan Mendoza", "Pedro Arredondo", "Jose Cisneros", "Javier Ramirez", "Jose Donías", "Jorge Días", "Josue Esquivel", "Ivan Guajardo"]
            supervisor = st.selectbox("Supervisor:", supervisores_disponibles, key="supervisor_select")
    
            # --------- FORMULARIO PRINCIPAL ----------
            # Fuera del formulario: selector de Código de Tecnología (para que actualice la app al cambiar)
            codigos_falla_dict = {
                "Mechanical": ["Chumacera_01", "Motor_02", "Servomotor_17", "Tornilleria_04", "Rodamiento_05","Cople_06", "Husillo_07", "Transmisión_08","Rodillo _09","Soporte/Base_10","Clampeo_11","Balero Lineal_12","Guia Lineal_13","Ajuste Mecanico_14","Chiller/ HVAC_15","Extractor de polvo_16","Deasgaste_17","Lubricación_18" ],
                "Electrical": ["Suministro_energía_01", "Sensor_02", "RFID_11","Fuente Poder_24VAC_04","Relevador_MCR_05","Relevador_Control_06","Relevador_Seguridad_07","Desconectador_08","Transformador_09","Cable_10","Interlock_11","UPS_12","10_Tablero electrico"],
                "Pneumatic": ["Suministro aire_01", "FLR (Filer Lubricator Regulator)_02", "Cilindro_03", "Valvula_04", "Electrovalvula_05", "Maniful_Electrovalvula_06","Conector_07" "Manguera_08", "Ventosa_09", "Bomba_10"],
                "Hydraulic": ["Cilindro_01", "Electrovalvula_02", "Maniful_Electrovalvula_03", "Conector_04", "Manguera_05", "Filtro_06", "Solenoide_07"],
                "Control_and_PLC": ["PLC_01", "Tarjeta_IO_02", "Tarjeta_Comunicacion_03", "Camara_Vision_04", "Sensor_05", "Escaner_Codigo_06","Escaner_seguridad_07", "Variador_Frecuencia_08", "Master_IO_Link_09", "Slave_IO_Link_10", "RFID_11", "Transductor_12", "Herramienta_Torque_13", "Controlador_Torque_14", "Switch_Comunicacion_15", "HMI_16", "Servomotor_17", "Servocontrolador_18", "Seguridad_19", "Secuencia", "Pendant Controller_13", "Encoder_14", "Cable comunicación_15", "Controlador_16", "Drive_17"],
                "Dimensional": ["01_Material Fuera especificacion", "02_Escalon", "03_Traslape", "04_Desalineación lineal", "05_Falta Penetracion", "06_Socavado", "07_Poros", "08_Soldadura Incompleta", "09_Soldadura Desalineada", "10_Fisura", "11_Mat Expuesto", "12_Colapso_Canning"],
                "Laser_Welding": ["01_Generador", "02_Resonador", "03_Cabezal", "04_Fibra Optica", "05_Fuente Laser"],
                "Electrical_Welding": ["01_Tambo Microalambre", "02_Punta Contacto", "03_Tobera", "04_Difusor", "05_Linner", "06_Antorcha", "07_Robacta", "08_Alimentador", "09_Guia Microalambre (Conduit)", "10_Fuente poder", "11_Reamer"],
                "Inductor_Welding": ["01_Generador", "02_Inductor"],
                "Robot": ["01_S-Axis", "02_L-Axis", "03_U-Axis", "04_R-Axis", "05_B-Axis", "06_T-Axis", "07_I/O", "08_Seguridad", "09_Freno", "10_Tablero electrico", "11_Fusible", "12_Ventilador", "13_Pendant Controller", "14_Controller Battery", "15_Regenerative Resistor", "16_Capacitor Unit", "17_QS1", "18_Controller Door Open/Close", "19_Controller Power ON/OFF", "20_Extended Axis", "21_Program", "22_Colisión"],
                "Pin_Marking": ["Punta"],
                "Laser_Marking": ["Controlador", "Cabezal", "Tarjeta_IO_02", "Tarjeta_Comunicacion_03"],
                "Cambio de Modelo": ["Cambio programado"],
                "Spot Welding": ["01_Electrodo"],
                "Operación": ["Mala Operación"]
            }
            
            if tipo_mantenimiento == "Correctivo (DOWNTIME)":
                st.markdown("### Código de Tecnología:")
                codigo_tecnologia = st.selectbox(
                    "",
                    list(codigos_falla_dict.keys()),
                    key="codigo_tecnologia"
                )
            else:
                codigo_tecnologia = None
    
            with st.form("actividad_form", clear_on_submit=True):
                # Fecha
                fecha_actual = datetime.date.today()
                fecha_registro = st.date_input("Fecha:", value=fecha_actual, key="fecha_registro_form")
    
                # Semana
                semana_actual = fecha_registro.isocalendar()[1]
                st.number_input("Semana:", value=semana_actual, disabled=True, key="semana_input_form")
    
                # Fecha de inicio y Fecha de fin
                col_fechas1, col_fechas2 = st.columns(2)
                with col_fechas1:
                    fecha_inicio = st.date_input("Fecha de inicio", datetime.date.today(), key="fecha_inicio_form")
                    fecha_fin = st.date_input("Fecha de fin", fecha_inicio, key="fecha_fin_form")
                with col_fechas2:
                    hora_inicio = st.time_input("Hora de inicio", datetime.time(8, 0), key="hora_inicio_form")
                    hora_fin = st.time_input("Hora de fin", datetime.time(9, 0), key="hora_fin_form")
    
                # Duración total (minutos) calculada automáticamente
                inicio_dt = datetime.datetime.combine(fecha_inicio, hora_inicio)
                fin_dt = datetime.datetime.combine(fecha_fin, hora_fin)
                duracion_min = max(0, (fin_dt - inicio_dt).total_seconds() / 60)
                st.number_input("Duración total (min):", value=round(duracion_min, 2), disabled=True, key="duracion_total_form")
    
                # Orden de trabajo, clasificación y demás campos correctivo
                if tipo_mantenimiento == "Correctivo (DOWNTIME)":
                    orden_trabajo = st.text_input("Orden de Trabajo:", key="orden_trabajo_form")
                    clasificacion = st.selectbox("Clasificación:", ["MANTENIMIENTO", "TOOLING", "CONTROLES", "SETTER"], key="clasificacion_form")
    
                    codigos_falla = codigos_falla_dict.get(codigo_tecnologia, [])
                    opciones_falla = ["-- Seleccione --"] + codigos_falla
    
                    codigo_falla_sel = st.selectbox("Código de Falla:", opciones_falla, key="codigo_falla_form")
                    codigo_falla_val = "" if codigo_falla_sel == "-- Seleccione --" else codigo_falla_sel
    
                    descripcion_falla = st.text_area("Descripción de Falla:", key="descripcion_falla_form")
                    causa_raiz = st.text_area("Causa Raíz:", key="causa_raiz_form")
                    accion_inmediata = st.text_area("Acción Inmediata:", key="accion_inmediata_form")
                else:
                    orden_trabajo = ""
                    clasificacion = ""
                    codigo_falla_val = ""
                    descripcion_falla = ""
                    causa_raiz = ""
                    accion_inmediata = ""
    
                # Descripción de la actividad
                actividad = st.text_area("Descripción de la actividad", key="descripcion_actividad_form")
    
                submitted = st.form_submit_button("Guardar actividad")
    
                if submitted:
                    inicio = datetime.datetime.combine(fecha_inicio, hora_inicio)
                    fin = datetime.datetime.combine(fecha_fin, hora_fin)
                    if fin > inicio:
                        duracion_min = (fin - inicio).total_seconds() / 60
                        unique_id = f"ACT-{uuid.uuid4().hex[:6].upper()}"
    
                        nueva_fila = {
                            "ID": unique_id,
                            "Usuario": st.session_state["usuario"],
                            "Nombre": tecnico_nombre,
                            "Unidad": unidad,
                            "area": st.session_state["area"],
                            "Turno": turno,
                            "Supervisor": supervisor,
                            "Línea": linea,
                            "Actividad": actividad,
                            "Tipo": tipo_mantenimiento,
                            "Número de orden": orden_trabajo if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Máquina": maquina,
                            "Inicio": inicio.strftime("%d-%m-%Y %H:%M"),
                            "Fin": fin.strftime("%d-%m-%Y %H:%M"),
                            "Duración (min)": round(duracion_min, 2),
                            "Fecha": fecha_registro.strftime("%d-%m-%Y"),
                            "Semana": semana_actual,
                            "Clasificación": clasificacion if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Código de Tecnología": codigo_tecnologia if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Código de Falla": codigo_falla_val if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Descripción de Falla": descripcion_falla if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Causa Raíz": causa_raiz if tipo_mantenimiento == "Correctivo (DOWNTIME)" else "",
                            "Acción Inmediata": accion_inmediata if tipo_mantenimiento == "Correctivo (DOWNTIME)" else ""
                        }
    
                        guardar_registro_tecnico(nueva_fila)
                        st.success(f"Actividad registrada ✅ | ID: {unique_id} | Duración: {round(duracion_min,2)} min")
                        st.rerun()
                    else:
                        st.error("La hora de fin debe ser mayor que la hora de inicio ❌")
    
        # TAB 2: HISTORIAL TÉCNICO
        with tab2:
            st.subheader("📋 Historial de actividades")
            df = cargar_datos_tecnicos()
    
            if not df.empty:
                df["Fecha"] = pd.to_datetime(df["Inicio"], format="%d-%m-%Y %H:%M", errors='coerce').dt.date
    
                modo_historial = st.selectbox(
                    "Modo de historial",
                    ["Día específico", "Seleccionar rango"]
                )
    
                if modo_historial == "Día específico":
                    fecha_seleccionada = st.date_input("Selecciona un día", datetime.date.today(), key="fecha_historial")
                    df_filtrado = df[df["Fecha"] == fecha_seleccionada]
    
                elif modo_historial == "Seleccionar rango":
                    col1, col2 = st.columns(2)
                    with col1:
                        fecha_inicio = st.date_input("Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio_historial")
                    with col2:
                        fecha_fin = st.date_input("Fecha fin", datetime.date.today(), key="fecha_fin_historial")
                    df_filtrado = df[(df["Fecha"] >= fecha_inicio) & (df["Fecha"] <= fecha_fin)]
    
                df_filtrado = df_filtrado[df_filtrado["Usuario"] == st.session_state["usuario"]]
    
                if not df_filtrado.empty:
                    if modo_historial == "Día específico":
                        st.markdown(f"### Actividades del {fecha_seleccionada}")
                    else:
                        st.markdown(f"### Actividades del rango {fecha_inicio} a {fecha_fin}")
    
                    vista = st.selectbox(
                        "📌 Selecciona modo de vista:",
                        ["🔖 Ver por línea", "📑 Ver todo"],
                        index=0
                    )
    
                    if vista == "🔖 Ver por línea":
                        for linea, df_linea in df_filtrado.groupby("Línea"):
                            nombre_linea = linea if linea else "No especificada"
                            with st.expander(f"📁 Línea: {nombre_linea}"):
                                for idx, row in df_linea.iterrows():
                                    etiqueta_usuario = f"{row['Usuario']} | {row.get('Nombre','')}"
                                    with st.expander(f"📝 {etiqueta_usuario} | {row['Actividad']} | {row['Inicio']}"):
                                        st.markdown(f"""
                                        <div class="registro-container">
                                        <h4 style="color:#00BFFF; margin-bottom:10px;">Actividad: {row['Actividad']}</h4>
                                        <p><b>ID:</b> {row['ID']}</p>
                                        <p><b>Usuario (MX):</b> {row['Usuario']}</p>
                                        <p><b>Nombre:</b> {row.get('Nombre','')}</p>
                                        <p><b>Supervisor:</b> {row.get('Supervisor','No especificado')}</p>
                                        <p><b>Turno:</b> {row.get('Turno','No especificado')}</p>
                                        <p><b>Tipo:</b> {row['Tipo']}</p>
                                        <p><b>Unidad:</b> {row['Unidad'] if 'Unidad' in row and row['Unidad'] else 'No especificada'}</p>
                                        <p><b>Línea:</b> {row['Línea'] if row['Línea'] else 'No especificada'}</p>
                                        <p><b>Máquina:</b> {row['Máquina']}</p>
                                        <p><b>Inicio:</b> {row['Inicio']}</p>
                                        <p><b>Fin:</b> {row['Fin']}</p>
                                        <p><b>Duración (min):</b> {row['Duración (min)']}</p>
                                        """
                                        , unsafe_allow_html=True)
    
                                        # Mostrar campos adicionales si existen
                                        if row['Tipo'] == "Correctivo (DOWNTIME)":
                                         
                                            st.markdown(f"<p><b>Número de orden:</b> {row.get('Número de orden', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Clasificación:</b> {row.get('Clasificación', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Código de Tecnología:</b> {row.get('Código de Tecnología', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Código de Falla:</b> {row.get('Código de Falla', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Descripción de Falla:</b> {row.get('Descripción de Falla', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Causa Raíz:</b> {row.get('Causa Raíz', '')}</p>", unsafe_allow_html=True)
                                            st.markdown(f"<p><b>Acción Inmediata:</b> {row.get('Acción Inmediata', '')}</p>", unsafe_allow_html=True)
    
                                        st.markdown("<​​/div>", unsafe_allow_html=True)
    
                                        if st.button(f"Eliminar {row['ID']}", key=f"del_{row['ID']}_{idx}"):
                                            eliminar_registro_tecnico(row["ID"] )
    
                    else:
                        df_sorted = df_filtrado.sort_values(by="Inicio", ascending=False)
                        for idx, row in df_sorted.iterrows():
                            etiqueta_usuario = f"{row['Usuario']} | {row.get('Nombre','')}"
                            with st.expander(f"📝 {etiqueta_usuario} | {row['Actividad']} | {row['Inicio']}"):
                                st.markdown(f"""
                                <div class="registro-container">
                                <h4 style="color:#00BFFF; margin-bottom:10px;">Actividad: {row['Actividad']}</h4>
                                <p><b>ID:</b> {row['ID']}</p>
                                <p><b>Usuario (MX):</b> {row['Usuario']}</p>
                                <p><b>Nombre:</b> {row.get('Nombre','')}</p>
                                <p><b>Supervisor:</b> {row.get('Supervisor','No especificado')}</p>
                                <p><b>Turno:</b> {row.get('Turno','No especificado')}</p>
                                <p><b>Tipo:</b> {row['Tipo']}</p>
                                <p><b>Unidad:</b> {row['Unidad'] if 'Unidad' in row and row['Unidad'] else 'No especificada'}</p>
                                <p><b>Línea:</b> {row['Línea'] if row['Línea'] else 'No especificada'}</p>
                                <p><b>Máquina:</b> {row['Máquina']}</p>
                                <p><b>Inicio:</b> {row['Inicio']}</p>
                                <p><b>Fin:</b> {row['Fin']}</p>
                                <p><b>Duración (min):</b> {row['Duración (min)']}</p>
                                """
                                , unsafe_allow_html=True)
    
                                # Mostrar campos adicionales si existen
                                if row['Tipo'] == "Correctivo (DOWNTIME)":
                                    st.markdown(f"<p><b>Número de orden:</b> {row.get('Número de orden', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Clasificación:</b> {row.get('Clasificación', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Código de Tecnología:</b> {row.get('Código de Tecnología', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Código de Falla:</b> {row.get('Código de Falla', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Descripción de Falla:</b> {row.get('Descripción de Falla', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Causa Raíz:</b> {row.get('Causa Raíz', '')}</p>", unsafe_allow_html=True)
                                    st.markdown(f"<p><b>Acción Inmediata:</b> {row.get('Acción Inmediata', '')}</p>", unsafe_allow_html=True)
    
                                #st.markdown("<​​/div>", unsafe_allow_html=True)
                                st.markdown(unsafe_allow_html=True)
    
                                if st.button(f"Eliminar {row['ID']}", key=f"del_{row['ID']}_{idx}"):
                                    eliminar_registro_tecnico(row["ID"] )
                else:
                    if modo_historial == "Día específico":
                        st.info(f"No hay actividades registradas para el día {fecha_seleccionada}.")
                    else:
                        st.info(f"No hay actividades registradas entre {fecha_inicio} y {fecha_fin}.")
            else:
                st.info("No hay actividades registradas aún.")

        # TAB 3: RENDIMIENTO TÉCNICO
        with tab3:
            st.subheader("📊 Gráficas de productividad")

            df = cargar_datos_tecnicos()
            if not df.empty:
                modo = st.selectbox(
                    "Modo de rendimiento",
                    ["Hoy", "Semana actual", "Mes actual", "Fecha específica", "Seleccionar rango"]
                )

                if modo == "Hoy":
                    fecha_inicio = fecha_fin = datetime.date.today()
                elif modo == "Semana actual":
                    fecha_fin = datetime.date.today()
                    fecha_inicio = fecha_fin - datetime.timedelta(days=fecha_fin.weekday())
                elif modo == "Mes actual":
                    fecha_fin = datetime.date.today()
                    fecha_inicio = fecha_fin.replace(day=1)
                elif modo == "Fecha específica":
                    fecha_inicio = fecha_fin = st.date_input(
                    "Selecciona la fecha", datetime.date.today(), key="fecha_rendimiento"
                    )
                elif modo == "Seleccionar rango":
                    col1, col2 = st.columns(2)
                    with col1:
                        fecha_inicio = st.date_input(
                        "Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio"
                        )
                    with col2:
                        fecha_fin = st.date_input(
                        "Fecha fin", datetime.date.today(), key="fecha_fin"
                        )

                df["Fecha"] = pd.to_datetime(df["Inicio"], format="%d-%m-%Y %H:%M", errors='coerce').dt.date
                df_rango = df[(df["Fecha"] >= fecha_inicio) & (df["Fecha"] <= fecha_fin)]

                df_rango = df_rango[df_rango["Usuario"] == st.session_state["usuario"]]
                st.info("Solo puedes ver tu propio rendimiento por privacidad.")

                if not df_rango.empty:
                    resumen = df_rango.groupby(["Usuario","Nombre"]).agg({
                    "Duración (min)": "sum",
                    "ID": "count"
                    }).reset_index().rename(columns={"ID":"Actividades"})

                    resumen["Etiqueta"] = resumen["Usuario"] + " | " + resumen["Nombre"]

                    chart1 = alt.Chart(resumen).mark_bar().encode(
                    x="Etiqueta:N",
                    y="Duración (min):Q",
                    color=alt.Color("Etiqueta:N", scale=alt.Scale(scheme="dark2"))
                    ).properties(title="Minutos invertidos por usuario")

                    chart2 = alt.Chart(resumen).mark_bar().encode(
                    x="Actividades:Q",
                    y=alt.Y("Etiqueta:N", sort="-x"),
                    color=alt.Color("Etiqueta:N", scale=alt.Scale(scheme="set2")),
                    tooltip=["Etiqueta","Actividades"]
                    ).properties(title="Número de actividades por usuario")

                    labels = alt.Chart(resumen).mark_text(
                    align="left", baseline="middle", dx=3
                    ).encode(
                    x="Actividades:Q",
                    y="Etiqueta:N",
                    text="Actividades:Q"
                    )

                    st.altair_chart(chart1 | (chart2 + labels), use_container_width=True)
                else:
                    st.info("No hay datos en el rango seleccionado.")
            else:
                st.warning("No hay datos suficientes para calcular rendimiento.")

        # TAB 4: DESCARGAS TÉCNICO
        with tab4:
            st.subheader("📂 Descargar reportes")
            st.info("⚠️ Solo los administradores pueden descargar reportes.")

    # ==== PANEL SUPERVISOR ====
    # ==== PANEL SUPERVISOR ====
    # ==== PANEL SUPERVISOR ====
    # ==== PANEL SUPERVISOR ====
    elif st.session_state["role"] == "supervisor":
        tab1, tab2, tab3, tab4 = st.tabs(["📝 Registro", "📋 Historial", "📊 Rendimiento", "📂 Descargas"])
    
        # TAB 1: REGISTRO SUPERVISOR
        with tab1:
            st.subheader("Registro Diario de Plantilla")
    
            with st.form("supervisor_form", clear_on_submit=True):
                col1, col2 = st.columns(2)
    
                with col1:
                   # area_sup = st.selectbox("Área:", ["EPU2", "EPU3", "EPU4"])
                    zona_sup = st.selectbox("EPU (Zona):", ["EPU2", "EPU3", "EPU4", "FACILIDADES"], key="zona_sup_form")
                    supervisor_nombre = st.text_input("Supervisor:", value=st.session_state.get("nombre", ""), disabled=True)
                with col2:
                    fecha = st.date_input("Fecha:", datetime.date.today())
                    turno = st.selectbox("Turno:", ["1", "2", "3"])
                    jornada = st.selectbox("Jornada:", ["8 horas", "12 horas"])
    
                st.markdown("---")
                st.markdown("### Registra Cantidad de Técnicos en Turno")
    
                col3, col4 = st.columns(2)
    
                with col3:
                    st.markdown("**Tipo de Jornada**")
                    jornada_normal = st.number_input("Total Plantilla:", min_value=0, value=0, step=1)
                    cursos_otros = st.number_input("Cursos/Otros:", min_value=0, value=0, step=1)
                    inasistencias = st.number_input("Inasistencias (Falta, vacaciones, permisos, etc):", min_value=0, value=0, step=1)
    
                with col4:
                    st.markdown("**Técnicos**")
                    total_esperado = jornada_normal
                    st.metric("Total Esperado", total_esperado)
    
                    total_real = jornada_normal - cursos_otros - inasistencias
                    st.metric("Total Real", total_real)
    
                    multiplicador = 7.5 if jornada == "8 horas" else 11.5
                    horas_disponibles = total_real * multiplicador
                    st.metric("Total de Horas Disponibles", f"{horas_disponibles:.1f} hrs")
    
                st.markdown("---")
                st.markdown("### Registrar Eventos Programados")
                col5, col6 = st.columns(2)
                with col5:
                    evento_actividad = st.text_input("Actividad del evento:")
                with col6:
                    evento_duracion = st.number_input("Duración del evento (horas):", min_value=0.0, value=0.0, step=0.5)
    
                horas_evento = evento_duracion * total_real
                horas_disponibles_finales = horas_disponibles - horas_evento
    
                st.metric("Horas Disponibles después del evento", f"{horas_disponibles_finales:.1f} hrs")
    
                submitted = st.form_submit_button("Guardar registro")
    
                if submitted:
                    unique_id = f"SUP-{uuid.uuid4().hex[:6].upper()}"
                    nuevo_registro = {
                        "ID": unique_id,
                        "Usuario": st.session_state["usuario"],
                        "Nombre": st.session_state.get("nombre", ""),
                        "zona": zona_sup,
                        "Supervisor": st.session_state.get("nombre", ""),
                        "Fecha": fecha.strftime("%d-%m-%Y"),
                        "Turno": turno,
                        "Jornada": jornada,
                        "Jornada_Normal": jornada_normal,
                        "Cursos_Otros": cursos_otros,
                        "Inasistencias": inasistencias,
                        "Total_Esperado": total_esperado,
                        "Total_Real": total_real,
                        "Horas_Disponibles": round(horas_disponibles_finales, 2),
                        "Evento_Actividad": evento_actividad,
                        "Evento_Duracion": evento_duracion,
                        "Horas_Evento": round(horas_evento, 2)
                    }
                    guardar_registro_supervisor(nuevo_registro)
                    st.success(f"Registro guardado ✅ | ID: {unique_id}")
                    st.rerun()
    
        # TAB 2: HISTORIAL SUPERVISOR
        with tab2:
            st.subheader("📋 Historial de registros")
            df = cargar_datos_supervisores()
            if not df.empty:
                df["Fecha_obj"] = pd.to_datetime(df["Fecha"], format="%d-%m-%Y", errors='coerce').dt.date
                modo_historial = st.selectbox("Modo de historial", ["Día específico", "Seleccionar rango"])
    
                if modo_historial == "Día específico":
                    fecha_sel = st.date_input("Selecciona un día", datetime.date.today(), key="h_sup_1")
                    df_filtrado = df[df["Fecha_obj"] == fecha_sel]
                else:
                    c1, c2 = st.columns(2)
                    f_ini = c1.date_input("Inicio", datetime.date.today() - datetime.timedelta(days=7), key="h_sup_2")
                    f_fin = c2.date_input("Fin", datetime.date.today(), key="h_sup_3")
                    df_filtrado = df[(df["Fecha_obj"] >= f_ini) & (df["Fecha_obj"] <= f_fin)]
    
                if not df_filtrado.empty:
                    for idx, row in df_filtrado.iterrows():
                        with st.expander(f"📝 {row['area']} | Turno {row['Turno']} | {row['Fecha']}"):
                            st.write(f"**ID:** {row['ID']} | **Total Real:** {row['Total_Real']} técnicos | **Horas Disp:** {row['Horas_Disponibles']} hrs")
                            if st.button(f"Eliminar {row['ID']}", key=f"del_{row['ID']}"):
                                eliminar_registro_supervisor(row["ID"])
                                st.rerun()
                else:
                    st.info("No hay registros en este periodo.")
            else:
                st.info("No hay registros aún.")
    
        # TAB 3: RENDIMIENTO SUPERVISOR (MEJORADO, con gráficos individuales)
        # TAB 3: RENDIMIENTO SUPERVISOR (MEJORADO, con gráficos individuales) - CORREGIDO
        # TAB 3: RENDIMIENTO SUPERVISOR (MEJORADO, con filtros y gráfica dinámica)


        with tab3:
                    st.subheader("📊 Análisis de Rendimiento")
        
                    nombre_actual = st.session_state.get("nombre", "")
                    area_actual = st.session_state.get("area", "")
        
                    # Vista elegante con botones para seleccionar análisis
                    opcion_rendimiento = st.radio(
                        "Selecciona vista:",
                        ["Mi Gestión de Plantilla", "Rendimiento de Técnicos (Mi Equipo)"],
                        horizontal=True
                    )
        
                    # Paleta base (se usarán escalas categóricas para asignar color distinto por barra)
                    default_palette = "category20"  # Altair/Vega schemes: category20, tableau20, category10, etc.
        
                    # -----------------------
                    # Mi Gestión de Plantilla
                    # -----------------------
                    if opcion_rendimiento == "Mi Gestión de Plantilla":
                        df_sup = cargar_datos_supervisores()
                        if df_sup is None or df_sup.empty:
                            st.info("No hay registros de gestión de plantilla.")
                        else:
                            df_sup = df_sup.copy()
        
                            # Asegurarse de tener Fecha_obj: intentar normalizar desde 'Fecha' o 'Inicio'
                            if "Fecha_obj" not in df_sup.columns:
                                if "Fecha" in df_sup.columns:
                                    df_sup["Fecha_obj"] = pd.to_datetime(df_sup["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                                elif "Inicio" in df_sup.columns:
                                    df_sup["Fecha_obj"] = pd.to_datetime(df_sup["Inicio"], errors="coerce").dt.date
                                else:
                                    # Si no hay columna de fecha, crear columna vacía para evitar KeyError posteriores
                                    df_sup["Fecha_obj"] = pd.NaT
        
                            # Filtro de fechas con default = Hoy (por eso por defecto solo se muestra el registro del día actual)
                            modo_fecha_plantilla = st.selectbox(
                                "Selecciona rango de fechas:",
                                ["Hoy", "Semana actual", "Mes actual", "Fecha específica", "Seleccionar rango"],
                                index=0
                            )
                            hoy = datetime.date.today()
                            if modo_fecha_plantilla == "Hoy":
                                fecha_ini_plt = fecha_fin_plt = hoy
                            elif modo_fecha_plantilla == "Semana actual":
                                fecha_fin_plt = hoy
                                fecha_ini_plt = fecha_fin_plt - datetime.timedelta(days=fecha_fin_plt.weekday())
                            elif modo_fecha_plantilla == "Mes actual":
                                fecha_fin_plt = hoy
                                fecha_ini_plt = fecha_fin_plt.replace(day=1)
                            elif modo_fecha_plantilla == "Fecha específica":
                                fecha_ini_plt = fecha_fin_plt = st.date_input("Selecciona la fecha", hoy, key="plt_fecha_especifica")
                            else:
                                c1, c2 = st.columns(2)
                                with c1:
                                    fecha_ini_plt = st.date_input("Fecha inicio", hoy - datetime.timedelta(days=7), key="plt_fecha_ini")
                                with c2:
                                    fecha_fin_plt = st.date_input("Fecha fin", hoy, key="plt_fecha_fin")
        
                            # Filtrar sólo los registros del supervisor y por fechas (por defecto esto traerá solo hoy)
                            # Protegemos contra NaT en Fecha_obj usando .notna() y comparaciones seguras
                            df_mio = df_sup[
                                (df_sup.get("Supervisor") == nombre_actual) &
                                (df_sup["Fecha_obj"].notna()) &
                                (df_sup["Fecha_obj"] >= fecha_ini_plt) &
                                (df_sup["Fecha_obj"] <= fecha_fin_plt)
                            ].copy()
        
                            if df_mio.empty:
                                st.info("No hay registros de plantilla para el periodo seleccionado.")
                            else:
                                # Normalizar columnas numéricas
                                df_mio["Jornada_Normal"] = pd.to_numeric(df_mio.get("Jornada_Normal", 0), errors="coerce").fillna(0)
                                df_mio["Cursos_Otros"] = pd.to_numeric(df_mio.get("Cursos_Otros", 0), errors="coerce").fillna(0)
                                df_mio["Inasistencias"] = pd.to_numeric(df_mio.get("Inasistencias", 0), errors="coerce").fillna(0)
                                df_mio["Total_Esperado"] = pd.to_numeric(df_mio.get("Total_Esperado", df_mio["Jornada_Normal"]), errors="coerce").fillna(0)
                                df_mio["Total_Real"] = pd.to_numeric(df_mio.get("Total_Real", 0), errors="coerce").fillna(0)
                                df_mio["Horas_Disponibles"] = pd.to_numeric(df_mio.get("Horas_Disponibles", 0), errors="coerce").fillna(0)
                                df_mio["Horas_Evento"] = pd.to_numeric(df_mio.get("Horas_Evento", 0), errors="coerce").fillna(0)
        
                                # Métricas para el periodo seleccionado (por defecto hoy)
                                total_registros = len(df_mio)
                                total_esperado = int(df_mio["Total_Esperado"].sum())
                                total_real = int(df_mio["Total_Real"].sum())
                                total_horas_disponibles = df_mio["Horas_Disponibles"].sum()
                                total_horas_evento = df_mio["Horas_Evento"].sum()
                                total_cursos = int(df_mio["Cursos_Otros"].sum())
                                total_inasistencias = int(df_mio["Inasistencias"].sum())
                                avg_horas_por_tecnico = (total_horas_disponibles / total_real) if total_real else 0
        
                                # Mostrar métricas en forma de tarjetas
                                r1, r2, r3, r4 = st.columns(4)
                                r1.metric("Registros (periodo)", f"{total_registros}")
                                r2.metric("Técnicos esperados", f"{total_esperado}")
                                r3.metric("Técnicos reales", f"{total_real}")
                                r4.metric("Horas disponibles (total)", f"{total_horas_disponibles:.1f} hrs")
        
                                r5, r6, r7 = st.columns(3)
                                r5.metric("Horas en eventos", f"{total_horas_evento:.1f} hrs")
                                r6.metric("Cursos / Otros", f"{total_cursos}")
                                r7.metric("Inasistencias", f"{total_inasistencias}")
        
                                st.markdown("### Evolución y detalle")
        
                                # Serie temporal: Horas Disponibles y Técnicos Reales por día (aplica cuando rango > 1 día)
                                # Asegurarse de que Fecha_obj sea columna de serie_horas y series_tecnicos
                                if "Fecha_obj" in df_mio.columns:
                                    series_horas = df_mio.groupby("Fecha_obj")["Horas_Disponibles"].sum().reset_index().sort_values("Fecha_obj")
                                    series_tecnicos = df_mio.groupby("Fecha_obj")["Total_Real"].sum().reset_index().sort_values("Fecha_obj")
                                else:
                                    series_horas = pd.DataFrame()
                                    series_tecnicos = pd.DataFrame()
        
                                if not series_horas.empty:
                                    chart_horas = alt.Chart(series_horas).mark_bar().encode(
                                        x=alt.X("Fecha_obj:T", title="Fecha"),
                                        y=alt.Y("Horas_Disponibles:Q", title="Horas Disponibles"),
                                        color=alt.value("#1f77b4"),
                                        tooltip=[alt.Tooltip("Fecha_obj:T", title="Fecha"), alt.Tooltip("Horas_Disponibles:Q", title="Horas")]
                                    ).properties(height=260)
                                    st.altair_chart(chart_horas, use_container_width=True)
        
                                if not series_tecnicos.empty:
                                    chart_tt = alt.Chart(series_tecnicos).mark_line(point=True, color="#ff7f0e").encode(
                                        x=alt.X("Fecha_obj:T", title="Fecha"),
                                        y=alt.Y("Total_Real:Q", title="Técnicos Reales"),
                                        tooltip=[alt.Tooltip("Fecha_obj:T", title="Fecha"), alt.Tooltip("Total_Real:Q", title="Técnicos reales")]
                                    ).properties(height=220)
                                    st.altair_chart(chart_tt, use_container_width=True)
        
                                st.markdown("### Detalle de registros (ordenado)")
        
                                # Preparar columnas a mostrar
                                display_cols = ["Fecha", "area", "Turno", "Jornada", "Jornada_Normal", "Cursos_Otros", "Inasistencias", "Total_Esperado", "Total_Real", "Horas_Disponibles", "Evento_Actividad", "Evento_Duracion", "Horas_Evento"]
                                available_cols = [c for c in display_cols if c in df_mio.columns]
        
                                # Ordenar el DataFrame por Fecha_obj si existe; hacer la selección de columnas después de ordenar
                                df_display = df_mio.copy()
                                if "Fecha_obj" in df_display.columns:
                                    # Si 'Fecha' no está en available_cols pero existe en df_display, podemos incluirla para visualización
                                    if "Fecha" not in available_cols and "Fecha" in df_display.columns:
                                        available_cols = ["Fecha"] + available_cols
                                    df_display = df_display.sort_values("Fecha_obj", ascending=False).reset_index(drop=True)
                                else:
                                    # Intentar ordenar por 'Fecha' (string) convertida a fecha si Fecha_obj falta
                                    if "Fecha" in df_display.columns:
                                        try:
                                            df_display["__tmp_fecha"] = pd.to_datetime(df_display["Fecha"], format="%d-%m-%Y", errors="coerce")
                                            df_display = df_display.sort_values("__tmp_fecha", ascending=False).reset_index(drop=True)
                                            df_display.drop(columns="__tmp_fecha", inplace=True)
                                        except Exception:
                                            # no hay fecha válida para ordenar, dejar tal cual
                                            pass
        
                                # Finalmente mostrar sólo las columnas disponibles (si available_cols quedó vacía, mostrar todo el df_display)
                                if available_cols:
                                    st.dataframe(df_display[available_cols].reset_index(drop=True), use_container_width=True)
                                else:
                                    st.dataframe(df_display.reset_index(drop=True), use_container_width=True)
        
                    # ------------------------------------------------------
                    # Rendimiento de Técnicos (Mi Equipo) - vista detallada
                    # ------------------------------------------------------
                    else:
                                        st.markdown(f"#### 👷 Rendimiento de Plantilla asignada a: {nombre_actual}")
                        
                                        # Cargar actividades técnicos (robusto)
                                        try:
                                            df_actividades = cargar_datos_tecnicos()
                                        except Exception as e:
                                            st.error("Error cargando actividades de técnicos: " + str(e))
                                            df_actividades = pd.DataFrame()
                        
                                        if df_actividades is None or df_actividades.empty:
                                            st.warning("No se encontraron actividades de técnicos (la base puede estar vacía).")
                                        else:
                                            df_actividades = df_actividades.copy()
                        
                                            # Normalizar Tiempo_Minutos
                                            if "Tiempo_Minutos" not in df_actividades.columns:
                                                if "Duración (min)" in df_actividades.columns:
                                                    df_actividades["Tiempo_Minutos"] = pd.to_numeric(df_actividades["Duración (min)"], errors="coerce").fillna(0)
                                                elif "Duración_min" in df_actividades.columns:
                                                    df_actividades["Tiempo_Minutos"] = pd.to_numeric(df_actividades["Duración_min"], errors="coerce").fillna(0)
                                                else:
                                                    if "Inicio" in df_actividades.columns and "Fin" in df_actividades.columns:
                                                        df_actividades["Inicio_dt"] = pd.to_datetime(df_actividades["Inicio"], errors="coerce")
                                                        df_actividades["Fin_dt"] = pd.to_datetime(df_actividades["Fin"], errors="coerce")
                                                        df_actividades["Tiempo_Minutos"] = ((df_actividades["Fin_dt"] - df_actividades["Inicio_dt"]).dt.total_seconds() / 60).fillna(0)
                                                    else:
                                                        df_actividades["Tiempo_Minutos"] = 0
                        
                                            # Fecha
                                            if "Fecha" not in df_actividades.columns:
                                                if "Inicio" in df_actividades.columns:
                                                    df_actividades["Fecha"] = pd.to_datetime(df_actividades["Inicio"], errors="coerce").dt.strftime("%d-%m-%Y")
                                                else:
                                                    df_actividades["Fecha"] = pd.NA
                                            df_actividades["Fecha_obj"] = pd.to_datetime(df_actividades["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                        
                                            # Detectar columnas jerárquicas
                                            cols_lower = [c.lower() for c in df_actividades.columns]
                                            col_unidad = next((c for c in df_actividades.columns if c.lower() == "unidad"), None)
                                            col_linea = next((c for c in df_actividades.columns if c.lower() in ("línea", "linea")), None)
                                            col_maquina = next((c for c in df_actividades.columns if c.lower() in ("máquina", "maquina")), None)
                        
                                            # Opciones de vista
                                            opciones_filtro = ["Unidad", "Línea", "Máquina", "Todas"]
                                            filtro_vista = st.selectbox("Ver inversión de tiempo por:", opciones_filtro, index=3)
                        
                                            # Filtro fechas (por defecto Hoy)
                                            modo_fecha_act = st.selectbox(
                                                "Selecciona rango de fechas:",
                                                ["Hoy", "Semana actual", "Mes actual", "Fecha específica", "Seleccionar rango"],
                                                index=0
                                            )
                                            hoy = datetime.date.today()
                                            if modo_fecha_act == "Hoy":
                                                fecha_inicio = fecha_fin = hoy
                                            elif modo_fecha_act == "Semana actual":
                                                fecha_fin = hoy
                                                fecha_inicio = fecha_fin - datetime.timedelta(days=fecha_fin.weekday())
                                            elif modo_fecha_act == "Mes actual":
                                                fecha_fin = hoy
                                                fecha_inicio = fecha_fin.replace(day=1)
                                            elif modo_fecha_act == "Fecha específica":
                                                fecha_inicio = fecha_fin = st.date_input("Selecciona la fecha", hoy, key="act_fecha_especifica")
                                            else:
                                                cc1, cc2 = st.columns(2)
                                                with cc1:
                                                    fecha_inicio = st.date_input("Fecha inicio", hoy - datetime.timedelta(days=7), key="act_fecha_ini")
                                                with cc2:
                                                    fecha_fin = st.date_input("Fecha fin", hoy, key="act_fecha_fin")
                        
                                            # Filtrar por supervisor y rango
                                            # ✅ CORRECCIÓN: "zona" en lugar de "area"
                                            zona_supervisor = st.session_state.get("zona", st.session_state.get("area", ""))
                                            filtro_base = (
                                                (df_actividades["Fecha_obj"].notna()) &
                                                (df_actividades["Fecha_obj"] >= fecha_inicio) &
                                                (df_actividades["Fecha_obj"] <= fecha_fin)
                                            )
                                            # Filtrar por zona del supervisor si el campo existe
                                            if "zona" in df_actividades.columns and zona_supervisor:
                                                filtro_base = filtro_base & (df_actividades["zona"] == zona_supervisor)
                                            elif "Supervisor" in df_actividades.columns:
                                                filtro_base = filtro_base & (df_actividades["Supervisor"] == nombre_actual)
                        
                                            df_filtrado = df_actividades[filtro_base].copy()
                        
                                            if df_filtrado.empty:
                                                st.info("No hay actividades registradas para el rango y filtros seleccionados.")
                                            else:
                                                # Nombre y actividad columns
                                                nombre_col = "Nombre" if "Nombre" in df_filtrado.columns else (df_filtrado.columns[0] if len(df_filtrado.columns) else "Usuario")
                                                actividad_col = "Actividad" if "Actividad" in df_filtrado.columns else None
                                                df_filtrado[nombre_col] = df_filtrado[nombre_col].astype(str)
                        
                                                # Métricas generales
                                                active_tecnicos = df_filtrado[nombre_col].nunique()
                                                total_activities = df_filtrado.shape[0]
                                                total_minutes = pd.to_numeric(df_filtrado["Tiempo_Minutos"], errors="coerce").fillna(0).sum()
                                                total_hours = total_minutes / 60
                        
                                                m1, m2, m3, m4 = st.columns(4)
                                                m1.metric("Técnicos activos", f"{active_tecnicos}")
                                                m2.metric("Total actividades", f"{total_activities}")
                                                m3.metric("Minutos invertidos", f"{total_minutes:.0f} min")
                                                m4.metric("Horas invertidas", f"{total_hours:.2f} hrs")
                        
                                                # Si selecciona "Todas" mostrar todas las gráficas (Unidad, Línea, Máquina) en el mismo panel
                                                if filtro_vista == "Todas":
                                                    st.markdown("### Todas las gráficas (Unidad / Línea / Máquina) — cada barra con color distinto")
                        
                                                    # 1) Por Unidad (si existe)
                                                    if col_unidad:
                                                        df_u = df_filtrado.groupby(col_unidad)["Tiempo_Minutos"].sum().reset_index().sort_values("Tiempo_Minutos", ascending=False)
                                                        if not df_u.empty:
                                                            chart_u = alt.Chart(df_u).mark_bar().encode(
                                                                x=alt.X("Tiempo_Minutos:Q", title="Minutos"),
                                                                y=alt.Y(f"{col_unidad}:N", sort='-x', title="Unidad"),
                                                                color=alt.Color(f"{col_unidad}:N", scale=alt.Scale(scheme=default_palette), legend=None),
                                                                tooltip=[alt.Tooltip(f"{col_unidad}:N", title="Unidad"), alt.Tooltip("Tiempo_Minutos:Q", title="Minutos")]
                                                            ).properties(height=300)
                                                            st.altair_chart(chart_u, use_container_width=True)
                        
                                                    # 2) Por Línea (si existe)
                                                    if col_linea:
                                                        df_l = df_filtrado.groupby(col_linea)["Tiempo_Minutos"].sum().reset_index().sort_values("Tiempo_Minutos", ascending=False)
                                                        if not df_l.empty:
                                                            chart_l = alt.Chart(df_l).mark_bar().encode(
                                                                x=alt.X("Tiempo_Minutos:Q", title="Minutos"),
                                                                y=alt.Y(f"{col_linea}:N", sort='-x', title="Línea"),
                                                                color=alt.Color(f"{col_linea}:N", scale=alt.Scale(scheme=default_palette), legend=None),
                                                                tooltip=[alt.Tooltip(f"{col_linea}:N", title="Línea"), alt.Tooltip("Tiempo_Minutos:Q", title="Minutos")]
                                                            ).properties(height=300)
                                                            st.altair_chart(chart_l, use_container_width=True)
                        
                                                    # 3) Por Máquina (si existe)
                                                    if col_maquina:
                                                        df_m = df_filtrado.groupby(col_maquina)["Tiempo_Minutos"].sum().reset_index().sort_values("Tiempo_Minutos", ascending=False)
                                                        if not df_m.empty:
                                                            chart_m = alt.Chart(df_m).mark_bar().encode(
                                                                x=alt.X("Tiempo_Minutos:Q", title="Minutos"),
                                                                y=alt.Y(f"{col_maquina}:N", sort='-x', title="Máquina"),
                                                                color=alt.Color(f"{col_maquina}:N", scale=alt.Scale(scheme=default_palette), legend=None),
                                                                tooltip=[alt.Tooltip(f"{col_maquina}:N", title="Máquina"), alt.Tooltip("Tiempo_Minutos:Q", title="Minutos")]
                                                            ).properties(height=300)
                                                            st.altair_chart(chart_m, use_container_width=True)
                        
                                                    # Si ninguna dimensión existe, mostrar total general
                                                    if not (col_unidad or col_linea or col_maquina):
                                                        total_minutos = df_filtrado["Tiempo_Minutos"].sum()
                                                        st.metric("Minutos totales invertidos", f"{total_minutos:.1f} min")
                        
                                                else:
                                                    # Comportamiento para Unidad / Línea / Máquina individual
                                                    if filtro_vista == "Unidad" and col_unidad:
                                                        grupo_col = col_unidad
                                                    elif filtro_vista == "Línea" and col_linea:
                                                        grupo_col = col_linea
                                                    elif filtro_vista == "Máquina" and col_maquina:
                                                        grupo_col = col_maquina
                                                    else:
                                                        grupo_col = None
                        
                                                    if grupo_col:
                                                        df_agr = df_filtrado.groupby(grupo_col)["Tiempo_Minutos"].sum().reset_index().sort_values("Tiempo_Minutos", ascending=False)
                                                        df_agr["Tiempo_Minutos"] = pd.to_numeric(df_agr["Tiempo_Minutos"], errors="coerce").fillna(0)
                                                        chart = alt.Chart(df_agr.head(50)).mark_bar().encode(
                                                            x=alt.X("Tiempo_Minutos:Q", title="Minutos invertidos"),
                                                            y=alt.Y(f"{grupo_col}:N", sort='-x', title=filtro_vista),
                                                            color=alt.Color(f"{grupo_col}:N", scale=alt.Scale(scheme=default_palette), legend=None),
                                                            tooltip=[alt.Tooltip(f"{grupo_col}:N", title=filtro_vista), alt.Tooltip("Tiempo_Minutos:Q", title="Minutos")]
                                                        ).properties(height=420)
                                                        st.altair_chart(chart, use_container_width=True)
                                                    else:
                                                        total_minutos = df_filtrado["Tiempo_Minutos"].sum()
                                                        st.metric("Minutos totales invertidos", f"{total_minutos:.1f} min")
                        
                                                # Gráfica individual combinada (Actividades vs Minutos) por técnico - top 6 en gráfico con colores diferenciados
                                                st.markdown("### Rendimiento Individual por Técnico")
                                                if actividad_col:
                                                    df_resumen = df_filtrado.groupby(nombre_col).agg(Total_Actividades=(actividad_col, "count"), Minutos_Invertidos=("Tiempo_Minutos", "sum")).reset_index()
                                                else:
                                                    df_resumen = df_filtrado.groupby(nombre_col).agg(Total_Actividades=(nombre_col, "count"), Minutos_Invertidos=("Tiempo_Minutos", "sum")).reset_index()
                                                df_resumen["Total_Actividades"] = pd.to_numeric(df_resumen["Total_Actividades"], errors="coerce").fillna(0)
                                                df_resumen["Minutos_Invertidos"] = pd.to_numeric(df_resumen["Minutos_Invertidos"], errors="coerce").fillna(0)
                        
                                                top_tecnicos = df_resumen.sort_values("Minutos_Invertidos", ascending=False).head(6)[nombre_col].tolist()
                                                df_res_top = df_resumen[df_resumen[nombre_col].isin(top_tecnicos)].copy()
                        
                                                if not df_res_top.empty:
                                                    melt = df_res_top.melt(id_vars=[nombre_col], value_vars=["Total_Actividades", "Minutos_Invertidos"], var_name="Metrica", value_name="Valor")
                                                    color_scale = alt.Scale(domain=["Total_Actividades", "Minutos_Invertidos"], range=["#1f77b4", "#ff7f0e"])
                                                    chart_grouped = alt.Chart(melt).mark_bar().encode(
                                                        x=alt.X("Metrica:N", title="Métrica"),
                                                        y=alt.Y("Valor:Q", title="Valor"),
                                                        column=alt.Column(f"{nombre_col}:N", header=alt.Header(labelAngle=-45, labelOrient="bottom")),
                                                        color=alt.Color("Metrica:N", scale=color_scale, legend=alt.Legend(title="Métrica")),
                                                        tooltip=[alt.Tooltip(f"{nombre_col}:N", title="Técnico"), alt.Tooltip("Metrica:N", title="Métrica"), alt.Tooltip("Valor:Q", title="Valor")]
                                                    ).properties(height=240)
                                                    st.altair_chart(chart_grouped, use_container_width=True)
                                                else:
                                                    st.info("No hay suficientes datos para mostrar gráfico individual.")
                        
                                                st.markdown("### Tabla resumen completa por técnico")
                                                df_resumen["Horas_Invertidas"] = (df_resumen["Minutos_Invertidos"] / 60).round(2)
                                                st.dataframe(df_resumen.sort_values("Minutos_Invertidos", ascending=False).reset_index(drop=True), use_container_width=True)
                        
    
        # TAB 4: DESCARGAS SUPERVISOR (Formato ejecutivo ampliado con resumen y título)
        # TAB 4: DESCARGAS SUPERVISOR (Formato ejecutivo ampliado + Descarga de actividades de la plantilla)
        with tab4:
            st.subheader("📂 Descargar reportes")
            df = cargar_datos_supervisores()
            if df is None or df.empty:
                st.info("⚠️ No hay registros disponibles en la base de datos.")
            else:
                modo_descarga = st.selectbox("Modo de descarga (registros de supervisión)", ["Día específico", "Seleccionar rango"])
    
                def to_excel_supervisor(df_to_save, supervisor_name, area_name, periodo_ini, periodo_fin):
                    """
                    Genera un Excel de registros de supervisión con formato ejecutivo.
                    """
                    output = BytesIO()
                    sheet_name_excel = "REPORTE_SUPERVISOR"  # <= 31 chars
                    writer = pd.ExcelWriter(output, engine="xlsxwriter")
                    startrow = 4
                    df_to_save.to_excel(writer, sheet_name=sheet_name_excel, index=False, startrow=startrow)
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name_excel]
    
                    # Formatos
                    title_format = workbook.add_format({
                        "bold": True,
                        "font_size": 14,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#1F4E78",
                        "font_color": "white",
                        "border": 1
                    })
                    sub_header_format = workbook.add_format({
                        "bold": True,
                        "font_size": 10,
                        "align": "left",
                        "valign": "vcenter",
                        "font_color": "#1F4E78"
                    })
                    header_format = workbook.add_format({
                        "bold": True,
                        "text_wrap": True,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#4F81BD",
                        "font_color": "white",
                        "border": 1
                    })
                    cell_format = workbook.add_format({
                        "align": "left",
                        "valign": "vcenter",
                        "border": 1
                    })
                    num_format = workbook.add_format({
                        "num_format": "#,##0.00",
                        "align": "right",
                        "valign": "vcenter",
                        "border": 1
                    })
    
                    # Título ejecutivo (fila 0) - merge across all data columns
                    worksheet.merge_range(0, 0, 0, len(df_to_save.columns) - 1, "PANEL SUPERVISOR - PUREM - PLANTA RAMOS ARIZPE", title_format)
                    worksheet.set_row(0, 28)
    
                    # Resumen ejecutivo (filas 1-3)
                    gen_fecha = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                    worksheet.write(1, 0, "Supervisor:", sub_header_format)
                    worksheet.write(1, 1, supervisor_name, cell_format)
                    worksheet.write(2, 0, "Área:", sub_header_format)
                    worksheet.write(2, 1, area_name or "", cell_format)
                    worksheet.write(1, 3, "Generado el:", sub_header_format)
                    worksheet.write(1, 4, gen_fecha, cell_format)
                    worksheet.write(2, 3, "Periodo:", sub_header_format)
                    worksheet.write(2, 4, f"{periodo_ini} a {periodo_fin}", cell_format)
    
                    # Totales y agregados calculados
                    total_registros = len(df_to_save)
                    if "Horas_Disponibles" in df_to_save.columns:
                        total_horas = pd.to_numeric(df_to_save["Horas_Disponibles"], errors="coerce").fillna(0).sum()
                    else:
                        total_horas = ""
                    worksheet.write(3, 0, "Total registros en reporte:", sub_header_format)
                    worksheet.write(3, 1, total_registros, cell_format)
                    worksheet.write(3, 3, "Total horas (si aplica):", sub_header_format)
                    worksheet.write(3, 4, total_horas, num_format if isinstance(total_horas, (int, float)) else cell_format)
    
                    # Encabezados (escribir con formato)
                    for col_num, value in enumerate(df_to_save.columns.values):
                        worksheet.write(startrow - 1, col_num, value, header_format)
                        # Ajustar ancho
                        max_len = min(max(df_to_save[value].astype(str).map(len).max(), len(value)) + 2, 50)
                        worksheet.set_column(col_num, col_num, max_len)
    
                    writer.close()
                    return output.getvalue()
    
                # --- Descarga de registros de supervisión (existente) ---
                nombre_supervisor = st.session_state.get("nombre", "")
                area_actual = st.session_state.get("area", "")
                df_mio_descarga = df[df["Supervisor"] == nombre_supervisor].copy()
    
                if df_mio_descarga.empty:
                    st.info("No tienes registros propios para descargar.")
                else:
                    df_mio_descarga["Fecha_obj"] = pd.to_datetime(df_mio_descarga["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
    
                    if modo_descarga == "Día específico":
                        fecha_excel = st.date_input("Selecciona la fecha (registros)", datetime.date.today(), key="fecha_excel_sup_final")
                        df_dia = df_mio_descarga[df_mio_descarga["Fecha_obj"] == fecha_excel]
    
                        if not df_dia.empty:
                            excel_data = to_excel_supervisor(df_dia.drop(columns=["Fecha_obj"], errors="ignore").fillna(""), nombre_supervisor, area_actual, fecha_excel, fecha_excel)
                            st.download_button(
                                label=f"📥 Descargar Excel de REGISTROS del {fecha_excel}",
                                data=excel_data,
                                file_name=f"Reporte_Supervisor_Registros_{nombre_supervisor}_{fecha_excel}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.info(f"⚠️ No hay registros para el día {fecha_excel}.")
                    else:
                        col1, col2 = st.columns(2)
                        with col1:
                            fecha_inicio = st.date_input("Fecha inicio (registros)", datetime.date.today() - datetime.timedelta(days=7), key="f_ini_desc_sup")
                        with col2:
                            fecha_fin = st.date_input("Fecha fin (registros)", datetime.date.today(), key="f_fin_desc_sup")
    
                        df_rango = df_mio_descarga[(df_mio_descarga["Fecha_obj"] >= fecha_inicio) & (df_mio_descarga["Fecha_obj"] <= fecha_fin)]
    
                        if not df_rango.empty:
                            excel_data = to_excel_supervisor(df_rango.drop(columns=["Fecha_obj"], errors="ignore").fillna(""), nombre_supervisor, area_actual, fecha_inicio, fecha_fin)
                            st.download_button(
                                label=f"📥 Descargar Excel REGISTROS Rango {fecha_inicio} a {fecha_fin}",
                                data=excel_data,
                                file_name=f"Reporte_Supervisor_Registros_{nombre_supervisor}_{fecha_inicio}_{fecha_fin}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.info(f"⚠️ No hay registros en el rango seleccionado.")
    
                st.markdown("---")
                st.markdown("### 📥 Descargar actividades registradas por mi plantilla")
    
                # Función para crear Excel de actividades (misma estética, con resumen de actividades)
                def to_excel_actividades(df_acts, supervisor_name, area_name, periodo_ini, periodo_fin):
                    """
                    Genera un Excel con todas las actividades registradas por la plantilla del supervisor,
                    con resumen ejecutivo y los datos crudos.
                    """
                    output = BytesIO()
                    sheet_name_excel = "ACTIVIDADES_SUP"  # <= 31 chars
                    writer = pd.ExcelWriter(output, engine="xlsxwriter")
                    startrow = 6
                    df_acts.to_excel(writer, sheet_name=sheet_name_excel, index=False, startrow=startrow)
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name_excel]
    
                    # Formatos
                    title_format = workbook.add_format({
                        "bold": True,
                        "font_size": 14,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#1F4E78",
                        "font_color": "white",
                        "border": 1
                    })
                    sub_header_format = workbook.add_format({
                        "bold": True,
                        "font_size": 10,
                        "align": "left",
                        "valign": "vcenter",
                        "font_color": "#1F4E78"
                    })
                    header_format = workbook.add_format({
                        "bold": True,
                        "text_wrap": True,
                        "align": "center",
                        "valign": "vcenter",
                        "fg_color": "#4F81BD",
                        "font_color": "white",
                        "border": 1
                    })
                    cell_format = workbook.add_format({
                        "align": "left",
                        "valign": "vcenter",
                        "border": 1
                    })
                    num_format = workbook.add_format({
                        "num_format": "#,##0.00",
                        "align": "right",
                        "valign": "vcenter",
                        "border": 1
                    })
    
                    # Título ejecutivo (fila 0) - merge across all data columns
                    worksheet.merge_range(0, 0, 0, len(df_acts.columns) - 1, "PANEL SUPERVISOR - PUREM - PLANTA RAMOS ARIZPE", title_format)
                    worksheet.set_row(0, 28)
    
                    # Resumen ejecutivo (filas 1-4)
                    gen_fecha = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
                    worksheet.write(1, 0, "Supervisor:", sub_header_format)
                    worksheet.write(1, 1, supervisor_name, cell_format)
                    worksheet.write(2, 0, "Área:", sub_header_format)
                    worksheet.write(2, 1, area_name or "", cell_format)
                    worksheet.write(1, 3, "Generado el:", sub_header_format)
                    worksheet.write(1, 4, gen_fecha, cell_format)
                    worksheet.write(2, 3, "Periodo:", sub_header_format)
                    worksheet.write(2, 4, f"{periodo_ini} a {periodo_fin}", cell_format)
    
                    # Agregados específicos de actividades
                    total_activities = len(df_acts)
                    # Asegurar columna de minutos
                    if "Tiempo_Minutos" not in df_acts.columns:
                        if "Duración (min)" in df_acts.columns:
                            df_acts["Tiempo_Minutos"] = pd.to_numeric(df_acts["Duración (min)"], errors="coerce").fillna(0)
                        elif "Duración_min" in df_acts.columns:
                            df_acts["Tiempo_Minutos"] = pd.to_numeric(df_acts["Duración_min"], errors="coerce").fillna(0)
                        else:
                            if "Inicio" in df_acts.columns and "Fin" in df_acts.columns:
                                df_acts["Inicio_dt"] = pd.to_datetime(df_acts["Inicio"], errors="coerce")
                                df_acts["Fin_dt"] = pd.to_datetime(df_acts["Fin"], errors="coerce")
                                df_acts["Tiempo_Minutos"] = ((df_acts["Fin_dt"] - df_acts["Inicio_dt"]).dt.total_seconds() / 60).fillna(0)
                            else:
                                df_acts["Tiempo_Minutos"] = 0
                    total_minutos = pd.to_numeric(df_acts["Tiempo_Minutos"], errors="coerce").fillna(0).sum()
                    total_horas = total_minutos / 60
    
                    worksheet.write(4, 0, "Total actividades:", sub_header_format)
                    worksheet.write(4, 1, total_activities, cell_format)
                    worksheet.write(4, 3, "Minutos totales:", sub_header_format)
                    worksheet.write(4, 4, total_minutos, num_format)
                    worksheet.write(5, 3, "Horas totales:", sub_header_format)
                    worksheet.write(5, 4, total_horas, num_format)
    
                    # Encabezados (escribir con formato)
                    for col_num, value in enumerate(df_acts.columns.values):
                        worksheet.write(startrow - 1, col_num, value, header_format)
                        # Ajustar ancho
                        max_len = min(max(df_acts[value].astype(str).map(len).max(), len(value)) + 2, 60)
                        worksheet.set_column(col_num, col_num, max_len)
    
                    writer.close()
                    return output.getvalue()
    
                # --- Descargas de actividades de la plantilla ---
                # Cargar actividades robustamente
                try:
                    df_actividades = cargar_datos_tecnicos()
                except Exception as e:
                    st.error("Error cargando actividades para descarga: " + str(e))
                    df_actividades = pd.DataFrame()
    
                if df_actividades is None or df_actividades.empty:
                    st.info("No hay actividades registradas por la plantilla.")
                else:
                    df_actividades = df_actividades.copy()
                    # Normalizar fecha y minutos
                    if "Fecha" not in df_actividades.columns and "Inicio" in df_actividades.columns:
                        df_actividades["Fecha"] = pd.to_datetime(df_actividades["Inicio"], errors="coerce").dt.strftime("%d-%m-%Y")
                    df_actividades["Fecha_obj"] = pd.to_datetime(df_actividades["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
    
                    if "Tiempo_Minutos" not in df_actividades.columns:
                        if "Duración (min)" in df_actividades.columns:
                            df_actividades["Tiempo_Minutos"] = pd.to_numeric(df_actividades["Duración (min)"], errors="coerce").fillna(0)
                        elif "Duración_min" in df_actividades.columns:
                            df_actividades["Tiempo_Minutos"] = pd.to_numeric(df_actividades["Duración_min"], errors="coerce").fillna(0)
                        else:
                            if "Inicio" in df_actividades.columns and "Fin" in df_actividades.columns:
                                df_actividades["Inicio_dt"] = pd.to_datetime(df_actividades["Inicio"], errors="coerce")
                                df_actividades["Fin_dt"] = pd.to_datetime(df_actividades["Fin"], errors="coerce")
                                df_actividades["Tiempo_Minutos"] = ((df_actividades["Fin_dt"] - df_actividades["Inicio_dt"]).dt.total_seconds() / 60).fillna(0)
                            else:
                                df_actividades["Tiempo_Minutos"] = 0
    
                    # Filtrar actividades del supervisor
                    df_acts_supervisor = df_actividades[df_actividades.get("Supervisor") == nombre_supervisor].copy()
                    if df_acts_supervisor.empty:
                        st.info("No hay actividades registradas por tu plantilla.")
                    else:
                        # Selector de modo de descarga para actividades
                        modo_desc_act = st.selectbox("Modo de descarga (actividades)", ["Día específico", "Seleccionar rango"], index=0)
    
                        if modo_desc_act == "Día específico":
                            fecha_act = st.date_input("Selecciona la fecha (actividades)", datetime.date.today(), key="fecha_acts_dia")
                            df_act_dia = df_acts_supervisor[df_acts_supervisor["Fecha_obj"] == fecha_act]
                            if not df_act_dia.empty:
                                excel_acts = to_excel_actividades(df_act_dia.drop(columns=["Fecha_obj"], errors="ignore").fillna(""), nombre_supervisor, area_actual, fecha_act, fecha_act)
                                st.download_button(
                                    label=f"📥 Descargar ACTIVIDADES del {fecha_act}",
                                    data=excel_acts,
                                    file_name=f"Actividades_Plantilla_{nombre_supervisor}_{fecha_act}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.info(f"No hay actividades para el día {fecha_act}.")
                        else:
                            c1, c2 = st.columns(2)
                            with c1:
                                f_ini_act = st.date_input("Fecha inicio (actividades)", datetime.date.today() - datetime.timedelta(days=7), key="f_ini_acts")
                            with c2:
                                f_fin_act = st.date_input("Fecha fin (actividades)", datetime.date.today(), key="f_fin_acts")
    
                            df_act_rango = df_acts_supervisor[(df_acts_supervisor["Fecha_obj"] >= f_ini_act) & (df_acts_supervisor["Fecha_obj"] <= f_fin_act)]
                            if not df_act_rango.empty:
                                excel_acts = to_excel_actividades(df_act_rango.drop(columns=["Fecha_obj"], errors="ignore").fillna(""), nombre_supervisor, area_actual, f_ini_act, f_fin_act)
                                st.download_button(
                                    label=f"📥 Descargar ACTIVIDADES Rango {f_ini_act} a {f_fin_act}",
                                    data=excel_acts,
                                    file_name=f"Actividades_Plantilla_{nombre_supervisor}_{f_ini_act}_{f_fin_act}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            else:
                                st.info("No hay actividades en el rango seleccionado.")
                                
##########################################################################################################################################################################################################################################

    # ==== PANEL ADMIN ====

    elif st.session_state["role"] == "admin":
        import pandas as pd
        import numpy as np
        import altair as alt
        import datetime
        from io import BytesIO
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
        import matplotlib.cm as cm
    
        # ------------------ FUNCIONES AUXILIARES ------------------
    
        def find_col(df, candidates):
            """Busca la primera columna de df cuyo nombre coincida (case-insensitive) con alguno de candidates."""
            if df is None or df.empty:
                return None
            cols = df.columns
            for cand in candidates:
                for c in cols:
                    if str(c).lower() == str(cand).lower():
                        return c
            return None
    
        def excel_executive(df_to_save, titulo_doc, periodo_text, filtros_info=None):
            """Genera un Excel con formato ejecutivo (bytes)."""
            if filtros_info is None:
                filtros_info = {}
            output = BytesIO()
            writer = pd.ExcelWriter(output, engine="xlsxwriter")
            # Escribir datos empezando en row 7 (dejar espacio al header)
            df_to_save.to_excel(writer, sheet_name="REPORTE", index=False, startrow=7)
            workbook = writer.book
            worksheet = writer.sheets["REPORTE"]
    
            # Formatos
            fmt_title = workbook.add_format({"bold": True, "font_size": 14, "align": "center", "bg_color": "#1F4E78", "font_color":"white"})
            fmt_meta_label = workbook.add_format({"bold": True, "align": "left"})
            fmt_meta_value = workbook.add_format({"align": "left"})
            fmt_header = workbook.add_format({"bold": True, "bg_color": "#4F81BD", "font_color":"white", "align":"center", "border":1})
    
            max_col = len(df_to_save.columns) - 1 if len(df_to_save.columns) > 0 else 0
            worksheet.merge_range(0, 0, 0, max_col, "PANEL SUPERVISOR - PUREM - PLANTA RAMOS ARIZPE", fmt_title)
    
            row_meta = 2
            worksheet.write(row_meta, 0, "Reporte:", fmt_meta_label)
            worksheet.write(row_meta, 1, titulo_doc, fmt_meta_value)
            worksheet.write(row_meta+1, 0, "Periodo:", fmt_meta_label)
            worksheet.write(row_meta+1, 1, periodo_text, fmt_meta_value)
    
            # Escribir filtros (si se suministraron)
            r = row_meta + 3
            for k, v in filtros_info.items():
                worksheet.write(r, 0, f"{k}:", fmt_meta_label)
                worksheet.write(r, 1, str(v), fmt_meta_value)
                r += 1
    
            # Encabezados del DataFrame
            for i, col in enumerate(df_to_save.columns):
                worksheet.write(7, i, col, fmt_header)
                worksheet.set_column(i, i, max(15, len(str(col)) + 2))
    
            writer.close()
            return output.getvalue()
    
        def generar_pdf_gap(df_resum, titulo, periodo, filtros):
            """Genera PDF ejecutivo con la gráfica gap y una tabla resumen (bytes)."""
            fig, ax = plt.subplots(2, 1, figsize=(11, 10), gridspec_kw={'height_ratios': [3, 1]})
            fig.suptitle(titulo, fontsize=16, fontweight='bold')
    
            grouping_col = df_resum.columns[0]
            x = np.arange(len(df_resum))
    
            # Barras reales delante (Horas_Invertidas stacked ya calculadas por grupo)
            ax[0].bar(x, df_resum["Horas_Invertidas"], color='#0b3d91', label='Horas Invertidas (Técnicos)')
            # Barras fantasma detrás (Horas_Disponibles)
            for xi, hd in zip(x, df_resum["Horas_Disponibles"]):
                ax[0].bar(xi, hd, color='lightgray', alpha=0.6, zorder=0)
    
            for i, g in enumerate(df_resum["Gap"]):
                ax[0].text(i, max(df_resum["Horas_Disponibles"].max()*0.01, df_resum["Horas_Invertidas"].iloc[i]) + df_resum["Horas_Disponibles"].max()*0.03,
                           f"Gap: {int(round(g))} hrs", color='red', fontweight='bold', ha='center')
    
            ax[0].set_xticks(x)
            ax[0].set_xticklabels(df_resum[grouping_col], rotation=45, ha='right')
            ax[0].set_ylabel("Horas")
            ax[0].legend()
    
            # Tabla resumen
            ax[1].axis('off')
            table_df = df_resum[[grouping_col, "Horas_Disponibles", "Horas_Invertidas", "Gap"]].copy()
            table_df = table_df.rename(columns={grouping_col: "Grupo", "Horas_Disponibles": "Horas Disponibles", "Horas_Invertidas": "Horas Invertidas", "Gap": "Gap (hrs)"})
            cell_text = []
            for _, row in table_df.iterrows():
                cell_text.append([str(row[c]) for c in table_df.columns])
            table = ax[1].table(cellText=cell_text, colLabels=table_df.columns, loc='center')
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1, 1.2)
    
            plt.subplots_adjust(hspace=0.6, top=0.88)
            filtros_text = " | ".join([f"{k}: {v}" for k, v in (filtros.items() if filtros else {}) if v and v != "Todos"])
            fig.text(0.01, 0.01, f"Periodo: {periodo}    Filtros: {filtros_text}", fontsize=9, color='gray')
    
            pdf_buf = BytesIO()
            with PdfPages(pdf_buf) as pdf:
                pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)
            return pdf_buf.getvalue()
    
        # ------------------ TABS ------------------
        tab1, tab2, tab3, tab4 = st.tabs(["📋 Historial", "📊 Rendimiento", "📂 Descargas", "👥 Usuarios activos"])
    
        # Inicializar estado del selector (vista)
        if "vista_admin" not in st.session_state:
            st.session_state.vista_admin = "tecnicos"
    
        # ---------------- TAB 1: HISTORIAL ADMIN ----------------
        with tab1:
                    st.subheader("📋 Historial de actividades")
                
                    # Selector dual elegante (uso width='stretch' para evitar advertencias)
                    col_sel1, col_sel2 = st.columns(2)
                    with col_sel1:
                        if st.button("🛠️ TÉCNICOS", width='stretch', type="primary" if st.session_state.vista_admin == "tecnicos" else "secondary"):
                            st.session_state.vista_admin = "tecnicos"
                            st.rerun()
                    with col_sel2:
                        if st.button("📋 SUPERVISORES", width='stretch', type="primary" if st.session_state.vista_admin == "supervisores" else "secondary"):
                            st.session_state.vista_admin = "supervisores"
                            st.rerun()
                
                    st.markdown("---")
                
                    # VISTA TÉCNICOS
                    if st.session_state.vista_admin == "tecnicos":
                        st.markdown("### 🛠️ Registros de Técnicos")
                        try:
                            df = cargar_datos_tecnicos()
                        except Exception as e:
                            st.error("Error cargando actividades de técnicos: " + str(e))
                            df = pd.DataFrame()
                
                        if df is None:
                            df = pd.DataFrame()
                
                        if not df.empty:
                            # --- CORRECCIÓN: Normalizar Fecha usando primero la columna 'Fecha' de Firestore ---
                            # Los registros en Firestore ya tienen 'Fecha' en formato "DD-MM-YYYY"
                            # Solo si no existe o está vacía, se extrae de 'Inicio'
                            if "Fecha" in df.columns:
                                # Intentar parsear con formato DD-MM-YYYY (formato guardado por el técnico)
                                df["Fecha"] = pd.to_datetime(df["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                                # Para los NaT que quedaron, intentar extraer de 'Inicio' como fallback
                                mask_nat = pd.Series([v is None or str(v) == 'NaT' for v in df["Fecha"]], index=df.index)
                                if mask_nat.any() and "Inicio" in df.columns:
                                    df.loc[mask_nat, "Fecha"] = pd.to_datetime(
                                        df.loc[mask_nat, "Inicio"].astype(str).str[:10],
                                        format="%d-%m-%Y", errors="coerce"
                                    ).dt.date
                            elif "Inicio" in df.columns:
                                # Fallback completo: extraer fecha de 'Inicio'
                                df["Inicio"] = pd.to_datetime(df["Inicio"], errors="coerce")
                                try:
                                    df["Inicio"] = df["Inicio"].dt.tz_convert("America/Mexico_City").dt.tz_localize(None)
                                except Exception:
                                    try:
                                        df["Inicio"] = df["Inicio"].dt.tz_localize(None)
                                    except Exception:
                                        pass
                                df["Fecha"] = df["Inicio"].dt.date
                            else:
                                df["Fecha"] = pd.NaT
                
                            modo_historial = st.selectbox(
                                "Modo de historial",
                                ["Día específico", "Seleccionar rango"],
                                key="modo_hist_tec"
                            )
                
                            # Inicializar variables por seguridad (evita NameError si se consultan más abajo)
                            fecha_seleccionada = None
                            fecha_inicio = None
                            fecha_fin = None
                            df_filtrado = pd.DataFrame()
                
                            if modo_historial == "Día específico":
                                fecha_seleccionada = st.date_input("Selecciona un día", datetime.date.today(), key="fecha_historial_tec_admin")
                                # Normalizar tipo si devuelve datetime
                                if isinstance(fecha_seleccionada, datetime.datetime):
                                    fecha_seleccionada = fecha_seleccionada.date()
                
                                # Comparación segura: convertir ambas a string ISO (YYYY-MM-DD)
                                df_filtrado = df[df["Fecha"].astype(str) == str(fecha_seleccionada)].copy()
                            else:
                                col1, col2 = st.columns(2)
                                with col1:
                                    fecha_inicio = st.date_input("Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio_historial_tec_admin")
                                with col2:
                                    fecha_fin = st.date_input("Fecha fin", datetime.date.today(), key="fecha_fin_historial_tec_admin")
                                # Asegurar que las fechas son datetime.date
                                if isinstance(fecha_inicio, datetime.datetime):
                                    fecha_inicio = fecha_inicio.date()
                                if isinstance(fecha_fin, datetime.datetime):
                                    fecha_fin = fecha_fin.date()
                
                                # Validación para evitar errores y fechas mal seleccionadas
                                if fecha_inicio is None or fecha_fin is None:
                                    st.error("Por favor selecciona ambas fechas para el rango.")
                                    df_filtrado = pd.DataFrame()
                                elif fecha_inicio > fecha_fin:
                                    st.error("La fecha de inicio no puede ser mayor que la fecha fin.")
                                    df_filtrado = pd.DataFrame()
                                else:
                                    # Filtrar usando la columna Fecha normalizada (que es datetime.date)
                                    df_filtrado = df[(df["Fecha"] >= fecha_inicio) & (df["Fecha"] <= fecha_fin)].copy()
                
                            # Mostrar resultados o mensajes según corresponda
                            if not df_filtrado.empty:
                                if modo_historial == "Día específico":
                                    st.markdown(f"#### Actividades del {fecha_seleccionada}")
                                else:
                                    st.markdown(f"#### Actividades del rango {fecha_inicio} a {fecha_fin}")
                
                                vista = st.selectbox(
                                    "📌 Selecciona modo de vista:",
                                    ["🔖 Ver por línea", "📑 Ver todo"],
                                    index=0,
                                    key="vista_tec_admin"
                                )
                
                                if vista == "🔖 Ver por línea":
                                    linea_col = find_col(df_filtrado, ["Línea", "Linea", "linea"])
                                    if linea_col is None:
                                        for idx, row in df_filtrado.sort_values("Inicio", ascending=False).iterrows():
                                            etiqueta_usuario = f"{row.get('Usuario','')} | {row.get('Nombre','')}"
                                            with st.expander(f"📝 {etiqueta_usuario} | {row.get('Actividad','')} | {row.get('Inicio','')}"):
                                                st.markdown(f"""
                                                <div class="registro-container">
                                                <h4 style="color:#00BFFF; margin-bottom:10px;">Actividad: {row.get('Actividad','')}</h4>
                                                <p><b>ID:</b> {row.get('ID','')}</p>
                                                <p><b>Usuario (MX):</b> {row.get('Usuario','')}</p>
                                                <p><b>Nombre:</b> {row.get('Nombre','')}</p>
                                                <p><b>Supervisor:</b> {row.get('Supervisor','No especificado')}</p>
                                                <p><b>Turno:</b> {row.get('Turno','No especificado')}</p>
                                                <p><b>Tipo:</b> {row.get('Tipo','')}</p>
                                                <p><b>Línea:</b> {row.get('Línea','No especificada')}</p>
                                                <p><b>Máquina:</b> {row.get('Máquina','')}</p>
                                                <p><b>Inicio:</b> {row.get('Inicio','')}</p>
                                                <p><b>Fin:</b> {row.get('Fin','')}</p>
                                                <p><b>Duración (min):</b> {row.get('Duración (min)','')}</p>
                                                </div>
                                                """, unsafe_allow_html=True)
                                                if st.button(f"Eliminar {row.get('ID','')}", key=f"del_tec_admin_{row.get('ID','')}_{idx}", width='stretch'):
                                                    eliminar_registro_tecnico(row.get("ID"))
                                    else:
                                        for linea, df_linea in df_filtrado.groupby(linea_col):
                                            nombre_linea = linea if pd.notna(linea) and str(linea).strip() != "" else "No especificada"
                                            with st.expander(f"📁 Línea: {nombre_linea}"):
                                                for idx, row in df_linea.sort_values("Inicio", ascending=False).iterrows():
                                                    etiqueta_usuario = f"{row.get('Usuario','')} | {row.get('Nombre','')}"
                                                    with st.expander(f"📝 {etiqueta_usuario} | {row.get('Actividad','')} | {row.get('Inicio','')}"):
                                                        st.markdown(f"""
                                                        <div class="registro-container">
                                                        <h4 style="color:#00BFFF; margin-bottom:10px;">Actividad: {row.get('Actividad','')}</h4>
                                                        <p><b>ID:</b> {row.get('ID','')}</p>
                                                        <p><b>Usuario (MX):</b> {row.get('Usuario','')}</p>
                                                        <p><b>Nombre:</b> {row.get('Nombre','')}</p>
                                                        <p><b>Supervisor:</b> {row.get('Supervisor','No especificado')}</p>
                                                        <p><b>Turno:</b> {row.get('Turno','No especificado')}</p>
                                                        <p><b>Tipo:</b> {row.get('Tipo','')}</p>
                                                        <p><b>Línea:</b> {row.get('Línea','No especificada')}</p>
                                                        <p><b>Máquina:</b> {row.get('Máquina','')}</p>
                                                        <p><b>Inicio:</b> {row.get('Inicio','')}</p>
                                                        <p><b>Fin:</b> {row.get('Fin','')}</p>
                                                        <p><b>Duración (min):</b> {row.get('Duración (min)','')}</p>
                                                        </div>
                                                        """, unsafe_allow_html=True)
                                                        if st.button(f"Eliminar {row.get('ID','')}", key=f"del_tec_admin_{row.get('ID','')}_{idx}", width='stretch'):
                                                            eliminar_registro_tecnico(row.get("ID"))
                                else:
                                    df_sorted = df_filtrado.sort_values(by="Inicio", ascending=False)
                                    for idx, row in df_sorted.iterrows():
                                        etiqueta_usuario = f"{row.get('Usuario','')} | {row.get('Nombre','')}"
                                        with st.expander(f"📝 {etiqueta_usuario} | {row.get('Actividad','')} | {row.get('Inicio','')}"):
                                            st.markdown(f"""
                                            <div class="registro-container">
                                            <h4 style="color:#00BFFF; margin-bottom:10px;">Actividad: {row.get('Actividad','')}</h4>
                                            <p><b>ID:</b> {row.get('ID','')}</p>
                                            <p><b>Usuario (MX):</b> {row.get('Usuario','')}</p>
                                            <p><b>Nombre:</b> {row.get('Nombre','')}</p>
                                            <p><b>Supervisor:</b> {row.get('Supervisor','No especificado')}</p>
                                            <p><b>Turno:</b> {row.get('Turno','No especificado')}</p>
                                            <p><b>Tipo:</b> {row.get('Tipo','')}</p>
                                            <p><b>Línea:</b> {row.get('Línea','No especificada')}</p>
                                            <p><b>Máquina:</b> {row.get('Máquina','')}</p>
                                            <p><b>Inicio:</b> {row.get('Inicio','')}</p>
                                            <p><b>Fin:</b> {row.get('Fin','')}</p>
                                            <p><b>Duración (min):</b> {row.get('Duración (min)','')}</p>
                                            </div>
                                            """, unsafe_allow_html=True)
                                            if st.button(f"Eliminar {row.get('ID','')}", key=f"del_tec_admin_all_{row.get('ID','')}_{idx}", width='stretch'):
                                                eliminar_registro_tecnico(row.get("ID"))
                            else:
                                if modo_historial == "Día específico":
                                    st.info(f"No hay actividades registradas para el día {fecha_seleccionada}.")
                                else:
                                    st.info(f"No hay actividades registradas entre {fecha_inicio} y {fecha_fin}.")
                        else:
                            st.info("No hay actividades registradas aún.")
                
                    # VISTA SUPERVISORES
                    else:
                        st.markdown("### 📋 Registros de Supervisores")
                        try:
                            df = cargar_datos_supervisores()
                        except Exception as e:
                            st.error("Error cargando registros de supervisores: " + str(e))
                            df = pd.DataFrame()
                
                        if df is None:
                            df = pd.DataFrame()
                
                        if not df.empty:
                            # Normalizar Fecha_obj (crear si no existe)
                            if "Fecha_obj" not in df.columns:
                                if "Fecha" in df.columns:
                                    try:
                                        df["Fecha_obj"] = pd.to_datetime(df["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                                        if df["Fecha_obj"].isna().all():
                                            df["Fecha_obj"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
                                    except Exception:
                                        df["Fecha_obj"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
                                elif "Inicio" in df.columns:
                                    df["Fecha_obj"] = pd.to_datetime(df["Inicio"], errors="coerce").dt.date
                                else:
                                    df["Fecha_obj"] = pd.NaT
                            else:
                                df["Fecha_obj"] = pd.to_datetime(df["Fecha_obj"], errors="coerce").dt.date
                
                            modo_historial = st.selectbox(
                                "Modo de historial",
                                ["Día específico", "Seleccionar rango"],
                                key="modo_hist_sup"
                            )
                
                            fecha_seleccionada = None
                            fecha_inicio = None
                            fecha_fin = None
                            df_filtrado = pd.DataFrame()
                
                            if modo_historial == "Día específico":
                                fecha_seleccionada = st.date_input("Selecciona un día", datetime.date.today(), key="fecha_historial_sup_admin")
                                if isinstance(fecha_seleccionada, datetime.datetime):
                                    fecha_seleccionada = fecha_seleccionada.date()
                                df_filtrado = df[df["Fecha_obj"] == fecha_seleccionada].copy()
                            else:
                                col1, col2 = st.columns(2)
                                with col1:
                                    fecha_inicio = st.date_input("Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio_historial_sup_admin")
                                with col2:
                                    fecha_fin = st.date_input("Fecha fin", datetime.date.today(), key="fecha_fin_historial_sup_admin")
                                if isinstance(fecha_inicio, datetime.datetime):
                                    fecha_inicio = fecha_inicio.date()
                                if isinstance(fecha_fin, datetime.datetime):
                                    fecha_fin = fecha_fin.date()
                
                                if fecha_inicio is None or fecha_fin is None:
                                    st.error("Por favor selecciona ambas fechas para el rango.")
                                    df_filtrado = pd.DataFrame()
                                elif fecha_inicio > fecha_fin:
                                    st.error("La fecha de inicio no puede ser mayor que la fecha fin.")
                                    df_filtrado = pd.DataFrame()
                                else:
                                    df_filtrado = df[(df["Fecha_obj"] >= fecha_inicio) & (df["Fecha_obj"] <= fecha_fin)].copy()
                
                            if not df_filtrado.empty:
                                if modo_historial == "Día específico":
                                    st.markdown(f"#### Registros del {fecha_seleccionada}")
                                else:
                                    st.markdown(f"#### Registros del rango {fecha_inicio} a {fecha_fin}")
                
                                sup_col = find_col(df_filtrado, ["Supervisor", "supervisor"])
                                if sup_col is None:
                                    st.dataframe(df_filtrado.reset_index(drop=True), use_container_width=True)
                                else:
                                    for supervisor, df_supervisor in df_filtrado.groupby(sup_col):
                                        with st.expander(f"📁 {supervisor}"):
                                            cols_show = [c for c in ["Fecha", "Turno", "Jornada_Normal", "Cursos_Otros", "Inasistencias", "Total_Real", "Horas_Disponibles", "Evento_Actividad"] if c in df_supervisor.columns]
                
                                            if "Fecha_obj" not in df_supervisor.columns:
                                                if "Fecha" in df_supervisor.columns:
                                                    df_supervisor["Fecha_obj"] = pd.to_datetime(df_supervisor["Fecha"], errors="coerce").dt.date
                                                elif "Inicio" in df_supervisor.columns:
                                                    df_supervisor["Fecha_obj"] = pd.to_datetime(df_supervisor["Inicio"], errors="coerce").dt.date
                                                else:
                                                    df_supervisor["Fecha_obj"] = pd.NaT
                
                                            if "Fecha_obj" in df_supervisor.columns and "Fecha_obj" not in cols_show:
                                                cols_show = ["Fecha_obj"] + cols_show
                
                                            if "Fecha_obj" in df_supervisor.columns:
                                                try:
                                                    st.dataframe(df_supervisor[cols_show].sort_values("Fecha_obj", ascending=False).reset_index(drop=True), use_container_width=True)
                                                except Exception:
                                                    st.dataframe(df_supervisor[cols_show].reset_index(drop=True), use_container_width=True)
                                            else:
                                                st.dataframe(df_supervisor[cols_show].reset_index(drop=True), use_container_width=True)
                            else:
                                if modo_historial == "Día específico":
                                    st.info(f"No hay registros para el día {fecha_seleccionada}.")
                                else:
                                    st.info(f"No hay registros entre {fecha_inicio} y {fecha_fin}.")
                        else:
                            st.info("No hay registros aún.")
    
        # ---------------- TAB 2: RENDIMIENTO ADMIN ----------------

        with tab2:
                                    st.subheader("📊 Gráficas de productividad (Admin)")
            
                                    # ---- CARGAR DATOS (una sola vez) ----
                                    try:
                                        df_sup_all = cargar_datos_supervisores_cached(
                                            st.session_state.usuario,
                                            st.session_state.role,
                                            st.session_state.area
                                        )
                                    except Exception as e:
                                        st.error("Error cargando registros de supervisores: " + str(e))
                                        df_sup_all = pd.DataFrame()
                                    if df_sup_all is None:
                                        df_sup_all = pd.DataFrame()
            
                                    try:
                                        df_act_all = cargar_datos_tecnicos_cached(
                                            st.session_state.usuario,
                                            st.session_state.role,
                                            st.session_state.area
                                        )
                                    except Exception as e:
                                        st.error("Error cargando actividades de técnicos: " + str(e))
                                        df_act_all = pd.DataFrame()
                                    if df_act_all is None:
                                        df_act_all = pd.DataFrame()
            
                                    # ---- Selector de vista ----
                                    col_sel1, col_sel2 = st.columns(2)
                                    with col_sel1:
                                        if st.button("🛠️ TÉCNICOS", width='stretch',
                                                     type="primary" if st.session_state.vista_admin == "tecnicos" else "secondary",
                                                     key="btn_rend_tec"):
                                            st.session_state.vista_admin = "tecnicos"
                                            st.rerun()
                                    with col_sel2:
                                        if st.button("📋 SUPERVISORES", width='stretch',
                                                     type="primary" if st.session_state.vista_admin == "supervisores" else "secondary",
                                                     key="btn_rend_sup"):
                                            st.session_state.vista_admin = "supervisores"
                                            st.rerun()
            
                                    st.markdown("---")
            
                                    # ---- HELPER: parsear fecha robustamente ----
                                    def _parse_fecha_col(series):
                                        parsed = pd.to_datetime(
                                            series.astype(str).str.strip(),
                                            format="%d-%m-%Y", errors="coerce"
                                        )
                                        mask_na = parsed.isna()
                                        if mask_na.any():
                                            parsed = parsed.copy()
                                            parsed[mask_na] = pd.to_datetime(
                                                series[mask_na].astype(str).str.strip(),
                                                errors="coerce"
                                            )
                                        return parsed.dt.date
            
                                    # ---- NORMALIZACIÓN SUPERVISORES ----
                                    if not df_sup_all.empty:
                                        df_sup_all = df_sup_all.copy()
                                        df_sup_all.columns = [c.strip() for c in df_sup_all.columns]
                                        if "Fecha" in df_sup_all.columns:
                                            df_sup_all["Fecha_obj"] = _parse_fecha_col(df_sup_all["Fecha"])
                                        elif "Inicio" in df_sup_all.columns:
                                            df_sup_all["Fecha_obj"] = pd.to_datetime(df_sup_all["Inicio"], errors="coerce").dt.date
                                        else:
                                            df_sup_all["Fecha_obj"] = None
                                        col_aliases_sup = {
                                            "Horas Disponibles": "Horas_Disponibles",
                                            "Total Esperado":    "Total_Esperado",
                                            "Jornada Normal":    "Jornada_Normal",
                                        }
                                        for col_orig, col_norm in col_aliases_sup.items():
                                            if col_orig in df_sup_all.columns:
                                                if col_norm not in df_sup_all.columns or df_sup_all[col_norm].eq(0).all():
                                                    df_sup_all[col_norm] = df_sup_all[col_orig]
                                        for c in ["Horas_Disponibles", "Total_Real", "Inasistencias", "Jornada_Normal", "Total_Esperado"]:
                                            if c in df_sup_all.columns:
                                                df_sup_all[c] = pd.to_numeric(df_sup_all[c], errors="coerce").fillna(0)
            
                                    # ---- NORMALIZACIÓN TÉCNICOS ----
                                    if not df_act_all.empty:
                                        df_act_all = df_act_all.copy()
                                        df_act_all.columns = [c.strip() for c in df_act_all.columns]
                                        if "Tiempo_Minutos" not in df_act_all.columns:
                                            if "Duración (min)" in df_act_all.columns:
                                                df_act_all["Tiempo_Minutos"] = pd.to_numeric(df_act_all["Duración (min)"], errors="coerce").fillna(0)
                                            elif "Duración_min" in df_act_all.columns:
                                                df_act_all["Tiempo_Minutos"] = pd.to_numeric(df_act_all["Duración_min"], errors="coerce").fillna(0)
                                            else:
                                                if "Inicio" in df_act_all.columns and "Fin" in df_act_all.columns:
                                                    df_act_all["Inicio_dt"] = pd.to_datetime(df_act_all["Inicio"], errors="coerce")
                                                    df_act_all["Fin_dt"]    = pd.to_datetime(df_act_all["Fin"],    errors="coerce")
                                                    df_act_all["Tiempo_Minutos"] = (
                                                        (df_act_all["Fin_dt"] - df_act_all["Inicio_dt"]).dt.total_seconds() / 60
                                                    ).fillna(0)
                                                else:
                                                    df_act_all["Tiempo_Minutos"] = 0
                                        if "Fecha" in df_act_all.columns:
                                            df_act_all["Fecha_obj"] = _parse_fecha_col(df_act_all["Fecha"])
                                        elif "Inicio" in df_act_all.columns:
                                            df_act_all["Fecha_obj"] = pd.to_datetime(df_act_all["Inicio"], errors="coerce").dt.date
                                        else:
                                            df_act_all["Fecha_obj"] = None
            
                                    # ---- FUNCIÓN: normaliza EPU/Facilidades ----
                                    def map_to_epu_short(val):
                                        if pd.isna(val) or str(val).strip() == "":
                                            return None
                                        s = str(val).upper().strip()
                                        if "EPU2" in s: return "EPU2"
                                        if "EPU3" in s: return "EPU3"
                                        if "EPU4" in s: return "EPU4"
                                        if "FACILIDAD" in s: return "FACILIDADES"
                                        return None
            
                                    # ---- DETECTAR COLUMNAS ----
                                    turno_col = find_col(df_act_all, ["Turno", "turno"]) or find_col(df_sup_all, ["Turno", "turno"])
                                    grouping_candidate = find_col(df_act_all, ["Unidad", "unidad", "EPU", "Epu"])
                                    if not grouping_candidate:
                                        grouping_candidate = find_col(df_act_all, ["Línea", "Linea", "linea"])
                                    tipo_col = find_col(df_act_all, ["Tipo", "Tipo_Mantenimiento", "Tipo de mantenimiento", "Mantenimiento", "tipo"])
                                    if tipo_col is None and not df_act_all.empty:
                                        df_act_all["Tipo"] = "Sin especificar"
                                        tipo_col = "Tipo"
                                    tecnico_col = find_col(df_act_all, ["Técnico", "Tecnico", "tecnico", "usuario", "Usuario", "Nombre", "nombre"])
            
                                    # ---- FILTROS DE FECHA Y TURNO ----
                                    turnos_vals = set()
                                    if turno_col:
                                        if not df_act_all.empty and turno_col in df_act_all.columns:
                                            turnos_vals.update(df_act_all[turno_col].dropna().unique().tolist())
                                        if not df_sup_all.empty and turno_col in df_sup_all.columns:
                                            turnos_vals.update(df_sup_all[turno_col].dropna().unique().tolist())
                                    turnos_list = ["Todos"] + sorted([v for v in turnos_vals if pd.notna(v)]) if turnos_vals else ["Todos"]
            
                                    modo_fecha = st.selectbox(
                                        "Rango de fechas (para las gráficas):",
                                        ["Hoy", "Semana actual", "Mes actual", "Fecha específica", "Seleccionar rango"],
                                        index=0
                                    )
                                    hoy = datetime.date.today()
                                    if modo_fecha == "Hoy":
                                        f_ini, f_fin = hoy, hoy
                                    elif modo_fecha == "Semana actual":
                                        f_fin = hoy
                                        f_ini = f_fin - datetime.timedelta(days=f_fin.weekday())
                                    elif modo_fecha == "Mes actual":
                                        f_fin = hoy
                                        f_ini = f_fin.replace(day=1)
                                    elif modo_fecha == "Fecha específica":
                                        f_ini = f_fin = st.date_input("Selecciona la fecha", hoy, key="admin_fecha_esp")
                                    else:
                                        c1, c2 = st.columns(2)
                                        with c1:
                                            f_ini = st.date_input("Fecha inicio", hoy - datetime.timedelta(days=7), key="admin_fecha_ini")
                                        with c2:
                                            f_fin = st.date_input("Fecha fin", hoy, key="admin_fecha_fin")
            
                                    if isinstance(f_ini, datetime.datetime): f_ini = f_ini.date()
                                    if isinstance(f_fin, datetime.datetime): f_fin = f_fin.date()
            
                                    turno_sel = st.selectbox("Filtrar por Turno (GAP):", turnos_list, index=0)
            
                                    # ---- APLICAR FILTRO DE FECHA ----
                                    def _filtrar_por_fecha(df, f_ini, f_fin):
                                        if df.empty or "Fecha_obj" not in df.columns:
                                            return df.copy() if not df.empty else pd.DataFrame()
                                        fecha_series = pd.to_datetime(df["Fecha_obj"], errors="coerce")
                                        f_ini_ts = pd.Timestamp(f_ini)
                                        f_fin_ts = pd.Timestamp(f_fin)
                                        mask = fecha_series.notna() & (fecha_series >= f_ini_ts) & (fecha_series <= f_fin_ts)
                                        return df[mask].copy()
            
                                    df_sup_date = _filtrar_por_fecha(df_sup_all, f_ini, f_fin)
                                    df_act_date = _filtrar_por_fecha(df_act_all, f_ini, f_fin)
            
                                    if turno_sel != "Todos" and turno_col:
                                        if turno_col in df_sup_date.columns:
                                            df_sup_date = df_sup_date[df_sup_date[turno_col] == turno_sel].copy()
                                        if turno_col in df_act_date.columns:
                                            df_act_date = df_act_date[df_act_date[turno_col] == turno_sel].copy()
            
                                    # ============================
                                    # === GRÁFICA GAP ===
                                    # ============================
                                    st.markdown("### 🔎 Gap de Horas Disponibles vs Horas Invertidas por EPU")
            
                                    if grouping_candidate is None:
                                        st.warning("⚠️ No se encontró columna 'Unidad'/'EPU' en datos de técnicos.")
                                    else:
                                        sup_group_col = find_col(df_sup_date, [
                                            "zona", "Zona", "ZONA",
                                            "area", "Area", "AREA",
                                            "Unidad", "unidad", "EPU", "Epu",
                                            "Línea", "Linea"
                                        ])
            
                                        if not df_sup_date.empty and sup_group_col is not None:
                                            df_sup_date = df_sup_date.copy()
                                            df_sup_date["_EPU_Sup"] = df_sup_date[sup_group_col].apply(map_to_epu_short)
                                            epus_con_supervisor = set(df_sup_date[df_sup_date["_EPU_Sup"].notna()]["_EPU_Sup"].unique())
                                            epus_sin_supervisor = []
                                            if not df_act_date.empty and grouping_candidate in df_act_date.columns:
                                                epus_sin_supervisor = [
                                                    e for e in df_act_date[grouping_candidate].apply(map_to_epu_short).dropna().unique()
                                                    if e is not None and e not in epus_con_supervisor
                                                ]
                                            filas_sin_epu = df_sup_date["_EPU_Sup"].isna()
                                            if filas_sin_epu.any() and len(epus_sin_supervisor) > 0:
                                                horas_residuales = pd.to_numeric(
                                                    df_sup_date.loc[filas_sin_epu, "Horas_Disponibles"], errors="coerce"
                                                ).fillna(0).sum()
                                                horas_por_epu = horas_residuales / len(epus_sin_supervisor)
                                                filas_extra = pd.DataFrame({
                                                    "_EPU_Sup": epus_sin_supervisor,
                                                    "Horas_Disponibles": [horas_por_epu] * len(epus_sin_supervisor)
                                                })
                                                df_sup_date = pd.concat([df_sup_date[~filas_sin_epu], filas_extra], ignore_index=True)
                                            elif filas_sin_epu.any():
                                                df_sup_date = df_sup_date[~filas_sin_epu].copy()
                                            sum_disp = (
                                                df_sup_date[df_sup_date["_EPU_Sup"].notna()]
                                                .groupby("_EPU_Sup")["Horas_Disponibles"]
                                                .sum().reset_index()
                                                .rename(columns={"_EPU_Sup": "Grupo"})
                                            )
                                        else:
                                            sum_disp = pd.DataFrame(columns=["Grupo", "Horas_Disponibles"])
                                            if df_sup_date.empty:
                                                st.warning("⚠️ Sin registros del supervisor para este periodo.")
                                            else:
                                                st.warning("⚠️ No se encontró columna de agrupación en datos del supervisor.")
            
                                        if not df_act_date.empty and grouping_candidate in df_act_date.columns:
                                            df_act_date = df_act_date.copy()
                                            df_act_date["_EPU_Tec"] = df_act_date[grouping_candidate].apply(map_to_epu_short)
                                            df_act_date["Tiempo_Minutos"] = pd.to_numeric(
                                                df_act_date["Tiempo_Minutos"] if "Tiempo_Minutos" in df_act_date.columns
                                                else df_act_date.get("Duración (min)", 0),
                                                errors="coerce"
                                            ).fillna(0)
                                            df_act_date["Horas"] = df_act_date["Tiempo_Minutos"] / 60.0
                                            df_act_valido = df_act_date[df_act_date["_EPU_Tec"].notna()].copy()
                                            sum_inv_by_type = (
                                                df_act_valido
                                                .groupby(["_EPU_Tec", tipo_col], dropna=False)["Horas"]
                                                .sum().reset_index()
                                                .rename(columns={"_EPU_Tec": "Grupo", tipo_col: "Tipo"})
                                            )
                                            sum_inv_total = (
                                                df_act_valido.groupby("_EPU_Tec")["Horas"]
                                                .sum().reset_index()
                                                .rename(columns={"_EPU_Tec": "Grupo", "Horas": "Horas_Invertidas"})
                                            )
                                        else:
                                            sum_inv_by_type = pd.DataFrame(columns=["Grupo", "Tipo", "Horas"])
                                            sum_inv_total   = pd.DataFrame(columns=["Grupo", "Horas_Invertidas"])
                                            df_act_valido   = pd.DataFrame()
            
                                        grupos = sorted(list(set(
                                            sum_disp["Grupo"].dropna().astype(str).tolist() +
                                            sum_inv_total["Grupo"].dropna().astype(str).tolist()
                                        )))
            
                                        if len(grupos) == 0:
                                            st.info("ℹ️ No hay datos en el periodo/turno seleccionado para generar el GAP por EPU.")
                                        else:
                                            df_disp_full = pd.DataFrame({"Grupo": grupos})
                                            sum_disp = pd.merge(df_disp_full, sum_disp, on="Grupo", how="left").fillna(0) \
                                                if not sum_disp.empty else df_disp_full.assign(Horas_Disponibles=0.0)
                                            sum_inv_total = pd.merge(df_disp_full, sum_inv_total, on="Grupo", how="left").fillna(0) \
                                                if not sum_inv_total.empty else df_disp_full.assign(Horas_Invertidas=0.0)
            
                                            resumen_merge = pd.merge(sum_disp, sum_inv_total, on="Grupo", how="outer").fillna(0)
                                            resumen_merge["Horas_Disponibles"] = pd.to_numeric(resumen_merge["Horas_Disponibles"], errors="coerce").fillna(0)
                                            resumen_merge["Horas_Invertidas"]  = pd.to_numeric(resumen_merge["Horas_Invertidas"],  errors="coerce").fillna(0)
                                            resumen_merge["Gap"] = (resumen_merge["Horas_Disponibles"] - resumen_merge["Horas_Invertidas"]).clip(lower=0)
                                            resumen_merge = resumen_merge.sort_values("Horas_Disponibles", ascending=False).reset_index(drop=True)
                                            resumen_merge["Label"] = resumen_merge["Gap"].apply(lambda v: f"Gap: {v:.1f} hrs")
            
                                            if sum_inv_by_type.empty:
                                                sum_inv_by_type = pd.DataFrame({
                                                    "Grupo": grupos,
                                                    "Tipo":  ["Sin actividad"] * len(grupos),
                                                    "Horas": [0.0] * len(grupos)
                                                })
            
                                            ymax = max(
                                                resumen_merge["Horas_Disponibles"].max(),
                                                resumen_merge["Horas_Invertidas"].max()
                                            ) * 1.15
                                            if pd.isna(ymax) or ymax <= 0:
                                                ymax = 1.0
            
                                            order = resumen_merge["Grupo"].astype(str).tolist()
            
                                            ghost = alt.Chart(resumen_merge).mark_bar(opacity=0.25, color="#9ea7ad").encode(
                                                x=alt.X("Grupo:N", sort=order, title="Grupo"),
                                                y=alt.Y("Horas_Disponibles:Q", title="Horas", scale=alt.Scale(domain=[0, ymax])),
                                                tooltip=[
                                                    alt.Tooltip("Grupo:N", title="Grupo"),
                                                    alt.Tooltip("Horas_Disponibles:Q", title="Horas Disponibles (Supervisor)", format=".2f")
                                                ]
                                            )
                                            stacked = alt.Chart(sum_inv_by_type).mark_bar().encode(
                                                x=alt.X("Grupo:N", sort=order),
                                                y=alt.Y("Horas:Q", stack="zero", scale=alt.Scale(domain=[0, ymax])),
                                                color=alt.Color("Tipo:N", title="Tipo Mantenimiento",
                                                                scale=alt.Scale(scheme="category10")),
                                                tooltip=[
                                                    alt.Tooltip("Grupo:N", title="Grupo"),
                                                    alt.Tooltip("Tipo:N", title="Tipo"),
                                                    alt.Tooltip("Horas:Q", title="Horas", format=".2f")
                                                ]
                                            )
                                            labels = alt.Chart(resumen_merge).mark_text(
                                                color="red", fontWeight="bold", dy=-10, fontSize=12
                                            ).encode(
                                                x=alt.X("Grupo:N", sort=order),
                                                y=alt.Y("Horas_Disponibles:Q"),
                                                text=alt.Text("Label:N")
                                            )
                                            gap_chart = alt.layer(ghost, stacked, labels).properties(
                                                height=420,
                                                title="Gap: Horas disponibles (supervisor, fondo) vs Horas invertidas por tipo (técnicos, frente)"
                                            )
                                            st.altair_chart(gap_chart, use_container_width=True)
            
                                            # ---- TABLA RESUMEN CON SEMÁFORO VISUAL ----
                                            st.markdown("#### 📄 Tabla de Horas por Grupo")
            
                                            def _semaforo_icon(pct):
                                                if pct >= 80:   return "🟢"
                                                elif pct >= 50: return "🟡"
                                                else:           return "🔴"
            
                                            resumen_display = resumen_merge[["Grupo", "Horas_Disponibles", "Horas_Invertidas", "Gap"]].copy()
                                            resumen_display["% Utilización"] = resumen_display.apply(
                                                lambda r: (r["Horas_Invertidas"] / r["Horas_Disponibles"] * 100)
                                                if r["Horas_Disponibles"] > 0 else 0.0, axis=1
                                            )
                                            resumen_display["Semáforo"] = resumen_display["% Utilización"].apply(_semaforo_icon)
                                            resumen_display["% Utilización"] = resumen_display["% Utilización"].apply(lambda v: f"{v:.1f}%")
                                            resumen_display = resumen_display.sort_values("Horas_Disponibles", ascending=False).reset_index(drop=True)
                                            st.dataframe(
                                                resumen_display[["Semáforo", "Grupo", "Horas_Disponibles", "Horas_Invertidas", "Gap", "% Utilización"]],
                                                use_container_width=True
                                            )
            
                                            periodo_txt  = f"{f_ini} a {f_fin}"
                                            filtros_info = {"Agrupado por": "EPU/Facilidades", "Turno": turno_sel, "Periodo": periodo_txt}
            
                                            # ================================================================
                                            # EXCEL EJECUTIVO MEJORADO (4 hojas)
                                            # ================================================================
                                            def _generar_excel_gap(resumen_df, actividades_df, tipo_col_name, tecnico_col_name, periodo, filtros):
                                                import io
                                                from openpyxl import Workbook
                                                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
                                                wb = Workbook()
            
                                                COLOR_HEADER   = "1F3864"
                                                COLOR_SUBHEAD  = "2E75B6"
                                                COLOR_EPU2     = "1F6B3A"
                                                COLOR_EPU3     = "7B3F00"
                                                COLOR_EPU4     = "1A3A6B"
                                                COLOR_FAC      = "6A0572"
                                                COLOR_DEFAULT  = "2E75B6"
                                                COLOR_GAP_OK   = "C6EFCE"
                                                COLOR_GAP_WARN = "FFEB9C"
                                                COLOR_GAP_BAD  = "FFC7CE"
                                                COLOR_ALT_ROW  = "EBF3FB"
                                                COLOR_SECTION  = "D6E4F0"
            
                                                EPU_COLORS = {
                                                    "EPU2": COLOR_EPU2,
                                                    "EPU3": COLOR_EPU3,
                                                    "EPU4": COLOR_EPU4,
                                                    "FACILIDADES": COLOR_FAC,
                                                }
            
                                                thin  = Side(style="thin",   color="BFBFBF")
                                                thick = Side(style="medium", color="1F3864")
                                                border_thin  = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
                                                border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
            
                                                def header_cell(ws, row, col, value, bg=COLOR_HEADER, font_size=11, bold=True, color="FFFFFF"):
                                                    c = ws.cell(row=row, column=col, value=value)
                                                    c.fill = PatternFill("solid", fgColor=bg)
                                                    c.font = Font(bold=bold, color=color, size=font_size)
                                                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                                    c.border = border_thin
                                                    return c
            
                                                def data_cell(ws, row, col, value, bg=None, bold=False, fmt=None, align="center", color="000000"):
                                                    c = ws.cell(row=row, column=col, value=value)
                                                    if bg:
                                                        c.fill = PatternFill("solid", fgColor=bg)
                                                    c.font = Font(bold=bold, size=10, color=color)
                                                    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
                                                    c.border = border_thin
                                                    if fmt:
                                                        c.number_format = fmt
                                                    return c
            
                                                def section_title(ws, row, col_start, col_end, value):
                                                    ws.merge_cells(
                                                        start_row=row, start_column=col_start,
                                                        end_row=row, end_column=col_end
                                                    )
                                                    c = ws.cell(row=row, column=col_start, value=value)
                                                    c.fill = PatternFill("solid", fgColor=COLOR_SECTION)
                                                    c.font = Font(bold=True, size=11, color=COLOR_HEADER)
                                                    c.alignment = Alignment(horizontal="left", vertical="center")
                                                    c.border = border_thin
                                                    ws.row_dimensions[row].height = 20
                                                    return c
            
                                                # ── HOJA 1: RESUMEN GAP ──────────────────────────────────
                                                ws1 = wb.active
                                                ws1.title = "Resumen GAP"
                                                ws1.sheet_view.showGridLines = False
            
                                                ws1.merge_cells("A1:F1")
                                                t = ws1["A1"]
                                                t.value = "REPORTE EJECUTIVO – GAP DE HORAS POR ÁREA"
                                                t.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                                                t.font = Font(bold=True, color="FFFFFF", size=15)
                                                t.alignment = Alignment(horizontal="center", vertical="center")
                                                ws1.row_dimensions[1].height = 36
            
                                                ws1.merge_cells("A2:F2")
                                                s = ws1["A2"]
                                                s.value = f"Periodo: {periodo}  |  Turno: {filtros.get('Turno','Todos')}  |  Generado: {datetime.date.today().strftime('%d/%m/%Y')}"
                                                s.fill = PatternFill("solid", fgColor=COLOR_SUBHEAD)
                                                s.font = Font(bold=False, color="FFFFFF", size=11, italic=True)
                                                s.alignment = Alignment(horizontal="center", vertical="center")
                                                ws1.row_dimensions[2].height = 22
                                                ws1.row_dimensions[3].height = 10
            
                                                headers = ["Área / EPU", "Horas Disponibles", "Horas Invertidas", "GAP (hrs)", "% Utilización", "Estado"]
                                                for ci, h in enumerate(headers, 1):
                                                    header_cell(ws1, 4, ci, h, bg=COLOR_SUBHEAD, font_size=10)
                                                ws1.row_dimensions[4].height = 24
            
                                                for ri, row in resumen_df.iterrows():
                                                    excel_row = ri + 5
                                                    grupo  = str(row["Grupo"])
                                                    h_disp = float(row["Horas_Disponibles"])
                                                    h_inv  = float(row["Horas_Invertidas"])
                                                    gap    = float(row["Gap"])
                                                    pct    = (h_inv / h_disp * 100) if h_disp > 0 else 0.0
                                                    epu_bg = EPU_COLORS.get(grupo.upper(), COLOR_DEFAULT)
                                                    row_bg = COLOR_ALT_ROW if ri % 2 == 0 else None
            
                                                    c = ws1.cell(row=excel_row, column=1, value=grupo)
                                                    c.fill = PatternFill("solid", fgColor=epu_bg)
                                                    c.font = Font(bold=True, color="FFFFFF", size=11)
                                                    c.alignment = Alignment(horizontal="center", vertical="center")
                                                    c.border = border_thin
            
                                                    data_cell(ws1, excel_row, 2, h_disp, bg=row_bg, fmt='0.00 "hrs"')
                                                    data_cell(ws1, excel_row, 3, h_inv,  bg=row_bg, fmt='0.00 "hrs"')
            
                                                
                                                    
                                                    if pct >= 80:
                                                        gap_color  = COLOR_GAP_OK;   estado = "✅ ÓPTIMO";   est_bg = "C6EFCE"; est_color = "375623"
                                                    elif pct >= 50:
                                                        gap_color  = COLOR_GAP_WARN; estado = "⚡ MODERADO"; est_bg = "FFEB9C"; est_color = "7F6000"
                                                    else:
                                                        gap_color  = COLOR_GAP_BAD;  estado = "⚠️ CRÍTICO";  est_bg = "FFC7CE"; est_color = "9C0006"
        
                                                    data_cell(ws1, excel_row, 4, gap, bg=gap_color, bold=True, fmt='0.00 "hrs"')
                                                    data_cell(ws1, excel_row, 5, pct / 100, bg=row_bg, fmt='0.0%')
        
                                                    c_est = ws1.cell(row=excel_row, column=6, value=estado)
                                                    c_est.fill = PatternFill("solid", fgColor=est_bg)
                                                    c_est.font = Font(bold=True, size=10, color=est_color)
                                                    c_est.alignment = Alignment(horizontal="center", vertical="center")
                                                    c_est.border = border_thin
                                                    ws1.row_dimensions[excel_row].height = 20
                                                    
            
                                                total_row  = len(resumen_df) + 5
                                                total_disp = resumen_df["Horas_Disponibles"].sum()
                                                total_inv  = resumen_df["Horas_Invertidas"].sum()
                                                total_gap  = resumen_df["Gap"].sum()
                                                total_pct  = (total_inv / total_disp * 100) if total_disp > 0 else 0.0
                                                ws1.row_dimensions[total_row].height = 22
                                                header_cell(ws1, total_row, 1, "TOTAL",                  bg=COLOR_HEADER, font_size=10)
                                                header_cell(ws1, total_row, 2, f"{total_disp:.2f} hrs",  bg=COLOR_HEADER, font_size=10)
                                                header_cell(ws1, total_row, 3, f"{total_inv:.2f} hrs",   bg=COLOR_HEADER, font_size=10)
                                                header_cell(ws1, total_row, 4, f"{total_gap:.1f} hrs",   bg=COLOR_HEADER, font_size=10)
                                                header_cell(ws1, total_row, 5, f"{total_pct:.1f}%",      bg=COLOR_HEADER, font_size=10)
                                                header_cell(ws1, total_row, 6, "",                        bg=COLOR_HEADER, font_size=10)
            
                                                legend_row = total_row + 2
                                                ws1.merge_cells(f"A{legend_row}:F{legend_row}")
                                                ws1[f"A{legend_row}"].value = "Semáforo:  ⚠️ CRÍTICO = < 50% utilizado  |  ⚡ MODERADO = 50–80%  |  ✅ ÓPTIMO = > 80%"
                                                ws1[f"A{legend_row}"].font = Font(italic=True, size=9, color="595959")
                                                ws1[f"A{legend_row}"].alignment = Alignment(horizontal="left")
                                                ws1.row_dimensions[legend_row].height = 16
            
                                                for col, w in zip(["A","B","C","D","E","F"], [18,20,20,16,16,16]):
                                                    ws1.column_dimensions[col].width = w
            
                                                # ── HOJA 2: ACTIVIDADES DETALLE ──────────────────────────
                                                ws2 = wb.create_sheet("Actividades Detalle")
                                                ws2.sheet_view.showGridLines = False
            
                                                ws2.merge_cells("A1:H1")
                                                t2 = ws2["A1"]
                                                t2.value = "DETALLE DE ACTIVIDADES REGISTRADAS POR TÉCNICOS"
                                                t2.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                                                t2.font = Font(bold=True, color="FFFFFF", size=13)
                                                t2.alignment = Alignment(horizontal="center", vertical="center")
                                                ws2.row_dimensions[1].height = 30
            
                                                ws2.merge_cells("A2:H2")
                                                s2 = ws2["A2"]
                                                s2.value = f"Periodo: {periodo}  |  Turno: {filtros.get('Turno','Todos')}"
                                                s2.fill = PatternFill("solid", fgColor=COLOR_SUBHEAD)
                                                s2.font = Font(color="FFFFFF", size=10, italic=True)
                                                s2.alignment = Alignment(horizontal="center")
                                                ws2.row_dimensions[2].height = 18
                                                ws2.row_dimensions[3].height = 10
            
                                                det_headers = ["Área / EPU", "Técnico", "Tipo de Mantenimiento",
                                                               "Actividad / Descripción", "Fecha", "Horas", "Turno", "Estado"]
                                                for ci, h in enumerate(det_headers, 1):
                                                    header_cell(ws2, 4, ci, h, bg=COLOR_SUBHEAD, font_size=10)
                                                ws2.row_dimensions[4].height = 24
            
                                                if not actividades_df.empty:
                                                    act_col      = find_col(actividades_df, ["Actividad", "actividad", "Descripcion",
                                                                                              "descripcion", "Descripción", "Tarea", "tarea"])
                                                    fecha_col_det = find_col(actividades_df, ["Fecha", "fecha"])
            
                                                    act_sorted = actividades_df.copy()
                                                    if "_EPU_Tec" in act_sorted.columns:
                                                        act_sorted = act_sorted.sort_values("_EPU_Tec", na_position="last").reset_index(drop=True)
            
                                                    prev_epu = None
                                                    current_excel_row = 5
            
                                                    for _, row2 in act_sorted.iterrows():
                                                        grupo2  = str(row2.get("_EPU_Tec", "")) if "_EPU_Tec" in act_sorted.columns else ""
                                                        tecnico = str(row2[tecnico_col_name]) if tecnico_col_name and tecnico_col_name in row2 else "—"
                                                        tipo    = str(row2[tipo_col_name])    if tipo_col_name    and tipo_col_name    in row2 else "—"
                                                        act_val = str(row2[act_col])          if act_col          and act_col          in row2 else "—"
                                                        fecha_v = str(row2[fecha_col_det])    if fecha_col_det    and fecha_col_det    in row2 else "—"
                                                        horas_v = float(row2.get("Horas", 0))
                                                        turno_v = str(row2[turno_col])        if turno_col        and turno_col        in row2 else "—"
                                                        epu_bg2 = EPU_COLORS.get(grupo2.upper(), COLOR_DEFAULT)
            
                                                        if grupo2 != prev_epu and grupo2 != "":
                                                            section_title(ws2, current_excel_row, 1, 8, f"  ▶  {grupo2}")
                                                            current_excel_row += 1
                                                            prev_epu = grupo2
            
                                                        alt_bg = COLOR_ALT_ROW if (current_excel_row % 2 == 0) else None
            
                                                        c2a = ws2.cell(row=current_excel_row, column=1, value=grupo2)
                                                        c2a.fill = PatternFill("solid", fgColor=epu_bg2)
                                                        c2a.font = Font(bold=True, color="FFFFFF", size=10)
                                                        c2a.alignment = Alignment(horizontal="center", vertical="center")
                                                        c2a.border = border_thin
            
                                                        data_cell(ws2, current_excel_row, 2, tecnico, bg=alt_bg, align="left")
                                                        data_cell(ws2, current_excel_row, 3, tipo,    bg=alt_bg, align="left")
                                                        data_cell(ws2, current_excel_row, 4, act_val, bg=alt_bg, align="left")
                                                        data_cell(ws2, current_excel_row, 5, fecha_v, bg=alt_bg)
                                                        data_cell(ws2, current_excel_row, 6, horas_v, bg=alt_bg, fmt='0.00 "hrs"')
                                                        data_cell(ws2, current_excel_row, 7, turno_v, bg=alt_bg)
            
                                                        if horas_v >= 8:
                                                            est2 = "✅ Completa"; est_bg2 = COLOR_GAP_OK
                                                        elif horas_v >= 4:
                                                            est2 = "⚡ Parcial";  est_bg2 = COLOR_GAP_WARN
                                                        else:
                                                            est2 = "⚠️ Breve";   est_bg2 = COLOR_GAP_BAD
                                                        data_cell(ws2, current_excel_row, 8, est2, bg=est_bg2, bold=True)
                                                        ws2.row_dimensions[current_excel_row].height = 18
                                                        current_excel_row += 1
            
                                                    current_excel_row += 1
                                                    section_title(ws2, current_excel_row, 1, 8, "  📊  SUBTOTALES POR ÁREA")
                                                    current_excel_row += 1
            
                                                    if "_EPU_Tec" in act_sorted.columns:
                                                        subtotales = (
                                                            act_sorted[act_sorted["_EPU_Tec"].notna()]
                                                            .groupby("_EPU_Tec")
                                                            .agg(Actividades=("Horas", "count"), Total_Horas=("Horas", "sum"))
                                                            .reset_index()
                                                            .sort_values("Total_Horas", ascending=False)
                                                        )
                                                        for _, st_row in subtotales.iterrows():
                                                            epu_st    = str(st_row["_EPU_Tec"])
                                                            epu_bg_st = EPU_COLORS.get(epu_st.upper(), COLOR_DEFAULT)
                                                            c_st = ws2.cell(row=current_excel_row, column=1, value=epu_st)
                                                            c_st.fill = PatternFill("solid", fgColor=epu_bg_st)
                                                            c_st.font = Font(bold=True, color="FFFFFF", size=10)
                                                            c_st.alignment = Alignment(horizontal="center", vertical="center")
                                                            c_st.border = border_thin
                                                            ws2.merge_cells(
                                                                start_row=current_excel_row, start_column=2,
                                                                end_row=current_excel_row, end_column=5
                                                            )
                                                            data_cell(ws2, current_excel_row, 2,
                                                                      f"{int(st_row['Actividades'])} actividades registradas",
                                                                      bg=COLOR_ALT_ROW, bold=True, align="left")
                                                            data_cell(ws2, current_excel_row, 6,
                                                                      float(st_row["Total_Horas"]),
                                                                      bg=COLOR_ALT_ROW, bold=True, fmt='0.00 "hrs"')
                                                            ws2.row_dimensions[current_excel_row].height = 20
                                                            current_excel_row += 1
            
                                                for col, w in zip(["A","B","C","D","E","F","G","H"], [16,24,26,38,14,12,12,14]):
                                                    ws2.column_dimensions[col].width = w
            
                                                # ── HOJA 3: HORAS POR TIPO ───────────────────────────────
                                                ws3 = wb.create_sheet("Horas por Tipo")
                                                ws3.sheet_view.showGridLines = False
            
                                                ws3.merge_cells("A1:E1")
                                                t3 = ws3["A1"]
                                                t3.value = "HORAS INVERTIDAS POR TIPO DE MANTENIMIENTO Y ÁREA"
                                                t3.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                                                t3.font = Font(bold=True, color="FFFFFF", size=13)
                                                t3.alignment = Alignment(horizontal="center", vertical="center")
                                                ws3.row_dimensions[1].height = 30
                                                ws3.row_dimensions[2].height = 10
            
                                                tipo_headers = ["Área / EPU", "Tipo de Mantenimiento", "Cantidad Actividades", "Total Horas", "% del Total"]
                                                for ci, h in enumerate(tipo_headers, 1):
                                                    header_cell(ws3, 3, ci, h, bg=COLOR_SUBHEAD, font_size=10)
                                                ws3.row_dimensions[3].height = 24
            
                                                if not actividades_df.empty and tipo_col_name and tipo_col_name in actividades_df.columns:
                                                    epu_col_det = "_EPU_Tec" if "_EPU_Tec" in actividades_df.columns else None
                                                    if epu_col_det:
                                                        total_horas_global = actividades_df["Horas"].sum() if "Horas" in actividades_df.columns else 1
                                                        grp3 = (
                                                            actividades_df[actividades_df[epu_col_det].notna()]
                                                            .groupby([epu_col_det, tipo_col_name])
                                                            .agg(Cantidad=("Horas", "count"), Horas=("Horas", "sum"))
                                                            .reset_index()
                                                            .sort_values([epu_col_det, "Horas"], ascending=[True, False])
                                                        )
                                                        prev_epu3    = None
                                                        current_row3 = 4
                                                        for _, row3 in grp3.iterrows():
                                                            grupo3  = str(row3[epu_col_det])
                                                            tipo3   = str(row3[tipo_col_name])
                                                            cant3   = int(row3["Cantidad"])
                                                            hrs3    = float(row3["Horas"])
                                                            pct3    = (hrs3 / total_horas_global * 100) if total_horas_global > 0 else 0.0
                                                            epu_bg3 = EPU_COLORS.get(grupo3.upper(), COLOR_DEFAULT)
            
                                                            if grupo3 != prev_epu3:
                                                                section_title(ws3, current_row3, 1, 5, f"  ▶  {grupo3}")
                                                                current_row3 += 1
                                                                prev_epu3 = grupo3
            
                                                            alt_bg3 = COLOR_ALT_ROW if (current_row3 % 2 == 0) else None
                                                            c3a = ws3.cell(row=current_row3, column=1, value=grupo3)
                                                            c3a.fill = PatternFill("solid", fgColor=epu_bg3)
                                                            c3a.font = Font(bold=True, color="FFFFFF", size=10)
                                                            c3a.alignment = Alignment(horizontal="center", vertical="center")
                                                            c3a.border = border_thin
                                                            data_cell(ws3, current_row3, 2, tipo3,       bg=alt_bg3, align="left")
                                                            data_cell(ws3, current_row3, 3, cant3,       bg=alt_bg3)
                                                            data_cell(ws3, current_row3, 4, hrs3,        bg=alt_bg3, fmt='0.00 "hrs"')
                                                            data_cell(ws3, current_row3, 5, pct3 / 100, bg=alt_bg3, fmt='0.0%')
                                                            ws3.row_dimensions[current_row3].height = 18
                                                            current_row3 += 1
            
                                                for col, w in zip(["A","B","C","D","E"], [16,30,22,16,14]):
                                                    ws3.column_dimensions[col].width = w
            
                                                # ── HOJA 4: TÉCNICOS ─────────────────────────────────────
                                                ws4 = wb.create_sheet("Técnicos")
                                                ws4.sheet_view.showGridLines = False
            
                                                ws4.merge_cells("A1:F1")
                                                t4 = ws4["A1"]
                                                t4.value = "RESUMEN DE PRODUCTIVIDAD POR TÉCNICO"
                                                t4.fill = PatternFill("solid", fgColor=COLOR_HEADER)
                                                t4.font = Font(bold=True, color="FFFFFF", size=13)
                                                t4.alignment = Alignment(horizontal="center", vertical="center")
                                                ws4.row_dimensions[1].height = 30
                                                ws4.row_dimensions[2].height = 10
            
                                                tec_headers = ["Técnico", "Área / EPU", "Actividades", "Horas Totales", "Promedio hrs/act", "Tipos realizados"]
                                                for ci, h in enumerate(tec_headers, 1):
                                                    header_cell(ws4, 3, ci, h, bg=COLOR_SUBHEAD, font_size=10)
                                                ws4.row_dimensions[3].height = 24
            
                                                if not actividades_df.empty and tecnico_col_name and tecnico_col_name in actividades_df.columns:
                                                    epu_col_tec = "_EPU_Tec" if "_EPU_Tec" in actividades_df.columns else None
                                                    grp_tec = (
                                                        actividades_df.groupby(tecnico_col_name)
                                                        .agg(Actividades=("Horas", "count"), Horas_Totales=("Horas", "sum"))
                                                        .reset_index()
                                                        .sort_values("Horas_Totales", ascending=False)
                                                    )
                                                    if epu_col_tec:
                                                        epu_por_tec = (
                                                            actividades_df[actividades_df[epu_col_tec].notna()]
                                                            .groupby(tecnico_col_name)[epu_col_tec]
                                                            .agg(lambda x: x.value_counts().index[0] if len(x) > 0 else "—")
                                                            .reset_index()
                                                            .rename(columns={epu_col_tec: "EPU_Principal"})
                                                        )
                                                        grp_tec = pd.merge(grp_tec, epu_por_tec, on=tecnico_col_name, how="left")
                                                    else:
                                                        grp_tec["EPU_Principal"] = "—"
            
                                                    if tipo_col_name and tipo_col_name in actividades_df.columns:
                                                        tipos_por_tec = (
                                                            actividades_df.groupby(tecnico_col_name)[tipo_col_name]
                                                            .agg(lambda x: ", ".join(sorted(x.dropna().unique())))
                                                            .reset_index()
                                                            .rename(columns={tipo_col_name: "Tipos"})
                                                        )
                                                        grp_tec = pd.merge(grp_tec, tipos_por_tec, on=tecnico_col_name, how="left")
                                                    else:
                                                        grp_tec["Tipos"] = "—"
            
                                                    for ri4, row4 in grp_tec.iterrows():
                                                        er4       = ri4 + 4
                                                        tec_name  = str(row4[tecnico_col_name])
                                                        epu_name  = str(row4.get("EPU_Principal", "—"))
                                                        act_cnt   = int(row4["Actividades"])
                                                        hrs_tot   = float(row4["Horas_Totales"])
                                                        prom_hrs  = hrs_tot / act_cnt if act_cnt > 0 else 0.0
                                                        tipos_str = str(row4.get("Tipos", "—"))
                                                        epu_bg4   = EPU_COLORS.get(epu_name.upper(), COLOR_DEFAULT)
                                                        alt_bg4   = COLOR_ALT_ROW if (ri4 % 2 == 0) else None
            
                                                        data_cell(ws4, er4, 1, tec_name, bg=alt_bg4, bold=True, align="left")
                                                        c4b = ws4.cell(row=er4, column=2, value=epu_name)
                                                        c4b.fill = PatternFill("solid", fgColor=epu_bg4)
                                                        c4b.font = Font(bold=True, color="FFFFFF", size=10)
                                                        c4b.alignment = Alignment(horizontal="center", vertical="center")
                                                        c4b.border = border_thin
                                                        data_cell(ws4, er4, 3, act_cnt,   bg=alt_bg4)
                                                        data_cell(ws4, er4, 4, hrs_tot,   bg=alt_bg4, fmt='0.00 "hrs"')
                                                        data_cell(ws4, er4, 5, prom_hrs,  bg=alt_bg4, fmt='0.00 "hrs"')
                                                        data_cell(ws4, er4, 6, tipos_str, bg=alt_bg4, align="left")
                                                        ws4.row_dimensions[er4].height = 18
            
                                                for col, w in zip(["A","B","C","D","E","F"], [26,16,16,16,18,36]):
                                                    ws4.column_dimensions[col].width = w
            
                                                buf = io.BytesIO()
                                                wb.save(buf)
                                                buf.seek(0)
                                                return buf.read()
            
                                            # ================================================================
                                            # PDF EJECUTIVO MEJORADO
                                            # ================================================================
                                            def _generar_pdf_gap_ejecutivo(resumen_df, actividades_df, tipo_col_name,
                                                                            tecnico_col_name, titulo, periodo, filtros):
                                                from reportlab.lib.pagesizes import A4, landscape
                                                from reportlab.lib import colors
                                                from reportlab.lib.units import cm
                                                from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                                                                Paragraph, Spacer, PageBreak)
                                                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                                                from reportlab.lib.enums import TA_CENTER, TA_LEFT
                                                from reportlab.graphics.shapes import Drawing, Rect, String, Line
                                                import io as _io
            
                                                buf = _io.BytesIO()
                                                doc = SimpleDocTemplate(
                                                    buf, pagesize=landscape(A4),
                                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                                    topMargin=1.5*cm, bottomMargin=1.5*cm
                                                )
            
                                                C_DARK   = colors.HexColor("#1F3864")
                                                C_MED    = colors.HexColor("#2E75B6")
                                                C_LIGHT  = colors.HexColor("#EBF3FB")
                                                C_WHITE  = colors.white
                                                C_GREEN  = colors.HexColor("#C6EFCE")
                                                C_YELLOW = colors.HexColor("#FFEB9C")
                                                C_RED    = colors.HexColor("#FFC7CE")
            
                                                EPU_PDF_COLORS = {
                                                    "EPU2":        colors.HexColor("#1F6B3A"),
                                                    "EPU3":        colors.HexColor("#7B3F00"),
                                                    "EPU4":        colors.HexColor("#1A3A6B"),
                                                    "FACILIDADES": colors.HexColor("#6A0572"),
                                                }
            
                                                style_title = ParagraphStyle(
                                                    "title", fontSize=20, textColor=C_WHITE,
                                                    alignment=TA_CENTER, fontName="Helvetica-Bold"
                                                )
                                                style_sub = ParagraphStyle(
                                                    "sub", fontSize=10, textColor=C_WHITE,
                                                    alignment=TA_CENTER, fontName="Helvetica"
                                                )
                                                style_section = ParagraphStyle(
                                                    "section", fontSize=13, textColor=C_DARK,
                                                    fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6
                                                )
                                                style_metric_label = ParagraphStyle(
                                                    "ml", fontSize=8, textColor=colors.HexColor("#595959"),
                                                    fontName="Helvetica", alignment=TA_CENTER
                                                )
                                                style_metric_value = ParagraphStyle(
                                                    "mv", fontSize=20, textColor=C_DARK,
                                                    fontName="Helvetica-Bold", alignment=TA_CENTER
                                                )
                                                style_note = ParagraphStyle(
                                                    "note", fontSize=8, textColor=colors.HexColor("#595959"),
                                                    fontName="Helvetica-Oblique", alignment=TA_LEFT
                                                )
            
                                                story = []
            
                                                story.append(Table(
                                                    [[Paragraph(f"<b>{titulo.upper()}</b>", style_title)]],
                                                    colWidths=[26*cm],
                                                    style=[("BACKGROUND",(0,0),(-1,-1),C_DARK),
                                                           ("TOPPADDING",(0,0),(-1,-1),12),
                                                           ("BOTTOMPADDING",(0,0),(-1,-1),8)]
                                                ))
                                                story.append(Table(
                                                    [[Paragraph(
                                                        f"Periodo: {periodo}  |  Turno: {filtros.get('Turno','Todos')}  |  "
                                                        f"Generado: {datetime.date.today().strftime('%d/%m/%Y')}",
                                                        style_sub
                                                    )]],
                                                    colWidths=[26*cm],
                                                    style=[("BACKGROUND",(0,0),(-1,-1),C_MED),
                                                           ("TOPPADDING",(0,0),(-1,-1),5),
                                                           ("BOTTOMPADDING",(0,0),(-1,-1),5)]
                                                ))
                                                story.append(Spacer(1, 0.5*cm))
            
                                                story.append(Paragraph("📈 Métricas Ejecutivas del Periodo", style_section))
            
                                                total_disp = float(resumen_df["Horas_Disponibles"].sum())
                                                total_inv  = float(resumen_df["Horas_Invertidas"].sum())
                                                total_gap  = float(resumen_df["Gap"].sum())
                                                total_pct  = (total_inv / total_disp * 100) if total_disp > 0 else 0.0
                                                num_acts   = len(actividades_df) if not actividades_df.empty else 0
                                                num_tecs   = (actividades_df[tecnico_col_name].nunique()
                                                              if (not actividades_df.empty and tecnico_col_name
                                                                  and tecnico_col_name in actividades_df.columns) else 0)
            
                                                if total_pct >= 80:
                                                    pct_color = C_RED;    pct_label = "✅ ÓPTIMO"
                                                elif total_pct >= 50:
                                                    pct_color = C_YELLOW; pct_label = "⚡ MODERADO"
                                                else:
                                                    pct_color = C_GREEN;  pct_label = "⚠️ CRÍTICO"
            
                                                def _card(label, value, sub, bg):
                                                    return Table(
                                                        [[Paragraph(label, style_metric_label)],
                                                         [Paragraph(f"<b>{value}</b>", style_metric_value)],
                                                         [Paragraph(sub, style_metric_label)]],
                                                        colWidths=[4.1*cm],
                                                        style=[("BACKGROUND",(0,0),(-1,-1),bg),
                                                               ("BOX",(0,0),(-1,-1),1,C_MED),
                                                               ("TOPPADDING",(0,0),(-1,-1),6),
                                                               ("BOTTOMPADDING",(0,0),(-1,-1),6),
                                                               ("ALIGN",(0,0),(-1,-1),"CENTER")]
                                                    )
            
                                                metrics_row = [[
                                                    _card("Horas Disponibles", f"{total_disp:.1f}", "hrs totales",    C_LIGHT),
                                                    _card("Horas Invertidas",  f"{total_inv:.1f}",  "hrs trabajadas", C_LIGHT),
                                                    _card("GAP Total",         f"{total_gap:.1f}",  "hrs sin cubrir",
                                                          C_RED if total_gap > 0 else C_GREEN),
                                                    _card("% Utilización",     f"{total_pct:.1f}%", pct_label,        pct_color),
                                                    _card("Técnicos Activos",  str(num_tecs),       "en el periodo",  C_LIGHT),
                                                    _card("Actividades",       str(num_acts),       "registradas",    C_LIGHT),
                                                ]]
                                                story.append(Table(
                                                    metrics_row, colWidths=[4.33*cm]*6,
                                                    style=[("ALIGN",(0,0),(-1,-1),"CENTER"),
                                                           ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                                           ("LEFTPADDING",(0,0),(-1,-1),3),
                                                           ("RIGHTPADDING",(0,0),(-1,-1),3)]
                                                ))
                                                story.append(Spacer(1, 0.5*cm))
            
                                                # ── GRÁFICA DE BARRAS VISUAL ─────────────────────────────
                                                story.append(Paragraph("📊 Gráfica GAP por Área / EPU", style_section))
            
                                                if not resumen_df.empty:
                                                    bar_w    = 26 * cm
                                                    bar_h    = 7  * cm
                                                    margin_l = 2.5 * cm
                                                    margin_b = 1.2 * cm
                                                    plot_w   = bar_w - margin_l - 0.5*cm
                                                    plot_h   = bar_h - margin_b - 0.8*cm
            
                                                    max_val = max(
                                                        resumen_df["Horas_Disponibles"].max(),
                                                        resumen_df["Horas_Invertidas"].max()
                                                    ) * 1.15
                                                    if max_val <= 0:
                                                        max_val = 1.0
            
                                                    n_grupos  = len(resumen_df)
                                                    group_w   = plot_w / n_grupos
                                                    bar_width = group_w * 0.35
            
                                                    d = Drawing(bar_w, bar_h)
                                                    d.add(Rect(0, 0, bar_w, bar_h, fillColor=colors.white, strokeColor=None))
            
                                                    for i in range(6):
                                                        y_val = max_val * i / 5
                                                        y_px  = margin_b + (y_val / max_val) * plot_h
                                                        d.add(Line(margin_l, y_px, bar_w - 0.5*cm, y_px,
                                                                   strokeColor=colors.HexColor("#E0E0E0"), strokeWidth=0.5))
                                                        d.add(String(margin_l - 4, y_px - 3, f"{y_val:.0f}",
                                                                     fontSize=6, fillColor=colors.HexColor("#595959"),
                                                                     textAnchor="end"))
            
                                                    d.add(Line(margin_l, margin_b, bar_w - 0.5*cm, margin_b,
                                                               strokeColor=colors.HexColor("#BFBFBF"), strokeWidth=1))
            
                                                    EPU_DRAW_COLORS = {
                                                        "EPU2":        colors.HexColor("#1F6B3A"),
                                                        "EPU3":        colors.HexColor("#7B3F00"),
                                                        "EPU4":        colors.HexColor("#1A3A6B"),
                                                        "FACILIDADES": colors.HexColor("#6A0572"),
                                                    }
            
                                                    for idx, row_g in resumen_df.reset_index(drop=True).iterrows():
                                                        grupo_g  = str(row_g["Grupo"])
                                                        h_disp_g = float(row_g["Horas_Disponibles"])
                                                        h_inv_g  = float(row_g["Horas_Invertidas"])
                                                        gap_g    = float(row_g["Gap"])
                                                        pct_g    = (h_inv_g / h_disp_g * 100) if h_disp_g > 0 else 0.0
            
                                                        cx      = margin_l + (idx + 0.5) * group_w
                                                        epu_c_g = EPU_DRAW_COLORS.get(grupo_g.upper(), C_MED)
            
                                                        h_disp_px = (h_disp_g / max_val) * plot_h
                                                        d.add(Rect(cx - bar_width, margin_b, bar_width * 2, h_disp_px,
                                                                   fillColor=colors.HexColor("#9ea7ad"), strokeColor=None,
                                                                   fillOpacity=0.3))
            
                                                        h_inv_px = (h_inv_g / max_val) * plot_h
                                                        d.add(Rect(cx - bar_width * 0.7, margin_b, bar_width * 1.4, h_inv_px,
                                                                   fillColor=epu_c_g, strokeColor=None))
            
                                                        gap_color_draw = (
                                                            colors.HexColor("#C00000") if pct_g >= 80
                                                            else colors.HexColor("#7F6000") if pct_g >= 50
                                                            else colors.HexColor("#375623")
                                                        )
                                                        d.add(String(cx, margin_b + h_disp_px + 3,
                                                                     f"Gap:{gap_g:.1f}h",
                                                                     fontSize=6.5, fillColor=gap_color_draw,
                                                                     fontName="Helvetica-Bold", textAnchor="middle"))
                                                        d.add(String(cx, margin_b - 10, grupo_g,
                                                                     fontSize=7, fillColor=colors.HexColor("#1F3864"),
                                                                     fontName="Helvetica-Bold", textAnchor="middle"))
            
                                                    story.append(d)
                                                    story.append(Spacer(1, 0.2*cm))
            
                                                    legend_d = Drawing(26*cm, 1.2*cm)
                                                    x_leg = 0
                                                    legend_d.add(Rect(x_leg, 2, 12, 10,
                                                                      fillColor=colors.HexColor("#9ea7ad"), strokeColor=None))
                                                    legend_d.add(String(x_leg + 16, 3, "Horas Disponibles (supervisor)",
                                                                        fontSize=8, fillColor=colors.HexColor("#595959")))
                                                    x_leg += 5.5*cm
                                                    for epu_name_l, epu_c_l in EPU_DRAW_COLORS.items():
                                                        legend_d.add(Rect(x_leg, 2, 12, 10, fillColor=epu_c_l, strokeColor=None))
                                                        legend_d.add(String(x_leg + 16, 3, f"Horas Invertidas – {epu_name_l}",
                                                                            fontSize=8, fillColor=colors.HexColor("#595959")))
                                                        x_leg += 5.5*cm
                                                    story.append(legend_d)
            
                                                story.append(Spacer(1, 0.3*cm))
            
                                                # ── TABLA GAP ────────────────────────────────────────────
                                                story.append(Paragraph("📋 Resumen de Horas por Área", style_section))
            
                                                gap_headers_pdf = ["Área / EPU", "Horas Disponibles", "Horas Invertidas",
                                                                   "GAP (hrs)", "% Utilización", "Estado"]
                                                gap_rows_pdf  = [gap_headers_pdf]
                                                gap_style_pdf = [
                                                    ("BACKGROUND",    (0,0), (-1,0), C_MED),
                                                    ("TEXTCOLOR",     (0,0), (-1,0), C_WHITE),
                                                    ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                                                    ("FONTSIZE",      (0,0), (-1,-1), 9),
                                                    ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                                                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                                                    ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#BFBFBF")),
                                                    ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_LIGHT]),
                                                    ("TOPPADDING",    (0,0), (-1,-1), 5),
                                                    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                                                ]
            
                                                for ri_p, row_p in resumen_df.iterrows():
                                                    grupo_p  = str(row_p["Grupo"])
                                                    h_disp_p = float(row_p["Horas_Disponibles"])
                                                    h_inv_p  = float(row_p["Horas_Invertidas"])
                                                    gap_p    = float(row_p["Gap"])
                                                    pct_p    = (h_inv_p / h_disp_p * 100) if h_disp_p > 0 else 0.0
                                                    epu_c_p  = EPU_PDF_COLORS.get(grupo_p.upper(), C_MED)
            
                                                    if pct_p >= 80:
                                                        gap_c_p = C_RED;    estado_p = "✅ ÓPTIMO"
                                                    elif pct_p >= 50:
                                                        gap_c_p = C_YELLOW; estado_p = "⚡ MODERADO"
                                                    else:
                                                        gap_c_p = C_GREEN;  estado_p = "⚠️ CRÍTICO"
            
                                                    gap_rows_pdf.append([
                                                        Paragraph(f"<b><font color='white'>{grupo_p}</font></b>",
                                                                  ParagraphStyle("ep_p", alignment=TA_CENTER,
                                                                                 fontSize=9, fontName="Helvetica-Bold")),
                                                        f"{h_disp_p:.2f} hrs",
                                                        f"{h_inv_p:.2f} hrs",
                                                        f"{gap_p:.1f} hrs",
                                                        f"{pct_p:.1f}%",
                                                        estado_p
                                                    ])
                                                    dr_p = len(gap_rows_pdf) - 1
                                                    gap_style_pdf.append(("BACKGROUND", (0, dr_p), (0, dr_p), epu_c_p))
                                                    gap_style_pdf.append(("BACKGROUND", (3, dr_p), (4, dr_p), gap_c_p))
                                                    gap_style_pdf.append(("BACKGROUND", (5, dr_p), (5, dr_p), gap_c_p))
            
                                                gap_rows_pdf.append([
                                                    Paragraph("<b>TOTAL</b>",
                                                              ParagraphStyle("tot_p", alignment=TA_CENTER,
                                                                             fontSize=9, fontName="Helvetica-Bold",
                                                                             textColor=C_WHITE)),
                                                    f"{total_disp:.2f} hrs",
                                                    f"{total_inv:.2f} hrs",
                                                    f"{total_gap:.1f} hrs",
                                                    f"{total_pct:.1f}%",
                                                    pct_label
                                                ])
                                                tr_p = len(gap_rows_pdf) - 1
                                                gap_style_pdf += [
                                                    ("BACKGROUND", (0, tr_p), (-1, tr_p), C_DARK),
                                                    ("TEXTCOLOR",  (0, tr_p), (-1, tr_p), C_WHITE),
                                                    ("FONTNAME",   (0, tr_p), (-1, tr_p), "Helvetica-Bold"),
                                                ]
            
                                                story.append(Table(gap_rows_pdf,
                                                                   colWidths=[4*cm, 4.5*cm, 4.5*cm, 4*cm, 4*cm, 5*cm],
                                                                   style=TableStyle(gap_style_pdf)))
                                                story.append(Spacer(1, 0.3*cm))
                                                story.append(Paragraph(
                                                    "Semáforo: ⚠️ CRÍTICO = < 50% utilizado  |  ⚡ MODERADO = 50–80%  |  ✅ ÓPTIMO = > 80%",
                                                    style_note
                                                ))
            
                                                # ── DETALLE DE ACTIVIDADES ───────────────────────────────
                                                if not actividades_df.empty:
                                                    story.append(PageBreak())
                                                    story.append(Paragraph("📋 Detalle de Actividades por Técnico", style_section))
            
                                                    act_col_pdf   = find_col(actividades_df, ["Actividad", "actividad",
                                                                                               "Descripcion", "descripcion",
                                                                                               "Descripción", "Tarea"])
                                                    fecha_col_pdf = find_col(actividades_df, ["Fecha", "fecha"])
            
                                                    act_headers_pdf = ["Área", "Técnico", "Tipo", "Actividad",
                                                                       "Fecha", "Horas", "Estado"]
                                                    act_rows_pdf    = [act_headers_pdf]
                                                    act_style_pdf   = [
                                                        ("BACKGROUND",    (0,0), (-1,0), C_MED),
                                                        ("TEXTCOLOR",     (0,0), (-1,0), C_WHITE),
                                                        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                                                        ("FONTSIZE",      (0,0), (-1,-1), 8),
                                                        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                                                        ("ALIGN",         (1,1), (3,-1), "LEFT"),
                                                        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                                                        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#BFBFBF")),
                                                        ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_LIGHT]),
                                                        ("TOPPADDING",    (0,0), (-1,-1), 4),
                                                        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                                                    ]
            
                                                    act_sorted_pdf = actividades_df.copy()
                                                    if "_EPU_Tec" in act_sorted_pdf.columns:
                                                        act_sorted_pdf = act_sorted_pdf.sort_values(
                                                            "_EPU_Tec", na_position="last"
                                                        ).reset_index(drop=True)
            
                                                    for ri_a, row_a in act_sorted_pdf.iterrows():
                                                        grupo_a = str(row_a.get("_EPU_Tec", "")) if "_EPU_Tec" in act_sorted_pdf.columns else ""
                                                        tec_a   = str(row_a[tecnico_col_name]) if tecnico_col_name and tecnico_col_name in row_a else "—"
                                                        tipo_a  = str(row_a[tipo_col_name])    if tipo_col_name    and tipo_col_name    in row_a else "—"
                                                        act_a   = str(row_a[act_col_pdf])      if act_col_pdf      and act_col_pdf      in row_a else "—"
                                                        fecha_a = str(row_a[fecha_col_pdf])    if fecha_col_pdf    and fecha_col_pdf    in row_a else "—"
                                                        horas_a = float(row_a.get("Horas", 0))
                                                        epu_c_a = EPU_PDF_COLORS.get(grupo_a.upper(), C_MED)
            
                                                        if horas_a >= 8:
                                                            est_a = "✅ Completa"; est_c_a = C_GREEN
                                                        elif horas_a >= 4:
                                                            est_a = "⚡ Parcial";  est_c_a = C_YELLOW
                                                        else:
                                                            est_a = "⚠️ Breve";   est_c_a = C_RED
            
                                                        act_rows_pdf.append([
                                                            Paragraph(f"<b><font color='white'>{grupo_a}</font></b>",
                                                                      ParagraphStyle("ep_a", alignment=TA_CENTER,
                                                                                     fontSize=8, fontName="Helvetica-Bold")),
                                                            tec_a, tipo_a,
                                                            Paragraph(act_a, ParagraphStyle("act_a", fontSize=7,
                                                                                             fontName="Helvetica")),
                                                            fecha_a, f"{horas_a:.2f}", est_a
                                                        ])
                                                        dr_a = len(act_rows_pdf) - 1
                                                        act_style_pdf.append(("BACKGROUND", (0, dr_a), (0, dr_a), epu_c_a))
                                                        act_style_pdf.append(("BACKGROUND", (6, dr_a), (6, dr_a), est_c_a))
            
                                                    story.append(Table(
                                                        act_rows_pdf,
                                                        colWidths=[3*cm, 4*cm, 3.5*cm, 7*cm, 2.8*cm, 2.2*cm, 3.5*cm],
                                                        style=TableStyle(act_style_pdf)
                                                    ))
            
                                                doc.build(story)
                                                buf.seek(0)
                                                return buf.read()
            
                                            # ---- BOTONES DE DESCARGA ----
                                            resumen_tab = resumen_merge.copy()
            
                                            col_dl1, col_dl2 = st.columns(2)
                                            with col_dl1:
                                                excel_bytes = _generar_excel_gap(
                                                    resumen_tab,
                                                    df_act_valido if not df_act_valido.empty else pd.DataFrame(),
                                                    tipo_col, tecnico_col,
                                                    periodo_txt, filtros_info
                                                )
                                                st.download_button(
                                                    "📥 Descargar Excel ejecutivo", excel_bytes,
                                                    file_name=f"REPORTE_GAP_{f_ini}_a_{f_fin}.xlsx",
                                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                    key="dl_gap_excel", use_container_width=True
                                                )
                                            with col_dl2:
                                                pdf_bytes = _generar_pdf_gap_ejecutivo(
                                                    resumen_tab,
                                                    df_act_valido if not df_act_valido.empty else pd.DataFrame(),
                                                    tipo_col, tecnico_col,
                                                    "Gap de Horas por Área / EPU",
                                                    periodo_txt, filtros_info
                                                )
                                                st.download_button(
                                                    "📄 Descargar PDF ejecutivo", pdf_bytes,
                                                    file_name=f"GAP_EPU_{f_ini}_a_{f_fin}.pdf",
                                                    mime="application/pdf",
                                                    key="dl_gap_pdf", use_container_width=True
                                                )
            
                                            st.markdown("---")
        
                                            # ============================
                                            # === NUEVA SECCIÓN: Comparativo por Turno y EPU (Desglosado) ===
                                            # ============================
            
                                            import unicodedata
                                            import pandas as pd
                                            import altair as alt
                                            import streamlit as st
                                            import difflib
                                            
                                            # =========================================================
                                            # GLOBAL: FUNCIONES DE NORMALIZACIÓN (Usadas por ambas vistas)
                                            # =========================================================
                                            def normalizar_texto_extremo(x):
                                                """
                                                Normalización general: quita espacios extras, pasa a ASCII sin tildes y a mayúsculas.
                                                Devuelve cadena vacía si el input es nulo/NaN/empty.
                                                Útil para columnas como EPU, Zona, Area, etc.
                                                """
                                                if pd.isna(x) or str(x).strip() == "" or str(x).lower() == "nan":
                                                    return ""
                                                s = str(x).strip().upper()
                                                s = " ".join(s.split())
                                                # Normalizar acentos y eliminar diacríticos
                                                s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
                                                return s
                                            
                                            def normalizar_nombre_key(x):
                                                """
                                                Normalización para llaves de supervisor: devuelve 'SIN ASIGNAR' si vacío.
                                                Esto evita introducir NaNs en los índices/etiquetas.
                                                """
                                                if pd.isna(x) or str(x).strip() == "" or str(x).lower() == "nan":
                                                    return "SIN ASIGNAR"
                                                s = str(x).strip().upper()
                                                s = " ".join(s.split())
                                                s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
                                                return s
                                            
                                            def mapeo_epu_pro(val):
                                                """
                                                Mapea valores normalizados a etiquetas EPU esperadas.
                                                """
                                                s = normalizar_texto_extremo(val)
                                                if not s:
                                                    return None
                                                if "EPU2" in s or "EPU 2" in s: return "EPU2"
                                                if "EPU3" in s or "EPU 3" in s: return "EPU3"
                                                if "EPU4" in s or "EPU 4" in s: return "EPU4"
                                                if "FACILIDAD" in s or "FACILIDADES" in s or s.startswith("FAC"): return "FACILIDADES"
                                                return None
                                            
                                            # =========================================================
                                            # === SECCIÓN: Comparativo de Horas Invertidas por Turno y EPU
                                            # =========================================================
                                            st.markdown("### 📊 Comparativo de Horas Invertidas por Turno y EPU (Desglosado)")
                                            
                                            # Hacemos copias locales para no modificar los DataFrames originales compartidos
                                            df_tec_chart = df_act_date.copy() if (('df_act_date' in globals() or 'df_act_date' in locals()) and not df_act_date.empty) else pd.DataFrame()
                                            df_sup_chart = df_sup_date.copy() if (('df_sup_date' in globals() or 'df_sup_date' in locals()) and not df_sup_date.empty) else pd.DataFrame()
                                            
                                            if df_tec_chart.empty:
                                                st.info("ℹ️ No hay datos de técnicos para mostrar la comparativa.")
                                            else:
                                                # Normalizar nombres de columnas (evita problemas por espacios)
                                                df_tec_chart.columns = [c.strip() for c in df_tec_chart.columns]
                                            
                                                # --- Normalización específica para la vista EPU (nombres únicos para no chocar) ---
                                                # Usamos sufijos de columna distintos para evitar sobrescribir variables que usa la vista supervisores.
                                                epu_group_col = None
                                                if 'grouping_candidate' in globals() and grouping_candidate and grouping_candidate in df_tec_chart.columns:
                                                    epu_group_col = grouping_candidate
                                                else:
                                                    # Fallback: buscar columnas comunes
                                                    for c_try in ["Zona", "zona", "EPU", "Epu", "area", "Area", "Unidad"]:
                                                        if c_try in df_tec_chart.columns:
                                                            epu_group_col = c_try
                                                            break
                                            
                                                if epu_group_col:
                                                    df_tec_chart["_EPU_Norm_tec"] = df_tec_chart[epu_group_col].apply(normalizar_texto_extremo).apply(mapeo_epu_pro)
                                                else:
                                                    df_tec_chart["_EPU_Norm_tec"] = None
                                            
                                                # Turno column (para técnicos)
                                                t_col_tec = None
                                                for t_try in ["Turno", "turno", "TURNO"]:
                                                    if t_try in df_tec_chart.columns:
                                                        t_col_tec = t_try
                                                        break
                                                if t_col_tec:
                                                    df_tec_chart["_Turno_Norm_tec"] = df_tec_chart[t_col_tec].astype(str).apply(lambda x: str(x).split('.')[0].strip() if str(x).strip() not in ["nan", "None", ""] else "Sin turno")
                                                else:
                                                    df_tec_chart["_Turno_Norm_tec"] = df_tec_chart.get("Turno", df_tec_chart.get("turno", "Sin turno")).astype(str).apply(lambda x: str(x).split('.')[0].strip() if str(x).strip() not in ["nan", "None", ""] else "Sin turno")
                                            
                                                # Horas técnicos: intentar convertir de varias columnas si es necesario
                                                if "Horas" not in df_tec_chart.columns:
                                                    candidate_min = None
                                                    for c_try in ["Tiempo_Minutos", "Minutos", "Duración (min)", "Duracion (min)"]:
                                                        if c_try in df_tec_chart.columns:
                                                            candidate_min = c_try
                                                            break
                                                    if candidate_min:
                                                        df_tec_chart["Horas"] = pd.to_numeric(df_tec_chart[candidate_min], errors="coerce").fillna(0) / 60.0
                                                    else:
                                                        # Si no hay columna de minutos, asegurar que la columna Horas exista
                                                        df_tec_chart["Horas"] = pd.to_numeric(df_tec_chart.get("Horas", 0), errors="coerce").fillna(0.0)
                                                else:
                                                    df_tec_chart["Horas"] = pd.to_numeric(df_tec_chart["Horas"], errors="coerce").fillna(0.0)
                                            
                                                # --- Preparación de datos de supervisores A REFERENCIA (solo para barra fantasma en EPU) ---
                                                if not df_sup_chart.empty:
                                                    df_sup_chart.columns = [c.strip() for c in df_sup_chart.columns]
                                                    # buscar columna que represente la EPU/zona en plantilla de supervisores
                                                    s_group_col = None
                                                    for c_try in ["zona", "Zona", "Unidad", "EPU", "Línea", "Area", "area"]:
                                                        if c_try in df_sup_chart.columns:
                                                            s_group_col = c_try
                                                            break
                                                    if s_group_col:
                                                        df_sup_chart["_EPU_Norm_sup"] = df_sup_chart[s_group_col].apply(normalizar_texto_extremo).apply(mapeo_epu_pro)
                                                    else:
                                                        df_sup_chart["_EPU_Norm_sup"] = None
                                            
                                                    # Turno en plantilla de supervisores (si existe)
                                                    t_col_sup = None
                                                    for t_try in ["Turno", "turno", "TURNO"]:
                                                        if t_try in df_sup_chart.columns:
                                                            t_col_sup = t_try
                                                            break
                                                    if t_col_sup:
                                                        df_sup_chart["_Turno_Norm_sup"] = df_sup_chart[t_col_sup].astype(str).apply(lambda x: str(x).split('.')[0].strip() if str(x).strip() not in ["nan", "None", ""] else "Sin turno")
                                                    else:
                                                        # Default: intentar usar la misma columna que técnicos si existe
                                                        df_sup_chart["_Turno_Norm_sup"] = df_tec_chart["_Turno_Norm_tec"] if "_Turno_Norm_tec" in df_tec_chart.columns else "Sin turno"
                                            
                                                    # Horas disponibles en plantilla (normalizar nombre de columna)
                                                    h_disp_col = None
                                                    for c_try in ["Horas_Disponibles", "DISPONIBLES", "Horas Disponibles"]:
                                                        if c_try in df_sup_chart.columns:
                                                            h_disp_col = c_try
                                                            break
                                                    if h_disp_col:
                                                        df_sup_chart["Horas_Disp"] = pd.to_numeric(df_sup_chart[h_disp_col], errors="coerce").fillna(0.0)
                                                    else:
                                                        df_sup_chart["Horas_Disp"] = 0.0
                                            
                                                # --- Pivots: técnicos y supervisores por EPU x Turno ---
                                                pivot_tec = pd.DataFrame()
                                                pivot_sup = pd.DataFrame(columns=["_EPU_Norm_sup", "_Turno_Norm_sup", "Horas_Disp"])
                                                try:
                                                    pivot_tec = df_tec_chart[df_tec_chart["_EPU_Norm_tec"].notna()].groupby(["_EPU_Norm_tec", "_Turno_Norm_tec"], dropna=False)["Horas"].sum().reset_index()
                                                except Exception:
                                                    pivot_tec = pd.DataFrame(columns=["_EPU_Norm_tec", "_Turno_Norm_tec", "Horas"])
                                                if not df_sup_chart.empty:
                                                    try:
                                                        pivot_sup = df_sup_chart[df_sup_chart["_EPU_Norm_sup"].notna()].groupby(["_EPU_Norm_sup", "_Turno_Norm_sup"], dropna=False)["Horas_Disp"].sum().reset_index()
                                                    except Exception:
                                                        pivot_sup = pd.DataFrame(columns=["_EPU_Norm_sup", "_Turno_Norm_sup", "Horas_Disp"])
                                            
                                                # --- RENDERIZADO ---
                                                EPU_COLORS = {"EPU2": "#1F6B3A", "EPU3": "#7B3F00", "EPU4": "#1A3A6B", "FACILIDADES": "#6A0572"}
                                                turnos_validos = ["1", "2", "3"]
                                            
                                                max_tec = pivot_tec["Horas"].max() if not pivot_tec.empty and "Horas" in pivot_tec.columns else 0
                                                max_sup = pivot_sup["Horas_Disp"].max() if not pivot_sup.empty and "Horas_Disp" in pivot_sup.columns else 0
                                                ymax_global = max(max_tec, max_sup, 1.0) * 1.2
                                            
                                                cols = st.columns(4)
                                                for i, epu in enumerate(["EPU2", "EPU3", "EPU4", "FACILIDADES"]):
                                                    with cols[i]:
                                                        plot_data = []
                                                        for t in turnos_validos:
                                                            try:
                                                                h_tec = pivot_tec[(pivot_tec["_EPU_Norm_tec"] == epu) & (pivot_tec["_Turno_Norm_tec"] == t)]["Horas"].sum()
                                                            except Exception:
                                                                h_tec = 0.0
                                                            try:
                                                                h_sup = pivot_sup[(pivot_sup["_EPU_Norm_sup"] == epu) & (pivot_sup["_Turno_Norm_sup"] == t)]["Horas_Disp"].sum() if not pivot_sup.empty else 0.0
                                                            except Exception:
                                                                h_sup = 0.0
                                                            plot_data.append({"Turno": t, "Invertidas": float(h_tec), "Disponibles": float(h_sup)})
                                            
                                                        df_plot = pd.DataFrame(plot_data)
                                            
                                                        if df_plot["Invertidas"].sum() == 0 and df_plot["Disponibles"].sum() == 0:
                                                            st.caption(f"Sin datos para {epu}")
                                                            continue
                                            
                                                        base = alt.Chart(df_plot).encode(x=alt.X("Turno:N", sort=turnos_validos, title="Turno"))
                                            
                                                        # Barra fantasma (Horas Disponibles - Supervisor)
                                                        ghost = base.mark_bar(opacity=0.2, color="#9ea7ad", size=40).encode(
                                                            y=alt.Y("Disponibles:Q", scale=alt.Scale(domain=[0, ymax_global]), title="Horas")
                                                        )
                                            
                                                        # Barra real (Horas Invertidas - Técnicos)
                                                        bars = base.mark_bar(size=25).encode(
                                                            y=alt.Y("Invertidas:Q"),
                                                            color=alt.value(EPU_COLORS.get(epu, "#2E75B6")),
                                                            tooltip=[
                                                                alt.Tooltip("Turno:N"),
                                                                alt.Tooltip("Invertidas:Q", title="Horas Reales", format=".1f"),
                                                                alt.Tooltip("Disponibles:Q", title="Horas Disponibles", format=".1f")
                                                            ]
                                                        )
                                            
                                                        # Etiqueta de texto
                                                        text = base.mark_text(dy=-10, fontWeight="bold").encode(
                                                            y=alt.Y("Invertidas:Q"),
                                                            text=alt.Text("Invertidas:Q", format=".1f")
                                                        )
                                            
                                                        st.altair_chart((ghost + bars + text).properties(title=f"Horas {epu}", height=260), use_container_width=True)
                                            
                                            st.markdown("---")
                                            
                                            # =========================================================
                                            # === VISTA TÉCNICOS (tab2) - MANTENIDO EXACTAMENTE COMO ESTABA
                                            # =========================================================
                                            if st.session_state.vista_admin == "tecnicos":
                                                st.markdown("### 🛠️ Rendimiento de Técnicos")
                                            
                                                if df_act_date.empty:
                                                    st.info("ℹ️ Sin actividades de técnicos en el periodo seleccionado.")
                                                else:
                                                    # Gráfica de horas por técnico
                                                    if tecnico_col and tecnico_col in df_act_date.columns:
                                                        df_act_date["Horas"] = pd.to_numeric(
                                                            df_act_date.get("Tiempo_Minutos", 0), errors="coerce"
                                                        ).fillna(0) / 60.0
                                            
                                                        horas_tec = (
                                                            df_act_date.groupby(tecnico_col)["Horas"]
                                                            .sum().reset_index()
                                                            .sort_values("Horas", ascending=False)
                                                        )
                                            
                                                        chart_tec = alt.Chart(horas_tec).mark_bar().encode(
                                                            x=alt.X(f"{tecnico_col}:N",
                                                                    sort=alt.EncodingSortField(field="Horas", order="descending"),
                                                                    title="Técnico"),
                                                            y=alt.Y("Horas:Q", title="Horas Invertidas"),
                                                            color=alt.Color(f"{tecnico_col}:N",
                                                                            scale=alt.Scale(scheme="tableau10"),
                                                                            legend=None),
                                                            tooltip=[
                                                                alt.Tooltip(f"{tecnico_col}:N", title="Técnico"),
                                                                alt.Tooltip("Horas:Q", title="Horas", format=".2f")
                                                            ]
                                                        ).properties(
                                                            height=350,
                                                            title="Horas invertidas por técnico"
                                                        )
                                                        st.altair_chart(chart_tec, use_container_width=True)
                                            
                                                    # Gráfica por tipo de mantenimiento
                                                    if tipo_col and tipo_col in df_act_date.columns:
                                                        horas_tipo = (
                                                            df_act_date.groupby(tipo_col)["Horas"]
                                                            .sum().reset_index()
                                                            .sort_values("Horas", ascending=False)
                                                        )
                                                        chart_tipo = alt.Chart(horas_tipo).mark_arc(innerRadius=50).encode(
                                                            theta=alt.Theta("Horas:Q"),
                                                            color=alt.Color(f"{tipo_col}:N",
                                                                            scale=alt.Scale(scheme="category10"),
                                                                            title="Tipo"),
                                                            tooltip=[
                                                                alt.Tooltip(f"{tipo_col}:N", title="Tipo"),
                                                                alt.Tooltip("Horas:Q", title="Horas", format=".2f")
                                                            ]
                                                        ).properties(
                                                            height=300,
                                                            title="Distribución por tipo de mantenimiento"
                                                        )
                                                        st.altair_chart(chart_tipo, use_container_width=True)
                                            
                                            # =========================================================
                                            # === VISTA SUPERVISORES (tab2) - IMPLEMENTACIÓN CORRECTA con mapeo tolerante
                                            # =========================================================
                                            elif st.session_state.vista_admin == "supervisores":
                                                st.markdown("### 📋 Rendimiento de Supervisores")
                                            
                                                if df_sup_date.empty:
                                                    st.info("ℹ️ Sin registros de supervisores en el periodo seleccionado.")
                                                else:
                                                    # Preparamos una copia local para no interferir con df_sup_date global
                                                    df_sup_work = df_sup_date.copy()
                                                    # Normalizar columnas numéricas si existen
                                                    cols_num_sup = ["Horas_Disponibles", "Total_Real", "Inasistencias", "Jornada_Normal", "Total_Esperado"]
                                                    for c in cols_num_sup:
                                                        if c in df_sup_work.columns:
                                                            df_sup_work[c] = pd.to_numeric(df_sup_work[c], errors="coerce").fillna(0)
                                            
                                                    # Identificar columna de nombre de supervisor en plantilla
                                                    sup_name_col = find_col(df_sup_work, ["Supervisor", "supervisor", "Nombre", "nombre", "Usuario"])
                                                    if sup_name_col is None:
                                                        st.warning("⚠️ No se encontró columna de nombre de supervisor en la plantilla.")
                                                        # Mostrar tabla cruda y salir
                                                        st.dataframe(df_sup_work.head(), use_container_width=True)
                                                    else:
                                                        # Crear llave de cruce limpia en plantilla (usa normalizar_nombre_key)
                                                        df_sup_work["_sup_key"] = df_sup_work[sup_name_col].apply(normalizar_nombre_key)
                                                        # Asegurar columna Horas_Disponibles
                                                        if "Horas_Disponibles" not in df_sup_work.columns:
                                                            # Intentar detectar columna alternativa
                                                            for alt_c in ["Horas Disponibles", "DISPONIBLES"]:
                                                                if alt_c in df_sup_work.columns:
                                                                    df_sup_work["Horas_Disponibles"] = pd.to_numeric(df_sup_work[alt_c], errors="coerce").fillna(0)
                                                                    break
                                                        df_sup_work["Horas_Disponibles"] = pd.to_numeric(df_sup_work.get("Horas_Disponibles", 0), errors="coerce").fillna(0)
                                            
                                                        # Métricas resumen
                                                        m1, m2, m3 = st.columns(3)
                                                        with m1:
                                                            st.metric("Total Horas Disponibles", f"{df_sup_work['Horas_Disponibles'].sum():.1f} hrs")
                                                        with m2:
                                                            # Calcular Horas Reales desde df_act_date (si no existe, 0)
                                                            total_h_reales = 0.0
                                                            if 'Horas' in df_act_date.columns:
                                                                total_h_reales = pd.to_numeric(df_act_date['Horas'], errors="coerce").fillna(0).sum()
                                                            else:
                                                                # Intentar calcular desde minutos
                                                                if "Tiempo_Minutos" in df_act_date.columns:
                                                                    total_h_reales = pd.to_numeric(df_act_date["Tiempo_Minutos"], errors="coerce").fillna(0).sum() / 60.0
                                                            st.metric("Total Horas Reales", f"{total_h_reales:.1f} hrs")
                                                        with m3:
                                                            st.metric("Total Inasistencias", f"{df_sup_work['Inasistencias'].sum():.0f}")
                                            
                                                        # Preparar actividades de técnicos (local copy)
                                                        sup_col_tec = find_col(df_act_date, ["Supervisor", "supervisor", "SUPERVISOR", "Procesado por", "Sup", "Jefe"])
                                                        if sup_col_tec is None:
                                                            st.warning("⚠️ No se encontró columna con el nombre del supervisor en los registros de técnicos.")
                                                            # Mostrar detalle de plantilla y salir
                                                            st.dataframe(df_sup_work[[sup_name_col, "Horas_Disponibles"]].head(), use_container_width=True)
                                                        else:
                                                            df_act_sup = df_act_date.copy()
                                                            # Asegurarse de calcular Horas si no existe
                                                            if "Horas" not in df_act_sup.columns:
                                                                if "Tiempo_Minutos" in df_act_sup.columns:
                                                                    df_act_sup["Horas"] = pd.to_numeric(df_act_sup["Tiempo_Minutos"], errors="coerce").fillna(0) / 60.0
                                                                elif "Duración (min)" in df_act_sup.columns or "Duracion (min)" in df_act_sup.columns:
                                                                    dur_col = "Duración (min)" if "Duración (min)" in df_act_sup.columns else "Duracion (min)"
                                                                    df_act_sup["Horas"] = pd.to_numeric(df_act_sup[dur_col], errors="coerce").fillna(0) / 60.0
                                                                else:
                                                                    df_act_sup["Horas"] = pd.to_numeric(df_act_sup.get("Horas", 0), errors="coerce").fillna(0.0)
                                                            else:
                                                                df_act_sup["Horas"] = pd.to_numeric(df_act_sup["Horas"], errors="coerce").fillna(0.0)
                                            
                                                            # Normalizar el nombre de supervisor que registra cada técnico
                                                            df_act_sup["_sup_key_raw"] = df_act_sup[sup_col_tec].astype(str).apply(lambda v: v if pd.notna(v) else "")
                                                            df_act_sup["_sup_key"] = df_act_sup["_sup_key_raw"].apply(normalizar_nombre_key)
                                            
                                                            # Construir lista de llaves válidas de la plantilla
                                                            lista_sups_validos = df_sup_work["_sup_key"].dropna().unique().tolist()
                                            
                                                            # Mapeo tolerante: para cada supervisor reportado por técnicos, encontrar mejor match en plantilla
                                                            # Si existe match exacto lo usa, si no, busca close match con difflib (cutoff 0.8). Ajusta cutoff si quieres ser más permisivo.
                                                            def map_to_plantilla_key(key_raw, key_norm, plantilla_keys, cutoff=0.8):
                                                                if key_norm in plantilla_keys:
                                                                    return key_norm
                                                                # intentar close match
                                                                matches = difflib.get_close_matches(key_norm, plantilla_keys, n=1, cutoff=cutoff)
                                                                if matches:
                                                                    return matches[0]
                                                                return None
                                            
                                                            # Crear columna mapeada
                                                            df_act_sup["_sup_key_mapped"] = df_act_sup.apply(
                                                                lambda r: map_to_plantilla_key(r["_sup_key_raw"], r["_sup_key"], lista_sups_validos, cutoff=0.8),
                                                                axis=1
                                                            )
                                            
                                                            # Para trazabilidad, podemos mantener original y mapped; filas sin mapped=None no serán contabilizadas en el stacked (evitan NAN)
                                                            # Agrupar horas invertidas por supervisor mapeado y por tipo
                                                            if tipo_col and tipo_col in df_act_sup.columns:
                                                                # Filtrar solo filas mapeadas y con horas > 0
                                                                df_act_mapped = df_act_sup[df_act_sup["_sup_key_mapped"].notna() & (df_act_sup["Horas"] > 0)].copy()
                                                                sum_inv_by_type = df_act_mapped.groupby(["_sup_key_mapped", tipo_col], dropna=False)["Horas"].sum().reset_index()
                                                                sum_inv_by_type.columns = ["Supervisor_Key", "Tipo", "Horas"]
                                                            else:
                                                                # Si no hay tipo_col, crear total por supervisor mapeado
                                                                df_act_mapped = df_act_sup[df_act_sup["_sup_key_mapped"].notna() & (df_act_sup["Horas"] > 0)].copy()
                                                                sum_inv_by_type = pd.DataFrame(columns=["Supervisor_Key", "Tipo", "Horas"])
                                            
                                                            # Totales por supervisor mapeado
                                                            sum_inv_total = df_act_sup[df_act_sup["_sup_key_mapped"].notna()].groupby("_sup_key_mapped")["Horas"].sum().reset_index().rename(columns={"_sup_key_mapped": "Supervisor_Key", "Horas": "Horas_Invertidas"})
                                            
                                                            # Agrupar horas disponibles por supervisor en plantilla
                                                            sum_disp = df_sup_work.groupby("_sup_key")["Horas_Disponibles"].sum().reset_index().rename(columns={"_sup_key": "Supervisor_Key", "Horas_Disponibles": "Horas_Disponibles"})
                                            
                                                            # Preparar resumen (merge)
                                                            resumen_sup = pd.merge(sum_disp, sum_inv_total, on="Supervisor_Key", how="left").fillna(0)
                                                            resumen_sup = resumen_sup.rename(columns={"Supervisor_Key": "Supervisor"})
                                                            # Si la columna Horas_Invertidas no existe, crearla
                                                            if "Horas_Invertidas" not in resumen_sup.columns:
                                                                resumen_sup["Horas_Invertidas"] = 0.0
                                                            resumen_sup["Gap"] = (resumen_sup["Horas_Disponibles"] - resumen_sup["Horas_Invertidas"]).clip(lower=0)
                                                            resumen_sup = resumen_sup.sort_values("Horas_Disponibles", ascending=False).reset_index(drop=True)
                                                            resumen_sup["Label"] = resumen_sup["Gap"].apply(lambda v: f"Gap: {v:.1f} hrs")
                                            
                                                            # Si sum_inv_by_type está vacío, crear fallback con zeros para que el stacked no falle
                                                            if sum_inv_by_type.empty:
                                                                sum_inv_by_type = pd.DataFrame({
                                                                    "Supervisor_Key": resumen_sup["Supervisor"].tolist(),
                                                                    "Tipo": ["Sin actividad"] * len(resumen_sup),
                                                                    "Horas": [0.0] * len(resumen_sup)
                                                                })
                                            
                                                            # Asegurar que los nombres de ejes coincidan: stacked usa Supervisor_Key (llave mapeada),
                                                            # mientras que resumen_sup usa columna 'Supervisor' que es la misma llave mapeada.
                                                            order_sup = resumen_sup["Supervisor"].astype(str).tolist()
                                                            ymax_sup = max(
                                                                resumen_sup["Horas_Disponibles"].max() if not resumen_sup["Horas_Disponibles"].isna().all() else 0,
                                                                resumen_sup["Horas_Invertidas"].max() if not resumen_sup["Horas_Invertidas"].isna().all() else 0,
                                                                1
                                                            ) * 1.15
                                            
                                                            ghost_sup = alt.Chart(resumen_sup).mark_bar(opacity=0.25, color="#9ea7ad").encode(
                                                                x=alt.X("Supervisor:N", sort=order_sup, title="Supervisor"),
                                                                y=alt.Y("Horas_Disponibles:Q", title="Horas", scale=alt.Scale(domain=[0, ymax_sup])),
                                                                tooltip=[alt.Tooltip("Supervisor:N", title="Supervisor"), alt.Tooltip("Horas_Disponibles:Q", title="Disponible", format=".1f")]
                                                            )
                                            
                                                            stacked_sup = alt.Chart(sum_inv_by_type).mark_bar().encode(
                                                                x=alt.X("Supervisor_Key:N", sort=order_sup, title="Supervisor"),
                                                                y=alt.Y("Horas:Q", stack="zero"),
                                                                color=alt.Color("Tipo:N", title="Tipo Mantenimiento", scale=alt.Scale(scheme="category10")),
                                                                tooltip=[alt.Tooltip("Supervisor_Key:N", title="Supervisor"), alt.Tooltip("Tipo:N", title="Tipo"), alt.Tooltip("Horas:Q", format=".1f")]
                                                            )
                                            
                                                            labels_sup = alt.Chart(resumen_sup).mark_text(color="red", fontWeight="bold", dy=-15, fontSize=12).encode(
                                                                x=alt.X("Supervisor:N", sort=order_sup),
                                                                y=alt.Y("Horas_Disponibles:Q"),
                                                                text=alt.Text("Label:N")
                                                            )
                                            
                                                            gap_sup_chart = alt.layer(ghost_sup, stacked_sup, labels_sup).properties(
                                                                height=450,
                                                                title="GAP por Supervisor: Horas Disponibles (plantilla) vs Horas Invertidas por Tipo (técnicos)"
                                                            )
                                                            st.altair_chart(gap_sup_chart, use_container_width=True)
                                            
                                                            # Tabla resumen con semáforo
                                                            st.markdown("#### 📄 Tabla de Rendimiento por Supervisor")
                                                            resumen_sup_display = resumen_sup[["Supervisor", "Horas_Disponibles", "Horas_Invertidas", "Gap"]].copy()
                                                            resumen_sup_display["% Utilización"] = resumen_sup_display.apply(
                                                                lambda r: (r["Horas_Invertidas"] / r["Horas_Disponibles"] * 100) if r["Horas_Disponibles"] > 0 else 0.0, axis=1
                                                            )
                                                            # Aplicar función de semáforo (se asume que _semaforo_icon está definida)
                                                            try:
                                                                resumen_sup_display["Semáforo"] = resumen_sup_display["% Utilización"].apply(_semaforo_icon)
                                                            except Exception:
                                                                # Si _semaforo_icon no existe, rellenar con placeholder
                                                                resumen_sup_display["Semáforo"] = "⚪"
                                            
                                                            resumen_sup_display["% Utilización"] = resumen_sup_display["% Utilización"].apply(lambda v: f"{v:.1f}%")
                                                            st.dataframe(
                                                                resumen_sup_display[["Semáforo", "Supervisor", "Horas_Disponibles", "Horas_Invertidas", "Gap", "% Utilización"]],
                                                                use_container_width=True
                                                            )
                                            
                                                        # Tabla detalle registros supervisores (mostrar df_sup_date original)
                                                        st.markdown("#### 🗂️ Detalle de Registros de Supervisores")
                                                        st.dataframe(
                                                            df_sup_date[[c for c in [sup_name_col, "Fecha", "Horas_Disponibles", "Total_Real", "Inasistencias", "Jornada_Normal"] if c and c in df_sup_date.columns]],
                                                            use_container_width=True
                                                        )

                          
        # ---------------- TAB 3: DESCARGAS ADMIN ----------------
        with tab3:
            st.subheader("📂 Descargar reportes")
    
            # Selector dual elegante
            col_sel1, col_sel2 = st.columns(2)
            with col_sel1:
                if st.button("🛠️ TÉCNICOS", width='stretch', type="primary" if st.session_state.vista_admin == "tecnicos" else "secondary", key="btn_desc_tec"):
                    st.session_state.vista_admin = "tecnicos"
                    st.rerun()
            with col_sel2:
                if st.button("📋 SUPERVISORES", width='stretch', type="primary" if st.session_state.vista_admin == "supervisores" else "secondary", key="btn_desc_sup"):
                    st.session_state.vista_admin = "supervisores"
                    st.rerun()
    
            st.markdown("---")
    
            # DESCARGAS - TÉCNICOS
            if st.session_state.vista_admin == "tecnicos":
                st.markdown("### 🛠️ Descargas de Técnicos")
                try:
                    df = cargar_datos_tecnicos()
                except Exception as e:
                    st.error("Error cargando actividades de técnicos: " + str(e))
                    df = pd.DataFrame()
    
                if df is None:
                    df = pd.DataFrame()
    
                if not df.empty:
                    modo_descarga = st.selectbox("Modo de descarga", ["Día específico", "Seleccionar rango"], key="modo_desc_tec")
    
                    def to_excel_tecnicos(df_to_write, title="Bitácora Técnicos", periodo=""):
                        return excel_executive(df_to_write.fillna(""), title, periodo, {"Area": st.session_state.get("area","")})
    
                    if modo_descarga == "Día específico":
                        fecha_excel = st.date_input("Selecciona la fecha", datetime.date.today(), key="fecha_excel_tec")
                        # Normalizar Fecha en df antes de filtrar
                        if "Fecha" not in df.columns:
                            if "Inicio" in df.columns:
                                df["Fecha"] = pd.to_datetime(df["Inicio"], errors="coerce").dt.date
                            else:
                                df["Fecha"] = pd.NaT
                        else:
                            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
    
                        if isinstance(fecha_excel, datetime.datetime):
                            fecha_excel = fecha_excel.date()
                        df_dia = df[df["Fecha"] == fecha_excel].copy()
                        if not df_dia.empty:
                            excel_data = to_excel_tecnicos(df_dia, f"Bitácora Técnicos - {st.session_state.get('area','')}", str(fecha_excel))
                            st.download_button(label=f"📥 Descargar Excel del {fecha_excel}", data=excel_data, file_name=f"Bitacora_Tecnicos_{st.session_state.get('area','')}_{fecha_excel}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tec_excel", width='stretch')
    
                            # PDF resumen
                            resumen = df_dia.groupby(["Usuario", "Nombre"]).agg({"Duración (min)": "sum", "ID": "count"}).reset_index().rename(columns={"ID":"Actividades"})
                            resumen["Etiqueta"] = resumen["Usuario"].astype(str) + " | " + resumen["Nombre"].astype(str)
                            def generar_pdf_tecnicos(resumen, titulo):
                                fig, axs = plt.subplots(1, 3, figsize=(16,6))
                                fig.suptitle(titulo, fontsize=16, fontweight="bold")
                                n = len(resumen)
                                colores = cm.get_cmap('tab20', max(1, n))(np.linspace(0,1,max(1,n)))
                                axs[0].bar(resumen["Etiqueta"], resumen["Duración (min)"], color=colores)
                                axs[0].set_title("Minutos invertidos por usuario")
                                axs[0].tick_params(axis='x', rotation=45)
                                axs[1].barh(resumen["Etiqueta"], resumen["Actividades"], color=colores)
                                axs[1].set_title("Número de actividades por usuario")
                                axs[2].scatter(resumen["Actividades"], resumen["Duración (min)"], s=120, color=colores)
                                axs[2].set_title("Diagrama de eficiencia")
                                axs[2].set_xlabel("Actividades")
                                axs[2].set_ylabel("Minutos")
                                plt.subplots_adjust(wspace=0.4)
                                pdf_buffer = BytesIO()
                                with PdfPages(pdf_buffer) as pdf:
                                    pdf.savefig(fig, bbox_inches='tight')
                                plt.close(fig)
                                return pdf_buffer.getvalue()
    
                            pdf_data = generar_pdf_tecnicos(resumen, f"Reporte de Productividad Técnicos - {fecha_excel}")
                            st.download_button(label=f"📄 Descargar PDF de rendimiento {fecha_excel}", data=pdf_data, file_name=f"Rendimiento_Tecnicos_{st.session_state.get('area','')}_{fecha_excel}.pdf", mime="application/pdf", key="dl_tec_pdf", width='stretch')
                        else:
                            st.info(f"⚠️ No hay actividades registradas para el día {fecha_excel}.")
                    else:
                        col1, col2 = st.columns(2)
                        with col1:
                            fecha_inicio = st.date_input("Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio_descarga_tec")
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", datetime.date.today(), key="fecha_fin_descarga_tec")
                        # Normalizar Fecha en df antes de filtrar por rango
                        if "Fecha" not in df.columns:
                            if "Inicio" in df.columns:
                                df["Fecha"] = pd.to_datetime(df["Inicio"], errors="coerce").dt.date
                            else:
                                df["Fecha"] = pd.NaT
                        else:
                            df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date
    
                        if isinstance(fecha_inicio, datetime.datetime):
                            fecha_inicio = fecha_inicio.date()
                        if isinstance(fecha_fin, datetime.datetime):
                            fecha_fin = fecha_fin.date()
    
                        df_rango = df[(df["Fecha"] >= fecha_inicio) & (df["Fecha"] <= fecha_fin)].copy()
                        if not df_rango.empty:
                            excel_data = to_excel_tecnicos(df_rango, f"Bitácora Técnicos - {st.session_state.get('area','')}", f"{fecha_inicio} a {fecha_fin}")
                            st.download_button(label=f"📥 Descargar Excel del rango {fecha_inicio} a {fecha_fin}", data=excel_data, file_name=f"Bitacora_Tecnicos_{st.session_state.get('area','')}_{fecha_inicio}_a_{fecha_fin}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tec_range_excel", width='stretch')
                            resumen = df_rango.groupby(["Usuario", "Nombre"]).agg({"Duración (min)": "sum", "ID": "count"}).reset_index().rename(columns={"ID":"Actividades"})
                            resumen["Etiqueta"] = resumen["Usuario"].astype(str) + " | " + resumen["Nombre"].astype(str)
                            pdf_data = generar_pdf_tecnicos(resumen, f"Reporte de Productividad Técnicos - {fecha_inicio} a {fecha_fin}")
                            st.download_button(label=f"📄 Descargar PDF de rendimiento {fecha_inicio} a {fecha_fin}", data=pdf_data, file_name=f"Rendimiento_Tecnicos_{st.session_state.get('area','')}_{fecha_inicio}_a_{fecha_fin}.pdf", mime="application/pdf", key="dl_tec_range_pdf", width='stretch')
                        else:
                            st.info(f"⚠️ No hay actividades registradas entre {fecha_inicio} y {fecha_fin}.")
                else:
                    st.info("⚠️ No hay registros disponibles para descargar.")
    
            # DESCARGAS - SUPERVISORES
            else:
                st.markdown("### 📋 Descargas de Supervisores")
                try:
                    df = cargar_datos_supervisores()
                except Exception as e:
                    st.error("Error cargando registros de supervisores: " + str(e))
                    df = pd.DataFrame()
                if df is None:
                    df = pd.DataFrame()
    
                if not df.empty:
                    modo_descarga = st.selectbox("Modo de descarga", ["Día específico", "Seleccionar rango"], key="modo_desc_sup")
                    def to_excel_supervisor(df_to_write, title="Registros Supervisor", periodo=""):
                        return excel_executive(df_to_write.fillna(""), title, periodo, {"Area": st.session_state.get("area","")})
    
                    if modo_descarga == "Día específico":
                        fecha_excel = st.date_input("Selecciona la fecha", datetime.date.today(), key="fecha_excel_sup_admin")
                        # Normalizar Fecha_obj
                        if "Fecha_obj" not in df.columns:
                            if "Fecha" in df.columns:
                                df["Fecha_obj"] = pd.to_datetime(df["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                            elif "Inicio" in df.columns:
                                df["Fecha_obj"] = pd.to_datetime(df["Inicio"], errors="coerce").dt.date
                            else:
                                df["Fecha_obj"] = pd.NaT
                        else:
                            df["Fecha_obj"] = pd.to_datetime(df["Fecha_obj"], errors="coerce").dt.date
    
                        if isinstance(fecha_excel, datetime.datetime):
                            fecha_excel = fecha_excel.date()
                        df_dia = df[df["Fecha_obj"] == fecha_excel].copy()
                        if not df_dia.empty:
                            excel_data = to_excel_supervisor(df_dia, "Registros Supervisor", str(fecha_excel))
                            st.download_button(label=f"📥 Descargar Excel del {fecha_excel}", data=excel_data, file_name=f"Supervisores_{st.session_state.get('area','')}_{fecha_excel}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_sup_excel", width='stretch')
                        else:
                            st.info(f"⚠️ No hay registros para el día {fecha_excel}.")
                    else:
                        col1, col2 = st.columns(2)
                        with col1:
                            fecha_inicio = st.date_input("Fecha inicio", datetime.date.today() - datetime.timedelta(days=7), key="fecha_inicio_descarga_sup_admin")
                        with col2:
                            fecha_fin = st.date_input("Fecha fin", datetime.date.today(), key="fecha_fin_descarga_sup_admin")
                        # Normalizar Fecha_obj
                        if "Fecha_obj" not in df.columns:
                            if "Fecha" in df.columns:
                                df["Fecha_obj"] = pd.to_datetime(df["Fecha"], format="%d-%m-%Y", errors="coerce").dt.date
                            elif "Inicio" in df.columns:
                                df["Fecha_obj"] = pd.to_datetime(df["Inicio"], errors="coerce").dt.date
                            else:
                                df["Fecha_obj"] = pd.NaT
                        else:
                            df["Fecha_obj"] = pd.to_datetime(df["Fecha_obj"], errors="coerce").dt.date
    
                        if isinstance(fecha_inicio, datetime.datetime):
                            fecha_inicio = fecha_inicio.date()
                        if isinstance(fecha_fin, datetime.datetime):
                            fecha_fin = fecha_fin.date()
    
                        df_rango = df[(df["Fecha_obj"] >= fecha_inicio) & (df["Fecha_obj"] <= fecha_fin)].copy()
                        if not df_rango.empty:
                            excel_data = to_excel_supervisor(df_rango, "Registros Supervisor", f"{fecha_inicio} a {fecha_fin}")
                            st.download_button(label=f"📥 Descargar Excel del rango {fecha_inicio} a {fecha_fin}", data=excel_data, file_name=f"Supervisores_{st.session_state.get('area','')}_{fecha_inicio}_a_{fecha_fin}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_sup_range_excel", width='stretch')
                        else:
                            st.info(f"⚠️ No hay registros entre {fecha_inicio} y {fecha_fin}.")
                else:
                    st.info("No hay registros disponibles para descargar.")
    

###################################################################################################################################################################################

        # TAB 4: USUARIOS ACTIVOS ADMIN
        with tab4:
                    st.subheader("👥 Usuarios activos")
        
                    empleados_ref = st.session_state.db.collection("empleados").stream()
                    empleados_data = [doc.to_dict() for doc in empleados_ref]
        
                    hoy_date = datetime.datetime.now().date()
        
                    def calcular_minutos_hoy(mx_id, actividades_hoy):
                        actos = [a for a in actividades_hoy if str(a.get("Usuario", "")).upper().strip() == mx_id]
                        total_min = 0
                        for a in actos:
                            try:
                                ini = pd.to_datetime(a.get("Inicio", ""), errors="coerce")
                                fin = pd.to_datetime(a.get("Fin", ""), errors="coerce")
                                if pd.notna(ini) and pd.notna(fin):
                                    total_min += (fin - ini).total_seconds() / 60
                            except:
                                pass
                        return round(total_min, 1)
        
                    try:
                        acts_ref = st.session_state.db.collection("actividades").stream()
                        actividades_hoy = []
                        for doc in acts_ref:
                            d = doc.to_dict()
                            fecha_act = d.get("Fecha", d.get("Inicio", ""))[:10] if (d.get("Fecha") or d.get("Inicio")) else ""
                            try:
                                fecha_obj = datetime.datetime.strptime(fecha_act, "%d-%m-%Y").date()
                            except:
                                try:
                                    fecha_obj = datetime.datetime.strptime(fecha_act, "%Y-%m-%d").date()
                                except:
                                    fecha_obj = None
                            if fecha_obj == hoy_date:
                                actividades_hoy.append(d)
                    except:
                        actividades_hoy = []
        
                    if empleados_data:
                        tecnicos = [e for e in empleados_data if e.get("role") == "tecnico"]
                        supervisores = [e for e in empleados_data if e.get("role") == "supervisor"]
        
                        # ── Selector con botones grandes ──────────────────────────────
                        if "tab4_vista" not in st.session_state:
                            st.session_state.tab4_vista = "tecnicos"
        
                        btn_col1, btn_col2 = st.columns(2)
                        with btn_col1:
                            activo_tec = st.session_state.tab4_vista == "tecnicos"
                            st.markdown(f"""
                            <div style="
                                background: {'linear-gradient(135deg,#0057B8,#0099FF)' if activo_tec else 'linear-gradient(135deg,#0d1b2a,#1a2f45)'};
                                border: {'2px solid #00BFFF' if activo_tec else '1px solid #00BFFF33'};
                                border-radius: 14px; padding: 18px 10px; text-align: center; cursor: pointer;
                                box-shadow: {'0 0 18px #00BFFF55' if activo_tec else 'none'};
                                transition: all 0.3s ease;">
                                <div style="font-size:2rem;">🛠️</div>
                                <div style="color:{'#fff' if activo_tec else '#00BFFF'};font-weight:700;font-size:1rem;margin-top:6px;">Plantilla Técnica</div>
                                <div style="color:{'#cce' if activo_tec else '#666'};font-size:0.78rem;margin-top:3px;">{len(tecnicos)} técnico(s) registrados</div>
                            </div>
                            """, unsafe_allow_html=True)
                            if st.button("Ver Técnicos", key="btn_vista_tec", use_container_width=True):
                                st.session_state.tab4_vista = "tecnicos"
                                st.rerun()
        
                        with btn_col2:
                            activo_sup = st.session_state.tab4_vista == "supervisores"
                            st.markdown(f"""
                            <div style="
                                background: {'linear-gradient(135deg,#0057B8,#0099FF)' if activo_sup else 'linear-gradient(135deg,#0d1b2a,#1a2f45)'};
                                border: {'2px solid #00BFFF' if activo_sup else '1px solid #00BFFF33'};
                                border-radius: 14px; padding: 18px 10px; text-align: center; cursor: pointer;
                                box-shadow: {'0 0 18px #00BFFF55' if activo_sup else 'none'};
                                transition: all 0.3s ease;">
                                <div style="font-size:2rem;">📋</div>
                                <div style="color:{'#fff' if activo_sup else '#00BFFF'};font-weight:700;font-size:1rem;margin-top:6px;">Cuerpo de Supervisión</div>
                                <div style="color:{'#cce' if activo_sup else '#666'};font-size:0.78rem;margin-top:3px;">{len(supervisores)} supervisor(es) registrados</div>
                            </div>
                            """, unsafe_allow_html=True)
                            if st.button("Ver Supervisores", key="btn_vista_sup", use_container_width=True):
                                st.session_state.tab4_vista = "supervisores"
                                st.rerun()
        
                        st.markdown("---")
        
                        # ══════════════════════════════════════════════════════════════
                        # VISTA TÉCNICOS
                        # ══════════════════════════════════════════════════════════════
                        if st.session_state.tab4_vista == "tecnicos":
                            st.markdown("### 🛠️ Plantilla Técnica")
                            busqueda_tec = st.text_input("🔍 Buscar por nombre o MX", placeholder="Escribe nombre o MX...", key="busq_tec")
        
                            if tecnicos:
                                tecnicos_filtrados = tecnicos
                                if busqueda_tec.strip():
                                    q = busqueda_tec.strip().lower()
                                    tecnicos_filtrados = [
                                        e for e in tecnicos
                                        if q in e.get("nombre", "").lower() or q in e.get("mx", "").lower()
                                    ]
        
                                if not tecnicos_filtrados:
                                    st.info("No se encontraron técnicos con ese criterio.")
                                else:
                                    st.markdown(f"<div style='color:#aaa;font-size:0.8rem;margin-bottom:8px;'>Mostrando {len(tecnicos_filtrados)} técnico(s)</div>", unsafe_allow_html=True)
                                    for i, user in enumerate(tecnicos_filtrados):
                                        nombre = user.get("nombre", "Sin nombre")
                                        mx_id = user.get("mx", f"MX_{i}").upper().strip()
                                        unidad = user.get("unidad", "No especificada")
                                        business_unit = user.get("business_unit", "No especificada")
                                        emp_no = user.get("emp_no", "N/A")
                                        area = user.get("zona", user.get("area", "General"))
                                        last_login = user.get("last_login", "Nunca")
        
                                        estado = "Desconectado"
                                        color = "🔴"
                                        if last_login and last_login != "Nunca":
                                            try:
                                                ultima_conexion = datetime.datetime.strptime(last_login, "%d-%m-%Y %H:%M")
                                                diferencia = datetime.datetime.now() - ultima_conexion
                                                if diferencia.total_seconds() < 700:
                                                    estado = "En línea"
                                                    color = "🟢"
                                            except Exception:
                                                estado = "Desconocido"
                                                color = "⚪"
        
                                        minutos_hoy = calcular_minutos_hoy(mx_id, actividades_hoy)
                                        horas_hoy = round(minutos_hoy / 60, 2)
                                        acts_count = len([a for a in actividades_hoy if str(a.get("Usuario", "")).upper().strip() == mx_id])
        
                                        col_info, col_btn = st.columns([10, 1])
                                        with col_info:
                                            st.markdown(f"""
                                            <div style="background:linear-gradient(135deg,#0d1b2a,#1a2f45);border:1px solid #00BFFF22;
                                            border-radius:10px;padding:10px 16px;margin-bottom:6px;display:flex;align-items:center;gap:20px;flex-wrap:wrap;">
                                              <div style="min-width:160px;">
                                                <div style="color:#00BFFF;font-weight:700;font-size:0.92rem;">{color} {nombre}</div>
                                                <div style="color:#888;font-size:0.75rem;">MX: {mx_id} &nbsp;|&nbsp; Emp#: {emp_no}</div>
                                              </div>
                                              <div style="color:#ccc;font-size:0.78rem;min-width:140px;">
                                                <div>📦 <b>Unidad:</b> {unidad}</div>
                                                <div>🏢 <b>BU:</b> {business_unit}</div>
                                                <div>📍 <b>Zona:</b> {area}</div>
                                              </div>
                                              <div style="color:#ccc;font-size:0.78rem;min-width:160px;">
                                                <div>🕐 <b>Último acceso:</b> {last_login}</div>
                                                <div>📶 <b>Estado:</b> {estado}</div>
                                              </div>
                                              <div style="margin-left:auto;text-align:right;font-size:0.78rem;">
                                                <div style="color:#7DF9AA;font-weight:600;">⏱ Hoy: {minutos_hoy} min / {horas_hoy} hrs</div>
                                                <div style="color:#aaa;">🔗 {acts_count} actividades hoy</div>
                                              </div>
                                            </div>
                                            """, unsafe_allow_html=True)
                                        with col_btn:
                                            if st.button("🗑️", key=f"del_tec_{mx_id}_{i}", help=f"Eliminar {nombre}"):
                                                st.session_state.db.collection("empleados").document(mx_id).delete()
                                                actividades = st.session_state.db.collection("actividades").where("Usuario", "==", mx_id).stream()
                                                for act in actividades:
                                                    st.session_state.db.collection("actividades").document(act.id).delete()
                                                st.success(f"Técnico {nombre} eliminado ✅")
                                                st.rerun()
                            else:
                                st.info("No hay técnicos registrados.")
        
                        # ══════════════════════════════════════════════════════════════
                        # VISTA SUPERVISORES
                        # ══════════════════════════════════════════════════════════════
                        elif st.session_state.tab4_vista == "supervisores":
                            st.markdown("### 📋 Cuerpo de Supervisión")
                            busqueda_sup = st.text_input("🔍 Buscar por nombre o MX", placeholder="Escribe nombre o MX...", key="busq_sup")
        
                            if supervisores:
                                supervisores_filtrados = supervisores
                                if busqueda_sup.strip():
                                    q = busqueda_sup.strip().lower()
                                    supervisores_filtrados = [
                                        e for e in supervisores
                                        if q in e.get("nombre", "").lower() or q in e.get("mx", "").lower()
                                    ]
        
                                if not supervisores_filtrados:
                                    st.info("No se encontraron supervisores con ese criterio.")
                                else:
                                    st.markdown(f"<div style='color:#aaa;font-size:0.8rem;margin-bottom:8px;'>Mostrando {len(supervisores_filtrados)} supervisor(es)</div>", unsafe_allow_html=True)
                                    for i, user in enumerate(supervisores_filtrados):
                                        nombre = user.get("nombre", "Sin nombre")
                                        mx_id = user.get("mx", f"MX_{i}").upper().strip()
                                        unidad = user.get("unidad", "No especificada")
                                        business_unit = user.get("business_unit", "No especificada")
                                        emp_no = user.get("emp_no", "N/A")
                                        area = user.get("zona", user.get("area", "General"))
                                        last_login = user.get("last_login", "Nunca")
        
                                        estado = "Desconectado"
                                        color = "🔴"
                                        if last_login and last_login != "Nunca":
                                            try:
                                                ultima_conexion = datetime.datetime.strptime(last_login, "%d-%m-%Y %H:%M")
                                                diferencia = datetime.datetime.now() - ultima_conexion
                                                if diferencia.total_seconds() < 700:
                                                    estado = "En línea"
                                                    color = "🟢"
                                            except Exception:
                                                estado = "Desconocido"
                                                color = "⚪"
        
                                        col_info, col_btn = st.columns([10, 1])
                                        with col_info:
                                            st.markdown(f"""
                                            <div style="background:linear-gradient(135deg,#0d1b2a,#1a2f45);border:1px solid #00BFFF22;
                                            border-radius:10px;padding:10px 16px;margin-bottom:6px;display:flex;align-items:center;gap:20px;flex-wrap:wrap;">
                                              <div style="min-width:160px;">
                                                <div style="color:#00BFFF;font-weight:700;font-size:0.92rem;">{color} {nombre}</div>
                                                <div style="color:#888;font-size:0.75rem;">MX: {mx_id} &nbsp;|&nbsp; Emp#: {emp_no}</div>
                                              </div>
                                              <div style="color:#ccc;font-size:0.78rem;min-width:140px;">
                                                <div>📦 <b>Unidad:</b> {unidad}</div>
                                                <div>🏢 <b>BU:</b> {business_unit}</div>
                                                <div>📍 <b>Zona:</b> {area}</div>
                                              </div>
                                              <div style="color:#ccc;font-size:0.78rem;min-width:160px;">
                                                <div>🕐 <b>Último acceso:</b> {last_login}</div>
                                                <div>📶 <b>Estado:</b> {estado}</div>
                                              </div>
                                            </div>
                                            """, unsafe_allow_html=True)
                                        with col_btn:
                                            if st.button("🗑️", key=f"del_sup_{mx_id}_{i}", help=f"Eliminar {nombre}"):
                                                st.session_state.db.collection("empleados").document(mx_id).delete()
                                                registros = st.session_state.db.collection("registros_supervisores").where("Usuario", "==", mx_id).stream()
                                                for reg in registros:
                                                    st.session_state.db.collection("registros_supervisores").document(reg.id).delete()
                                                st.success(f"Supervisor {nombre} eliminado ✅")
                                                st.rerun()
                            else:
                                st.info("No hay supervisores registrados.")
                    else:
                        st.info("⚠️ No hay empleados registrados en el sistema.")